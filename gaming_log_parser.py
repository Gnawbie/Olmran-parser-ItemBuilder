#!/usr/bin/env python3
"""
Gaming Log Parser - Built for Olmran/MUD-style text game logs
Parses Chat, Combat, and Loot from action and chat .log files
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import font as tkfont
import re, os, json, difflib
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Shown in the main window's title bar - bump this alongside the README
# Version History entry whenever a new version is cut.
VERSION = "5.1.2"

# ─────────────────────────────────────────────────────────────
#  AREA TO REALM MAPPING (from Olmran_Realm_Leveling.xlsx)
# ─────────────────────────────────────────────────────────────
AREA_TO_REALM = {
    'Abandoned Karimere Castle': 'Good', 'Abrishamkar Coast': 'Chaos', 'Abrishamkar Mountains': 'Chaos',
    'Acheronian Keep': 'Good', 'Allorien': 'Good', 'Ancient Crystalline Mines of Fala': 'Good',
    'Arachenlair Forest': 'Chaos', 'Arctic Glaciers': 'Evil', 'Arctic Glaciers Castle': 'Evil',
    'Azure Rainforest': 'Good', 'Barrows of Zden': 'Evil', 'Black Keep': 'Evil', 'Blackwoods': 'Chaos',
    'Bluffs of Zden': 'Evil', 'Carendel': 'Good', 'Carendel Nobles Graveyard': 'Good',
    'Carendel Potter\'s Field': 'Good', 'Cavern of Roots': 'Chaos', 'Cavern of Shadows': 'Good',
    'Caverns of Gaggsith': 'Chaos', 'Cerulean Lakes': 'Evil', 'Chasmanic Swamp': 'Good',
    'Chiligulla Mountains': 'Good', 'Chui Rainforest': 'Good', 'Chui Rainforest Treetops': 'Good',
    'Chui Savanna': 'Good', 'Citadel of Gnovormir': 'Chaos', 'Crescent Moon Gorge': 'Good',
    'Crystal Towers of Fala': 'Good', 'Crystalline Mines': 'Chaos', 'Curdled Blood Marsh': 'Chaos',
    'Cursed Woods': 'Evil', 'Cverick\'s Tower': 'Good', 'Darikor': 'Chaos', 'Darikor Caverns': 'Chaos',
    'Darikor Coast': 'Chaos', 'Darikor Forest': 'Chaos', 'Dark Forest': 'Evil',
    'Dark Heart of Karimere': 'Good', 'Deep Mountain Caves': 'Good', 'Diseased Temple': 'Evil',
    'Doral Coast': 'Chaos', 'Doral Sea': 'Chaos', 'Doral Straits': 'Chaos', 'Dragonclaw Island': 'Chaos',
    'Drakurat\'s Cave': 'Chaos', 'Drakurat\'s Spine': 'Chaos', 'Dunes of Kad\'iril': 'Evil',
    'Eastern Desert': 'Evil', 'Enchants Woods of Acornak': 'Chaos', 'Evermist Knolls': 'Good',
    'Fields of Mo\'Serat': 'Chaos', 'Forests of Shalifi\'s Demise': 'Good',
    'Forests of Shalifi\'s Demise - Mine': 'Good', 'Fort Darkbane': 'Chaos', 'Frozen Wastelands': 'Evil',
    'Gateway to Madness': 'Good', 'Grasslands': 'Evil', 'Grazzt\'s Refuge': 'Good',
    'Grazzt\'s Underground': 'Good', 'Great Desert': 'Evil', 'Great Swamp': 'Evil',
    'Greater Rulan Jungle': 'Chaos', 'Greater Spider Caves': 'Evil', 'Greater Wailing Woods': 'Chaos',
    'Greenmist Forest': 'Evil', 'Greenmist Forest Gateway': 'Evil', 'Greenmist Palisades': 'Evil',
    'Greenmist Palisades - Stockade': 'Evil', 'Grey Mountains': 'Evil', 'Grey Mountains Castle': 'Evil',
    'Grey Mountains Treehouse': 'Evil', 'Grey Mountains Tunnels': 'Evil', 'Greyknife Peaks': 'Evil',
    'Island of Mingo': 'Good', 'Karimere Crypt': 'Good', 'Karimere Highlands': 'Good',
    'Karimere Timber Forest': 'Good', 'Lake Dragonclaw': 'Chaos', 'Lesser Rulan Jungle': 'Chaos',
    'Lord Garon\'s Keep': 'Evil', 'Marshlands': 'Evil', 'Mausoleum': 'Good', 'Medoran Forest': 'Good',
    'Millers Pointe': 'Good', 'Mingo Jungle': 'Good', 'Mingo Mountain': 'Good',
    'Misery\'s Crossing': 'Chaos', 'Motapa Hills': 'Chaos', 'Mount Decadare': 'Chaos',
    'Mount Desparare': 'Evil', 'Mountain Caves': 'Good', 'Myrobi Hills': 'Good', 'Ninja Caves': 'Good',
    'Nogrim, the Deep City': 'Good', 'Nomadic Caravan': 'Chaos', 'Pine Forest/Tundra': 'Evil',
    'Pirate\'s Cove': 'Chaos', 'Plains of Hevak': 'Chaos', 'Poisoned Temple': 'Good',
    'Ravenwood Marsh': 'Chaos', 'Remorse Mountains': 'Good', 'Rigan Coast': 'Evil',
    'Ruins of Gnovormir': 'Chaos', 'Sanctum of the Exile': 'Chaos', 'Saurian Wastes': 'Chaos',
    'Sea of Grass': 'Chaos', 'Sea of Grass Underground': 'Chaos', 'Serpent\'s Pass': 'Evil',
    'Serpentine Mountains': 'Evil', 'Shady Brush Hills': 'Good', 'Shores of the Thundermist': 'Chaos',
    'Shrouded Castle of Craebaen': 'Good', 'Shrouded City of Craebaen': 'Good',
    'Shrouded Forest of Craebaen': 'Good', 'Spider Caves': 'Evil', 'Spider Caves Lower': 'Evil',
    'Stables of Yazik': 'Evil', 'Stonefist Temple': 'Chaos', 'Straits of Mingo': 'Good',
    'Straits of Mingo Islands': 'Good', 'Tamia': 'Evil', 'Tamia Local': 'Evil',
    'Tamian Foothills': 'Evil', 'Teeth of Heaven': 'Good', 'Temple of the Divine Horn': 'Evil',
    'Tenebrous Hollow': 'Chaos', 'Terandhel': 'Good', 'The Bladegrass': 'Good', 'The Catacombs': 'Evil',
    'The Deadlands': 'Chaos', 'The Endless Night': 'Evil', 'The Endless Night City of Zhak-Tor': 'Evil',
    'The Endless Night Sorceror\'s Tower': 'Evil', 'The Endless Night Temple of Klotha': 'Evil',
    'The Hive': 'Evil', 'The Orc Caverns': 'Evil', 'The Pyramid of the Sun': 'Evil',
    'The Rowangroves': 'Evil', 'The Southern Dunes': 'Chaos', 'The Underdark': 'Evil',
    'The Underground Temple': 'Evil', 'The Wailing Woods': 'Chaos', 'The Wildwood': 'Good',
    'The snarewoods': 'Evil', 'Thorn Forest': 'Good', 'Thorny Brush Hills': 'Chaos',
    'Throrfiril': 'Chaos', 'Throrfiril City of Trees': 'Chaos', 'Throrfiril Royal Citadel': 'Chaos',
    'Tomb of the Dragon King': 'Chaos', 'Twisted Jungle': 'Good', 'Unicorn Glen': 'Evil',
    'Unicorn Glen Basement': 'Evil', 'Upper Drakurat\'s Spine': 'Chaos', 'Utilla Town': 'Chaos',
    'Valley of Zden': 'Evil', 'Valley of the Giants': 'Evil', 'Valley of the Giants Volcano': 'Evil',
    'Venomvein Timberland': 'Good', 'Verulian Falls': 'Chaos', 'Verulian Forest': 'Chaos',
    'Village of Reilyn': 'Evil', 'Wastes of Olmran': 'Evil', 'Weald of Decdare': 'Chaos',
    'Wretched Forest': 'Chaos',
}

# ─────────────────────────────────────────────────────────────
#  PARSERS  (one class per log type, using actual file patterns)
# ─────────────────────────────────────────────────────────────

class ChatParser:
    """
    Parses CHAT log files.
    Line format examples:
      <00:23:29> You say '11'
      <00:32:27> [Plan] Terminator: 'yo'
      <00:34:26> Zealut sends, 'ball spells...'
      <00:34:34> You send, 'yes' to Zealut
    """
    CHANNEL_RE = re.compile(
        r'^<(\d{2}:\d{2}:\d{2})>\s+'           # timestamp
        r'(?:\[([^\]]+)\]\s+)?'                  # optional [channel]
        r'(.+?)(?:\s+sends?,\s*\'([^\']+)\'|'    # Name sends, '...'
        r'\s*:\s*\'([^\']+)\'|'                   # Name: '...'
        r'\s+says?\s+\'([^\']+)\')'              # Name says '...'
    )
    SEND_RE = re.compile(
        r'^<(\d{2}:\d{2}:\d{2})>\s+You send,\s*\'([^\']+)\'\s+to\s+(\w+)'
    )
    SAY_RE = re.compile(
        r'^<(\d{2}:\d{2}:\d{2})>\s+You say\s+\'([^\']+)\''
    )
    NAMED_RE = re.compile(
        r'^<(\d{2}:\d{2}:\d{2})>\s+(?:\[([^\]]+)\]\s+)?(\w+)\s*(?:sends?,\s*)?[\':]?\s*\'([^\']+)\''
    )

    @staticmethod
    def parse_file(filepath: str) -> list:
        rows = []
        filename = os.path.basename(filepath)
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            for raw in f:
                line = raw.strip()
                if not line:
                    continue
                ts_m = re.match(r'^<(\d{2}:\d{2}:\d{2})>', line)
                if not ts_m:
                    continue
                ts = ts_m.group(1)
                rest = line[len(ts_m.group(0)):].strip()

                # You send '...' to Name
                m = re.match(r"You send,?\s+'([^']+)'\s+to\s+(\w+)", rest, re.I)
                if m:
                    rows.append({'Timestamp': ts, 'Channel': 'Tell-Out',
                                 'Speaker': 'You', 'Target': m.group(2),
                                 'Message': m.group(1), 'File': filename})
                    continue

                # You say '...'
                m = re.match(r"You say\s+'([^']+)'", rest, re.I)
                if m:
                    rows.append({'Timestamp': ts, 'Channel': 'Say',
                                 'Speaker': 'You', 'Target': '',
                                 'Message': m.group(1), 'File': filename})
                    continue

                # [Channel] Name: '...' or Name sends, '...' or Name says '...'
                m = re.match(r"(?:\[([^\]]+)\]\s+)?(\w+)\s+(?:sends?,\s*)?[\"']([^\"']+)[\"']", rest, re.I)
                if m:
                    ch = m.group(1) or 'Say'
                    # Target from "to Name" suffix
                    tgt_m = re.search(r"\bto\s+(\w+)\s*$", rest, re.I)
                    rows.append({'Timestamp': ts, 'Channel': ch,
                                 'Speaker': m.group(2), 'Target': tgt_m.group(1) if tgt_m else '',
                                 'Message': m.group(3), 'File': filename})
                    continue

                # [Plan] Name: 'msg'   or   [Guild] Name: 'msg'
                m = re.match(r"\[([^\]]+)\]\s+(\w+):\s+'([^']+)'", rest, re.I)
                if m:
                    rows.append({'Timestamp': ts, 'Channel': m.group(1),
                                 'Speaker': m.group(2), 'Target': '',
                                 'Message': m.group(3), 'File': filename})
                    continue

        return rows


class CombatParser:
    """
    Parses ACTION log files for combat events.
    Location comes from [Room Name] lines immediately before the block.

    Extracts:
      - Player attacks mob (smash/cast lines)
      - Mob attacks player
      - Kill events + XP gained
      - Spell casts
    """
    TS_RE   = re.compile(r'^<(\d{2}:\d{2}:\d{2})>')
    LOC_RE  = re.compile(r'^\[([^\]]+)\]')

    MOB_ATK = re.compile(
        r'^(A |An |The )?([A-Za-z ]+?) attacks? (?:you|a knight of dread)(?: and (miss(?:es)?)| for (\d+) damage!)'
    )
    YOU_ATK = re.compile(
        r'^You (?:smash|strike|swing at) (?:a |an |the )?(.+?) (?:with .+? for (\d+) damage!|and miss)'
    )
    SPELL   = re.compile(
        r'^You cast a (.+?) spell!'
    )
    FREEZE  = re.compile(
        r'^(.+?) is hit for (\d+) damage!'
    )
    KILL    = re.compile(
        r'^(.+?) dies!'
    )
    XP      = re.compile(
        r'^You gain (\d+) \(\+(\d+)\) experience points\.'
    )
    HIT_REC = re.compile(
        r'^You are hit for (\d+) damage'
    )

    @staticmethod
    def parse_file(filepath: str) -> list:
        rows = []
        filename = os.path.basename(filepath)
        current_location = ''
        current_ts = ''

        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            raw_text = f.read()

        # Split into timestamped blocks
        blocks = re.split(r'\n(?=<\d{2}:\d{2}:\d{2}>)', raw_text)

        for block in blocks:
            lines = [l.rstrip('\r') for l in block.split('\n')]
            if not lines:
                continue

            ts_m = CombatParser.TS_RE.match(lines[0])
            if ts_m:
                current_ts = ts_m.group(1)

            # Check for location header in this block
            for l in lines:
                loc_m = CombatParser.LOC_RE.match(l.strip())
                if loc_m:
                    current_location = loc_m.group(1)
                    break

            # Parse each content line
            for l in lines[1:]:
                l = l.strip()
                if not l:
                    continue

                # Mob attacks you
                m = CombatParser.MOB_ATK.match(l)
                if m:
                    dmg = m.group(4) or '0'
                    hit = 'miss' if m.group(3) else 'hit'
                    rows.append({
                        'Timestamp': current_ts,
                        'Location': current_location,
                        'Attacker': (m.group(1) or '') + (m.group(2) or ''),
                        'Target': 'You',
                        'Action': hit,
                        'Damage': int(dmg) if dmg.isdigit() else 0,
                        'Spell': '',
                        'Event': 'mob_attack',
                        'Full': l,
                        'File': filename
                    })
                    continue

                # You attack a mob
                m = CombatParser.YOU_ATK.match(l)
                if m:
                    rows.append({
                        'Timestamp': current_ts,
                        'Location': current_location,
                        'Attacker': 'You',
                        'Target': m.group(1).strip(),
                        'Action': 'smash' if m.group(2) else 'miss',
                        'Damage': int(m.group(2)) if m.group(2) else 0,
                        'Spell': '',
                        'Event': 'player_attack',
                        'Full': l,
                        'File': filename
                    })
                    continue

                # Spell cast
                m = CombatParser.SPELL.match(l)
                if m:
                    rows.append({
                        'Timestamp': current_ts,
                        'Location': current_location,
                        'Attacker': 'You',
                        'Target': '',
                        'Action': 'cast',
                        'Damage': 0,
                        'Spell': m.group(1),
                        'Event': 'spell_cast',
                        'Full': l,
                        'File': filename
                    })
                    continue

                # Mob/target hit for damage (spell AoE)
                m = CombatParser.FREEZE.match(l)
                if m:
                    rows.append({
                        'Timestamp': current_ts,
                        'Location': current_location,
                        'Attacker': 'You (spell)',
                        'Target': m.group(1).strip(),
                        'Action': 'spell_hit',
                        'Damage': int(m.group(2)),
                        'Spell': '',
                        'Event': 'spell_damage',
                        'Full': l,
                        'File': filename
                    })
                    continue

                # Kill
                m = CombatParser.KILL.match(l)
                if m:
                    rows.append({
                        'Timestamp': current_ts,
                        'Location': current_location,
                        'Attacker': 'You',
                        'Target': m.group(1).strip(),
                        'Action': 'kill',
                        'Damage': 0,
                        'Spell': '',
                        'Event': 'kill',
                        'Full': l,
                        'File': filename
                    })
                    continue

                # XP gained
                m = CombatParser.XP.match(l)
                if m:
                    rows.append({
                        'Timestamp': current_ts,
                        'Location': current_location,
                        'Attacker': '',
                        'Target': '',
                        'Action': 'xp_gain',
                        'Damage': 0,
                        'Spell': '',
                        'Event': f"XP: {m.group(1)} (+{m.group(2)} bonus)",
                        'Full': l,
                        'File': filename
                    })
                    continue

                # Player takes damage (bleed, etc.)
                m = CombatParser.HIT_REC.match(l)
                if m:
                    rows.append({
                        'Timestamp': current_ts,
                        'Location': current_location,
                        'Attacker': 'Environment',
                        'Target': 'You',
                        'Action': 'damage_taken',
                        'Damage': int(m.group(1)),
                        'Spell': '',
                        'Event': 'env_damage',
                        'Full': l,
                        'File': filename
                    })
                    continue

        return rows


class LootParser:
    """
    Parses ACTION log files for loot/drop events.
    Location comes from [Room Name] bracket in same block.

    Patterns from actual log:
      A sand giant drops a shining cracked gauntlets of the sands.
      An Al'Tizor assassin drops a glowing mithril enforcer's bladed spear.
      [GSEND] 'Cardova has completed the 'Level 10 Sin Gear' player achievement!'
    """
    TS_RE   = re.compile(r'^<(\d{2}:\d{2}:\d{2})>')
    LOC_RE  = re.compile(r'^\[([^\]]+)\]')

    DROP_RE  = re.compile(r'^(A |An |The )?(.+?) drops (.+?)\.')
    ACHIEV_RE = re.compile(r"\[GSEND\] '(.+?) has completed the '(.+?)' player achievement!'")
    
    # Item quality and material prefixes to remove
    QUALITY_PREFIXES = re.compile(r'^(silvered|bright|shining|glowing|lustrous|brilliant)\s+', re.IGNORECASE)
    MATERIAL_PREFIXES = re.compile(r'^(bronze|iron|steel|alloy|mithril|laen|wool|cotton|silk|gossamer|wispweave|ebonweave|leather|rough|embossed|suede|wyvern scale|enchanted|maple|oak|yew|rosewood|ironwood|ebony)\s+', re.IGNORECASE)
    
    @staticmethod
    def clean_item_name(item_name: str) -> str:
        """Remove quality and material prefixes from item names."""
        original = item_name
        
        # Remove leading article (a, an, the)
        cleaned = re.sub(r'^(a|an|the)\s+', '', item_name, flags=re.IGNORECASE)
        
        # Try removing quality prefixes
        after_quality = LootParser.QUALITY_PREFIXES.sub('', cleaned)
        
        # Try removing material prefixes
        after_material = LootParser.MATERIAL_PREFIXES.sub('', after_quality)
        
        result = after_material.strip()
        
        # If we're down to just one word, that's probably wrong
        # Keep the material prefix instead
        if result and len(result.split()) == 1:
            # Just remove quality, keep material
            result = after_quality.strip()
        
        return result if result else cleaned.strip()
    
    @staticmethod
    def parse_delve_stats(lines: list) -> dict:
        """Extract stats from delve output in the log."""
        stats = {
            'Slot': '',
            'Type': '',
            'Spell': '',
            'Level': '',
            'Damage': '',
            'Timer': '',
            'Fumble': '',
            'Accuracy': '',
            'Defense': '',
            'Sigil': '',
            'SigilLvl': '',
            'Weight': '',
            'Holdable': False,
            'HoldableWhileCasting': False,
            'Hands': '',
            'Worth': ''
        }
        
        for line in lines:
            line = line.strip()
            
            # Weapon type: "This object is a weapon of the two-handed thrusting type."
            m = re.search(r'weapon of the (.+?)\s+type', line, re.I)
            if m:
                weapon_desc = m.group(1).strip()
                # Extract just the last word (the actual type: thrusting, slashing, etc.)
                type_match = re.search(r'(\w+)$', weapon_desc)
                if type_match:
                    stats['Type'] = type_match.group(1)
                stats['Slot'] = 'weapon'
                continue
            
            # Hands: "This object takes two hands to wield."
            if re.search(r'two hands to wield', line, re.I):
                stats['Hands'] = '2h'
                continue
            
            # Armor slot: "This object can be worn on the Hands."
            m = re.search(r'worn on the (\w+)', line, re.I)
            if m:
                stats['Slot'] = m.group(1).lower()
                continue
            
            # Armor type: "This object is armor of the cloth type."
            m = re.search(r'armor of the (\w+) type', line, re.I)
            if m:
                stats['Type'] = m.group(1)
                continue
            
            # Spell: "This object seems to cast the spell: Agility.II."
            # Preserve case to keep .I .II .III intact
            m = re.search(r'cast the spell:\s*(.+?)\.?\s*$', line, re.I)
            if m:
                spell_name = m.group(1).strip()
                # Don't add generic words as spells
                if spell_name.lower() not in ['melee', 'shield', 'armor', 'weapon']:
                    stats['Spell'] = spell_name  # Keep original case
                continue
            
            # Sigil: "This item seems to be enchanted with a tier 2 sigil of Psionic Protection!"
            m = re.search(r'tier (\d+) sigil of ([^!]+)', line, re.I)
            if m:
                stats['SigilLvl'] = m.group(1)
                stats['Sigil'] = m.group(2).strip()
                continue
            
            # Level: "This item seems to require level 60 to use effectively."
            m = re.search(r'require level (\d+)', line, re.I)
            if m:
                stats['Level'] = m.group(1)
                continue
            
            # Damage: "This item seems to do a very good amount of damage."
            m = re.search(r'do a (.+?) amount of damage', line, re.I)
            if m:
                stats['Damage'] = m.group(1).strip()
                continue
            
            # Fumble: "This item seems to fumble rarely."
            m = re.search(r'fumble (\w+)', line, re.I)
            if m:
                stats['Fumble'] = m.group(1).strip()
                continue
            
            # Accuracy: "This item seems to hit in combat better than normal."
            m = re.search(r'hit in combat (.+?)\.', line, re.I)
            if m:
                accuracy_desc = m.group(1).strip()
                # Clean up "better than normal" -> "better", "about like normal" -> "normal"
                accuracy_desc = accuracy_desc.replace(' than normal', '').replace('about like ', '')
                stats['Accuracy'] = accuracy_desc.strip()
                continue
            
            # Defense/Protection: "This item seems to protect in combat about like normal."
            m = re.search(r'protect in combat (.+?)\.', line, re.I)
            if m:
                defense_desc = m.group(1).strip()
                # Clean up "about like normal" -> "normal"
                defense_desc = defense_desc.replace('about like ', '').replace(' than normal', '')
                stats['Defense'] = defense_desc.strip()
                continue
            
            # Weight: "This item seems to weigh 6 pounds." or "This item seems to weigh 2 pounds."
            m = re.search(r'weigh (\d+)', line, re.I)
            if m:
                stats['Weight'] = m.group(1)
                continue
            
            # Holdable while casting: "This item seems to be holdable while casting spells."
            if re.search(r'holdable while casting', line, re.I):
                stats['HoldableWhileCasting'] = True
                stats['Holdable'] = True
                continue
            
            # Holdable: "This object can be held"
            if re.search(r'can be held|holdable', line, re.I):
                stats['Holdable'] = True
                continue
            
            # Worth/Value: "This item seems to be worth between 23000 and 24000 silver."
            m = re.search(r'worth between (\d+) and (\d+) silver', line, re.I)
            if m:
                # Use the average of the range
                low = int(m.group(1))
                high = int(m.group(2))
                avg = (low + high) // 2
                stats['Worth'] = str(avg)
                continue
        
        # Post-process: Build Type field for weapons
        if stats['Slot'] == 'weapon' and stats['Type']:
            # Build Type: [direct] <type> <hands>
            type_parts = []
            
            # Add "direct" prefix if holdable while casting
            if stats['HoldableWhileCasting']:
                type_parts.append('direct')
            
            # Add weapon type (slashing, crushing, thrusting, etc.)
            type_parts.append(stats['Type'])
            
            # Add hands (1h or 2h)
            if stats['Hands']:
                type_parts.append(stats['Hands'])
            else:
                # Default to 1h if not specified
                type_parts.append('1h')
            
            stats['Type'] = ' '.join(type_parts)
        
        return stats

    @staticmethod
    def parse_file(filepath: str) -> list:
        rows = []
        filename = os.path.basename(filepath)

        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            raw_text = f.read()

        # Split into timestamp blocks
        blocks = re.split(r'\n(?=<\d{2}:\d{2}:\d{2}>)', raw_text)
        
        # First pass: Find all delves and extract stats
        delved_items = {}  # clean_item_name -> delve_stats
        
        for block in blocks:
            lines = [l.rstrip('\r') for l in block.split('\n')]
            
            # Look for "You examine a <item> closely." which indicates delve output
            delve_item_name = None
            for line in lines:
                m = re.search(r'You examine (?:a |an |the )?(.+?) closely', line, re.I)
                if m:
                    full_item_name = m.group(1).strip()
                    clean_item_name = LootParser.clean_item_name(full_item_name)
                    delve_item_name = clean_item_name
                    break
            
            if delve_item_name:
                # Extract stats from this delve block
                stats = LootParser.parse_delve_stats(lines)
                delved_items[delve_item_name] = stats
        
        # Second pass: Find drops of delved items
        # Build a list of all area markers and drop events with their positions
        area_markers = []  # (line_index, area_name)
        drop_events = []   # (line_index, drop_line, item_clean, mob_name, timestamp)
        
        line_index = 0
        current_ts = ''
        
        for block in blocks:
            lines = [l.rstrip('\r') for l in block.split('\n')]
            
            # Get timestamp from first line
            ts_m = LootParser.TS_RE.match(lines[0])
            if ts_m:
                current_ts = ts_m.group(1)
            
            for line in lines:
                # Track area markers
                loc_m = LootParser.LOC_RE.match(line.strip())
                if loc_m:
                    area_markers.append((line_index, loc_m.group(1)))
                
                # Track drops
                m = LootParser.DROP_RE.match(line.strip())
                if m:
                    item_raw = m.group(3).strip()
                    
                    # Skip silver/coins
                    if re.match(r'a bag of (silver|coins?|gold)', item_raw, re.I):
                        line_index += 1
                        continue
                    
                    item_clean = LootParser.clean_item_name(item_raw)
                    
                    # Only care about delved items
                    if item_clean in delved_items:
                        mob_name = ((m.group(1) or '') + m.group(2)).strip()
                        mob_name = re.sub(r'^(A|An|The)\s+', '', mob_name, flags=re.IGNORECASE).strip()
                        drop_events.append((line_index, line.strip(), item_clean, mob_name, current_ts))
                
                line_index += 1
        
        # Third pass: Match drops to their nearest preceding area
        for drop_line_idx, drop_line, item_name, mob_name, timestamp in drop_events:
            # Find the closest area marker before this drop
            closest_area = ''
            for area_idx, area_name in area_markers:
                if area_idx < drop_line_idx:
                    closest_area = area_name
                else:
                    break  # Past the drop point
            
            # Get delve stats for this item
            delve_stats = delved_items.get(item_name, {})
            
            # Check if this is a fodder/crafting item (has only weight, no other stats)
            has_only_weight = (
                delve_stats.get('Weight') and
                not delve_stats.get('Slot') and
                not delve_stats.get('Type') and
                not delve_stats.get('Spell') and
                not delve_stats.get('Level') and
                not delve_stats.get('Damage') and
                not delve_stats.get('Defense')
            )
            
            if has_only_weight:
                if delve_stats.get('Holdable'):
                    # Holdable crafting item
                    delve_stats['Slot'] = 'possible'
                    delve_stats['Type'] = 'crafting'
                    delve_stats['Spell'] = 'item'
                else:
                    # Non-holdable fodder
                    delve_stats['Slot'] = 'fodder'
                    delve_stats['Type'] = 'fodder'
            
            realm = AREA_TO_REALM.get(closest_area, '') if closest_area else ''
            
            # Build notes with sell price if available
            notes = ''
            if delve_stats.get('Worth'):
                notes = f"sells for {delve_stats['Worth']}"
            
            rows.append({
                'Realm': realm,
                'Area': closest_area,
                'Mob': mob_name,
                'Item': item_name,
                'Slot': delve_stats.get('Slot', ''),
                'Type': delve_stats.get('Type', ''),
                'Spell': delve_stats.get('Spell', ''),
                'Level': delve_stats.get('Level', ''),
                'Damage': delve_stats.get('Damage', ''),
                'Timer': '',
                'Fumble': delve_stats.get('Fumble', ''),
                'Accuracy': delve_stats.get('Accuracy', ''),
                'Defense': delve_stats.get('Defense', ''),
                'Sigil': delve_stats.get('Sigil', ''),
                'SigilLvl': delve_stats.get('SigilLvl', ''),
                'Weight': delve_stats.get('Weight', ''),
                'Notes': notes,
                'Timestamp': timestamp,
                'Location': closest_area,
                'Source': mob_name,
                'Quantity': '1',
                'Silver': '',
                'Event': 'mob_drop',
                'Full': drop_line,
                'File': filename
            })

        # Deduplicate rows based on Realm, Area, Mob, Item
        # Keep first occurrence of each unique combination
        seen = set()
        deduplicated = []
        for row in rows:
            key = (row['Realm'], row['Area'], row['Mob'], row['Item'])
            if key not in seen:
                seen.add(key)
                deduplicated.append(row)

        return deduplicated


# ─────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────────────

COLORS = {
    'chat':   'FF4472C4',  # blue
    'combat': 'FFED7D31',  # orange
    'loot':   'FF70AD47',  # green
}

def write_sheet(ws, rows: list, fields: list, sheet_type: str):
    """Write rows to a worksheet using the field configuration."""
    if not rows:
        ws.append(['No data found for this log type.'])
        return

    accent = COLORS.get(sheet_type, 'FF4472C4')
    hdr_font  = Font(name='Arial', bold=True, color='FFFFFFFF', size=10)
    hdr_fill  = PatternFill(fill_type='solid', start_color=accent, end_color=accent)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Arial', size=10)
    thin      = Side(style='thin', color='FFD9D9D9')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Build header row using field positions
    # fields = list of {'label': str, 'source_key': str, 'col': int, 'row_mode': bool}
    max_col = max(f['col'] for f in fields) if fields else 1
    headers = [''] * max_col
    for f in fields:
        headers[f['col'] - 1] = f['label']

    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    ws.row_dimensions[1].height = 22

    # Data rows
    alt_fill = PatternFill(fill_type='solid', start_color='FFF2F2F2', end_color='FFF2F2F2')
    for ri, row_data in enumerate(rows, 2):
        values = [''] * max_col
        for f in fields:
            values[f['col'] - 1] = str(row_data.get(f['source_key'], ''))
        ws.append(values)
        for ci in range(1, max_col + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            if ri % 2 == 0:
                cell.fill = alt_fill

    ws.freeze_panes = 'A2'

    # Auto-width
    for ci in range(1, max_col + 1):
        col_letter = get_column_letter(ci)
        best = len(str(headers[ci - 1])) + 2
        for ri in range(2, min(len(rows) + 2, 200)):
            val = ws.cell(row=ri, column=ci).value or ''
            best = max(best, min(len(str(val)) + 2, 60))
        ws.column_dimensions[col_letter].width = best


# ─────────────────────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────────────────────

DEFAULT_FIELDS = {
    'chat': [
        {'label': 'Timestamp',  'source_key': 'Timestamp', 'col': 1},
        {'label': 'Channel',    'source_key': 'Channel',   'col': 2},
        {'label': 'Speaker',    'source_key': 'Speaker',   'col': 3},
        {'label': 'Target',     'source_key': 'Target',    'col': 4},
        {'label': 'Message',    'source_key': 'Message',   'col': 5},
        {'label': 'File',       'source_key': 'File',      'col': 6},
    ],
    'combat': [
        {'label': 'Timestamp',  'source_key': 'Timestamp', 'col': 1},
        {'label': 'Location',   'source_key': 'Location',  'col': 2},
        {'label': 'Attacker',   'source_key': 'Attacker',  'col': 3},
        {'label': 'Target',     'source_key': 'Target',    'col': 4},
        {'label': 'Action',     'source_key': 'Action',    'col': 5},
        {'label': 'Damage',     'source_key': 'Damage',    'col': 6},
        {'label': 'Spell',      'source_key': 'Spell',     'col': 7},
        {'label': 'Event',      'source_key': 'Event',     'col': 8},
        {'label': 'File',       'source_key': 'File',      'col': 9},
    ],
    'loot': [
        {'label': 'Realm',      'source_key': 'Realm',      'col': 1},
        {'label': 'Area',       'source_key': 'Area',       'col': 2},
        {'label': 'Mob',        'source_key': 'Mob',        'col': 3},
        {'label': 'Item',       'source_key': 'Item',       'col': 4},
        {'label': 'Slot',       'source_key': 'Slot',       'col': 5},
        {'label': 'Type',       'source_key': 'Type',       'col': 6},
        {'label': 'Spell',      'source_key': 'Spell',      'col': 7},
        {'label': 'Level',      'source_key': 'Level',      'col': 8},
        {'label': 'Damage',     'source_key': 'Damage',     'col': 9},
        {'label': 'Timer',      'source_key': 'Timer',      'col': 10},
        {'label': 'Fumble',     'source_key': 'Fumble',     'col': 11},
        {'label': 'Accuracy',   'source_key': 'Accuracy',   'col': 12},
        {'label': 'Defense',    'source_key': 'Defense',    'col': 13},
        {'label': 'Sigil',      'source_key': 'Sigil',      'col': 14},
        {'label': 'SigilLvl',   'source_key': 'SigilLvl',   'col': 15},
        {'label': 'Weight',     'source_key': 'Weight',     'col': 16},
        {'label': 'Notes',      'source_key': 'Notes',      'col': 17},
    ],
}

CHAT_SOURCES   = ['Timestamp','Channel','Speaker','Target','Message','File']
COMBAT_SOURCES = ['Timestamp','Location','Attacker','Target','Action','Damage','Spell','Event','Full','File']
LOOT_SOURCES   = ['Realm','Area','Mob','Item','Slot','Type','Spell','Level','Damage','Timer','Fumble','Accuracy','Defense','Sigil','SigilLvl','Weight','Notes','Timestamp','Location','Source','Quantity','Silver','Event','Full','File']

# Spell name options for the Build tab category dropdowns.
SPELL_CATEGORIES = {
    'Basic':          ['Agility', 'Bless', 'Dexterity', 'Constitution', 'Intelligence', 'Wisdom', 'Strength', 'Evade', 'Combat', 'Tough.skin'],
    'Class Specific': ['Reverb.enhance', 'Aura.enhance', 'Backstab.enhance', 'Bash.enhance', 'Berzerk.enhance',
                       'Crush.enhance', 'Direct.enhance', 'Double.enhance', 'Fired.enhance', 'Improve.enhance',
                       'Leathers.enhance', 'Lockpick.enhance', 'Mark.enhance', 'Martialarts.enhance', 'Maul.enhance',
                       'Melee.enhance', 'Pummel.enhance', 'Rake.enhance', 'Repair.enhance',
                       'Slash.enhance', 'Tap.enhance', 'Thrust.enhance', 'Track.enhance',
                       'Volley.enhance'],
    'Shields/Buffs':  ['Protect', 'Blur', 'Shield', 'Vitalize', 'Regenerate',
                      'Bleed.resist', 'Disease.resist', 'Poison.resist'],
    'General Skills': ['Platemail.enhance', 'Leathers.enhance', 'Thrust.enhance', 'Chaos.crush',
                       'Slash.enhance', 'All.weapons.enhance', 'Weapons.enhance',
                       'Climb.enhance', 'Hide.enhance', 'Jump.enhance', 'Swim.enhance',
                       'Percept.enhance', 'Sneak.enhance'],
    # Protects use minor/normal/improved instead of i/ii/iii - minor and
    # improved are a PREFIX on the spell (e.g. "minor.cold.protect"), not a
    # suffix; "normal" is the bare spell name with no prefix at all.
    'Protects':       ['Cold.protect', 'Earth.protect', 'Elemental.protect', 'Fire.protect',
                       'Lightning.protect', 'Mental.protect', 'Shock.protect', 'Water.protect'],
}

SPELL_TIERS = ['(any)', 'i', 'ii', 'iii']

# Spells (lowercase) whose tier dropdown is restricted (or extended) beyond
# the default SPELL_TIERS whenever that spell is selected.
PROTECT_SPELLS = ['cold.protect', 'earth.protect', 'elemental.protect', 'fire.protect',
                  'lightning.protect', 'mental.protect', 'shock.protect', 'water.protect']
# Protects show minor/normal/improved as friendlier labels, but these map
# directly onto the ordinary .i/.ii/.iii suffix everything else uses - no
# special parsing, so there's no confusion around the word "protect" itself.
PROTECT_TIERS = ['(any)', 'minor', 'normal', 'improved']
PROTECT_TIER_TO_SUFFIX = {'minor': 'i', 'normal': 'ii', 'improved': 'iii'}
# TODO: some improved.spell (tier iii) Protects aren't showing up in build
# search results - root cause not yet diagnosed, fix deferred (2026-07-05).

SPELL_TIER_RESTRICTIONS = {
    'agility': ['(any)', 'i', 'ii'],
    'bless': ['(any)', 'i', 'ii'],
    # disease.resist is the only spell that goes up to tier iv
    'disease.resist': ['(any)', 'i', 'ii', 'iii', 'iv'],
    **{spell: PROTECT_TIERS for spell in PROTECT_SPELLS},
    # Class Specific spells don't go up to tier iii at all.
    **{spell.lower(): ['(any)', 'i', 'ii'] for spell in SPELL_CATEGORIES['Class Specific']},
}

# A build can include at most this many items from the "Crafted" realm.
MAX_CRAFTED_ITEMS = 1

# Cap on how many alternate full-build variants "Generate multiple build
# options" will produce (in addition to the primary optimal build).
MAX_BUILD_VARIANTS = 10

# Spells (lowercase) whose underlying spell-column value differs from the
# display name itself, e.g. Evade is stored as "evade.enhance".
SPELL_VALUE_OVERRIDES = {
    'evade': 'evade.enhance',
}

_TIER_RANK = {'i': 1, 'ii': 2, 'iii': 3, 'iv': 4}

DEFENSE_LEVELS = ['much worse', 'worse', 'normal', 'better', 'much better']
DEFENSE_RANK = {level: i for i, level in enumerate(DEFENSE_LEVELS)}

SIGIL_TYPES = ['Cold', 'Earth', 'Fire', 'Lightning', 'Pain', 'Shock', 'Water']

# The 6 armor body slots that have their own per-slot Sigil dropdown in Armor
# Constraints (weapon/shield/jewel have their own separate Sigil concepts -
# see MELEE_SIGIL_VALUES/SHIELD_SIGIL_VALUES below) - also what Wanted
# Sigils is scoped to, since it's built as part of Armor Constraints.
ARMOR_SIGIL_SLOTS = ['head', 'cloak', 'body', 'hands', 'legs', 'feet']

# Weapon combo damage-type dropdowns (Dual-Wield 1h, 1h/Shield, 2h/Shield).
# The item list's Type column spells these as "slash"/"thrust"/"crush" (e.g.
# "offhand slash 1h"), not "slashing"/"thrusting"/"crushing", so matching
# needs the stem, not the display label, to actually hit anything.
WEAPON_DAMAGE_TYPES = ['Any', 'Slashing', 'Thrusting', 'Crushing']
WEAPON_DAMAGE_STEMS = {'Slashing': 'slash', 'Thrusting': 'thrust', 'Crushing': 'crush'}

# Melee Weapon Constraints' Damage dropdown - the item list's Damage column
# values, best to worst (matches the item list's actual lowercase spelling).
MELEE_DAMAGE_LEVELS = ['Any', 'Amazing', 'Very Good', 'Good', 'Small']

# Melee Weapon Constraints' Timer dropdown - every distinct value seen in the
# item list's Timer column (weapon speed in ms). Not all weapons have one -
# 'Any' covers "no restriction", not "blank" specifically.
MELEE_TIMER_VALUES = ['Any', '2200', '2300', '2400', '2500', '2600', '2700', '2800', '2900',
                      '3000', '3100', '3199', '3200', '3300', '3400', '3500', '3599', '3600',
                      '3700', '3800', '3900', '4000', '4200', '4400', '4500', 'normal']

# Melee Weapon Constraints' Fumble dropdown - the item list's Fumble column values.
MELEE_FUMBLE_VALUES = ['Any', 'Never', 'Rarely', 'Sometimes', 'Normal', 'Often']

# Melee Weapon Constraints' Accuracy dropdown - the item list's Accuracy
# column values, best to worst.
MELEE_ACCURACY_VALUES = ['Any', 'Much Better', 'Better', 'Normal', 'Worse', 'Much Worse']

# Melee Weapon Constraints' Sigil dropdown - same Sigil types as Armor
# Constraints' per-slot Sigil dropdowns, plus an explicit "None" option
# (require no Sigil at all) that Armor Constraints doesn't have.
MELEE_SIGIL_VALUES = ['Any', 'None'] + SIGIL_TYPES

# Shield Constraints' Defense dropdown - a single-value pick (unlike Armor
# Constraints' Min/Max range), fed into the same _defense_priority_score
# machinery as an exact one-value range.
SHIELD_DEFENSE_LEVELS = ['Any'] + DEFENSE_LEVELS

# Shield Constraints' Sigil dropdown - same list/behavior as every other
# Sigil dropdown in Armor Constraints (no "None" option, unlike Melee's).
SHIELD_SIGIL_VALUES = ['Any'] + SIGIL_TYPES


def _weapon_damage_matches(item_type, choice):
    """True if choice is 'Any' or its stem appears in item_type (already
    lowercased by the caller)."""
    if choice == 'Any':
        return True
    return WEAPON_DAMAGE_STEMS.get(choice, '\0') in item_type


def _weapon_style_matches(item_type, style):
    """True if the item's melee-vs-direct classification matches style
    ('Melee' or 'Direct'). Staff-type items never match either - callers
    that support a Parry option (currently only Two-Handed) handle staves
    separately, before calling this."""
    if 'staff' in item_type:
        return False
    is_direct = 'direct' in item_type
    return is_direct if style == 'Direct' else not is_direct


_KNOWN_WANTED_SPELL_BASES = {spell.lower() for spells in SPELL_CATEGORIES.values() for spell in spells}


def _direct_weapon_eligible(item_spell):
    """Direct weapons always carry a spell in-game, but it's typically a
    one-off effect that "does its own thing" rather than being one of the
    program's recognized/pickable spells (see SPELL_CATEGORIES) - so
    eligibility requires some spell to be present, but specifically NOT one
    that could be chosen from the Wanted Spells category dropdowns. Default
    policy until those one-off spells get added to the program; revisit
    then. Melee/Parry/Fired have no such requirement."""
    if not item_spell:
        return False
    return _spell_base(item_spell) not in _KNOWN_WANTED_SPELL_BASES


def _two_handed_matches(item_type, item_spell, style, damage):
    """Two-Handed eligibility, gated by its own Style (Melee/Direct/Parry/
    Fired) and Damage Type dropdowns - a first step toward per-combo style
    selection (see the comment above the UI). Parry (staff-type items) and
    Fired ("fired 2h") both have no damage-type concept in the item list, so
    each is matched on its own, ignoring the Damage Type dropdown - Parry
    also ignores the 2h requirement, since parry staves aren't tagged 2h."""
    if style == 'Parry':
        return 'staff' in item_type
    if '2h' not in item_type:
        return False
    if style == 'Fired':
        return 'fired' in item_type
    if style == 'Direct' and not _direct_weapon_eligible(item_spell):
        return False
    return _weapon_style_matches(item_type, style) and _weapon_damage_matches(item_type, damage)


def _spell_base(spell):
    """Strip a trailing .i/.ii/.iii/.iv tier suffix so different tiers of the
    same wanted spell are recognized as one requirement, not two.

    Also strips a leading minor./improved. prefix when what's left is a
    recognized Protect spell - real item data stores a Protect's tier as
    that prefix (see PROTECT_TIER_TO_SUFFIX), not the .i/.ii/.iii/.iv suffix
    everything else (and the internal Wanted Spell chip representation) uses,
    so without this an item like "minor.shock.protect" would never be
    recognized as the same base as a "shock.protect" want. Wanted-spell
    strings never carry this prefix themselves, so this is safe to apply
    unconditionally regardless of which side (item or wanted-chip) called it."""
    s = spell.strip().lower()
    m = re.match(r'^(minor|improved)\.(.+)$', s)
    if m and m.group(2) in PROTECT_SPELLS:
        return m.group(2)
    m = re.match(r'^(.*)\.(i|ii|iii|iv)$', s)
    return m.group(1) if m else s


def _spell_tier_rank(variant):
    """Rank of the tier suffix on a wanted-spell string (0 if untiered/any)."""
    m = re.match(r'^.*\.(i|ii|iii|iv)$', variant.strip().lower())
    return _TIER_RANK[m.group(1)] if m else 0


def _item_tier_rank(item_spell):
    """Tier rank of an actual item's Spell value. Ordinary tiered spells use
    the same .i/.ii/.iii/.iv suffix as _spell_tier_rank, but real Protect
    items store their tier differently: minor./improved. as a PREFIX, and no
    prefix at all for normal tier (ii) - see PROTECT_TIER_TO_SUFFIX. That
    bare-string convention only applies to actual item data, never to a
    Wanted Spell chip (where a bare "shock.protect" instead means "(any)"
    tier was picked) - so this is intentionally a separate function from
    _spell_tier_rank rather than an extension of it."""
    s = item_spell.strip().lower()
    m = re.match(r'^(minor|improved)\.(.+)$', s)
    if m and m.group(2) in PROTECT_SPELLS:
        return _TIER_RANK['i'] if m.group(1) == 'minor' else _TIER_RANK['iii']
    if s in PROTECT_SPELLS:
        return _TIER_RANK['ii']
    return _spell_tier_rank(s)


# Bank Build - parses a pasted bank/inventory listing (see
# App._find_bank_build) into a set of "owned item" keys that can be looked
# up against self.master_data. One line looks like:
#   12.) a softly glowing mushroom wreath necklace [60|Jewel|elementalprotect1]
# The bracket's tags are a numeric Level followed by a Slot name (or, for
# consumables/held junk that isn't equippable gear at all, something else
# entirely - "Held Left", "Uses N", or no numeric Level as the first tag -
# those are skipped outright rather than guessed at). The trailing tags
# after Level/Slot are the game client's own abbreviated stat codes (e.g.
# "CP2", "M5") - deliberately never decoded here: the same item name
# already exists in the master database with real Spell/Sigil/Type data,
# so matching by (cleaned name, Slot, Level) borrows that directly instead
# of trying to reverse-engineer the abbreviations.
_BANK_LINE_RE = re.compile(r'^\s*\d+\.\)\s*(.+?)\s*\[(.+)\]\s*$')
_BANK_SLOT_MAP = {
    'head': 'head', 'feet': 'feet', 'hands': 'hands', 'legs': 'legs',
    'jewel': 'jewel', 'cloak': 'cloak', 'body': 'body', 'weap': 'weapon',
    'shield': 'shield',
}


def _parse_bank_paste_text(text):
    """Parse a pasted bank/inventory listing into (owned_keys, recognized)
    - owned_keys is a set of (cleaned item name lower, canonical slot,
    level string) tuples, one per recognized equippable line; recognized
    is how many lines qualified (lines that don't match the "N.) name
    [tags]" format at all - like the "Items in Strongbox..." header - or
    that describe a consumable/held item rather than gear are silently
    skipped, not counted as an error)."""
    owned_keys = set()
    recognized = 0
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        m = _BANK_LINE_RE.match(line)
        if not m:
            continue
        raw_name, tags_str = m.groups()
        tags = [t.strip() for t in tags_str.split('|')]
        if len(tags) < 2 or not tags[0].isdigit():
            continue  # consumable/potion (e.g. "restpot5") - no numeric Level
        level = tags[0]
        slot = _BANK_SLOT_MAP.get(tags[1].strip().lower())
        if not slot:
            continue  # "Held Left", "Uses N", or an unrecognized slot tag
        cleaned = LootParser.clean_item_name(raw_name).strip().lower()
        if not cleaned:
            continue
        owned_keys.add((cleaned, slot, level))
        recognized += 1
    return owned_keys, recognized


def _bank_item_key(item):
    """Same (name, slot, level) shape as _parse_bank_paste_text's keys, but
    computed from a master_data item dict instead of a pasted line."""
    return (
        (item.get('Item') or '').strip().lower(),
        (item.get('Slot') or '').strip().lower(),
        str(item.get('Level') or '').strip(),
    )


class ItemMatchDialog(tk.Toplevel):
    """Modal dialog listing close-spelling matches for a searched item name,
    letting the user confirm which one they actually meant."""
    def __init__(self, parent, query: str, matches: list):
        super().__init__(parent)
        self.result = None
        self.title("Did you mean...?")
        self.resizable(False, False)
        self.grab_set()

        ttk.Label(self, text=f"No exact match for '{query}'. Closest matches:",
                 font=('Arial', 9)).pack(anchor='w', padx=10, pady=(10,4))

        self.listbox = tk.Listbox(self, height=min(6, len(matches)), width=50)
        for m in matches:
            self.listbox.insert(tk.END, m)
        self.listbox.selection_set(0)
        self.listbox.pack(padx=10, pady=4)

        bf = ttk.Frame(self)
        bf.pack(pady=10)
        ttk.Button(bf, text="Use Selected", command=self._confirm).pack(side='left', padx=8)
        ttk.Button(bf, text="Cancel", command=self.destroy).pack(side='left', padx=8)

    def _confirm(self):
        sel = self.listbox.curselection()
        if sel:
            self.result = self.listbox.get(sel[0])
        self.destroy()


class FieldEditorDialog(tk.Toplevel):
    """Modal dialog to add or edit a single field."""
    def __init__(self, parent, mode: str, existing: dict = None):
        super().__init__(parent)
        self.result = None
        self.mode = mode
        source_opts = {'chat': CHAT_SOURCES, 'combat': COMBAT_SOURCES, 'loot': LOOT_SOURCES}[mode]

        self.title("Edit Field" if existing else "Add Field")
        self.resizable(False, False)
        self.grab_set()

        pad = {'padx': 10, 'pady': 6}

        ttk.Label(self, text="Column label (header text):").grid(row=0, column=0, sticky='w', **pad)
        self.lbl_var = tk.StringVar(value=existing['label'] if existing else '')
        ttk.Entry(self, textvariable=self.lbl_var, width=26).grid(row=0, column=1, **pad)

        ttk.Label(self, text="Data source:").grid(row=1, column=0, sticky='w', **pad)
        self.src_var = tk.StringVar(value=existing['source_key'] if existing else source_opts[0])
        ttk.Combobox(self, textvariable=self.src_var, values=source_opts,
                     state='readonly', width=24).grid(row=1, column=1, **pad)

        ttk.Label(self, text="Column position (1=A, 2=B…):").grid(row=2, column=0, sticky='w', **pad)
        self.col_var = tk.IntVar(value=existing['col'] if existing else 1)
        ttk.Spinbox(self, textvariable=self.col_var, from_=1, to=30,
                    width=6).grid(row=2, column=1, sticky='w', **pad)

        bf = ttk.Frame(self)
        bf.grid(row=3, column=0, columnspan=2, pady=10)
        ttk.Button(bf, text="Save", command=self._save).pack(side='left', padx=8)
        ttk.Button(bf, text="Cancel", command=self.destroy).pack(side='left', padx=8)

    def _save(self):
        lbl = self.lbl_var.get().strip()
        if not lbl:
            messagebox.showwarning("Missing", "Enter a label.", parent=self)
            return
        self.result = {'label': lbl, 'source_key': self.src_var.get(), 'col': self.col_var.get()}
        self.destroy()


class SnapshotViewer(tk.Toplevel):
    """Shows the raw parsed rows for one sheet type so user can verify."""
    def __init__(self, parent, title: str, rows: list, fields: list):
        super().__init__(parent)
        self.title(f"Snapshot — {title}  ({len(rows)} rows)")
        self.geometry("1100x540")

        if not rows:
            ttk.Label(self, text="No data parsed yet.  Run Parse first.",
                      font=('Arial', 12)).pack(expand=True)
            return

        cols = [f['label'] for f in fields]
        keys = [f['source_key'] for f in fields]

        frame = ttk.Frame(self)
        frame.pack(fill='both', expand=True)

        tv = ttk.Treeview(frame, columns=cols, show='headings', height=22)
        for c in cols:
            tv.heading(c, text=c)
            tv.column(c, width=max(80, len(c) * 9), stretch=True)

        vsb = ttk.Scrollbar(frame, orient='vertical',   command=tv.yview)
        hsb = ttk.Scrollbar(frame, orient='horizontal', command=tv.xview)
        tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tv.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        for row in rows[:2000]:
            tv.insert('', 'end', values=[str(row.get(k, '')) for k in keys])

        ttk.Label(self, text=f"Showing up to 2 000 of {len(rows)} rows",
                  foreground='gray').pack(pady=4)


class LogSearchViewer(tk.Toplevel):
    """Shows every raw log line matching a search term, across all loaded files."""
    def __init__(self, parent, query: str, results: list):
        super().__init__(parent)
        self.title(f"Search Results — '{query}'  ({len(results)} matches)")
        self.geometry("1150x540")

        cols = ('File', 'Line #', 'Timestamp', 'Text')
        frame = ttk.Frame(self)
        frame.pack(fill='both', expand=True)

        tv = ttk.Treeview(frame, columns=cols, show='headings', height=22)
        widths = {'File': 220, 'Line #': 70, 'Timestamp': 90, 'Text': 700}
        for c in cols:
            tv.heading(c, text=c)
            tv.column(c, width=widths[c], anchor='center' if c == 'Line #' else 'w')

        vsb = ttk.Scrollbar(frame, orient='vertical',   command=tv.yview)
        hsb = ttk.Scrollbar(frame, orient='horizontal', command=tv.xview)
        tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tv.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        for file_name, line_num, timestamp, text in results[:5000]:
            tv.insert('', 'end', values=(file_name, line_num, timestamp, text))

        shown = min(len(results), 5000)
        ttk.Label(self, text=f"Showing {shown} of {len(results)} matching lines",
                  foreground='gray').pack(pady=4)


class DropSnapshotViewer(tk.Toplevel):
    """Lists every drop event found for an item; selecting one shows a snapshot
    of the raw log from the last timestamp through that drop line."""
    def __init__(self, parent, query: str, drops: list):
        super().__init__(parent)
        self.title(f"Drop Search — '{query}'  ({len(drops)} drops)")
        self.geometry("1150x650")
        self.drops = drops

        paned = ttk.PanedWindow(self, orient='vertical')
        paned.pack(fill='both', expand=True)

        list_frame = ttk.Frame(paned)
        paned.add(list_frame, weight=1)

        cols = ('File', 'Timestamp', 'Item', 'Mob')
        self.tv = ttk.Treeview(list_frame, columns=cols, show='headings', height=12)
        widths = {'File': 220, 'Timestamp': 90, 'Item': 320, 'Mob': 220}
        for c in cols:
            self.tv.heading(c, text=c)
            self.tv.column(c, width=widths[c], anchor='center' if c == 'Timestamp' else 'w')

        vsb = ttk.Scrollbar(list_frame, orient='vertical', command=self.tv.yview)
        self.tv.configure(yscrollcommand=vsb.set)
        self.tv.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        for i, d in enumerate(drops):
            self.tv.insert('', 'end', iid=str(i),
                           values=(d['file'], d['timestamp'], d['item'], d['mob']))
        self.tv.bind('<<TreeviewSelect>>', self._on_select)

        snapshot_frame = ttk.LabelFrame(paned, text="Snapshot (last timestamp → drop)", padding=6)
        paned.add(snapshot_frame, weight=1)

        self.snapshot_text = tk.Text(snapshot_frame, wrap='none', height=10, state='disabled')
        snap_vsb = ttk.Scrollbar(snapshot_frame, orient='vertical', command=self.snapshot_text.yview)
        snap_hsb = ttk.Scrollbar(snapshot_frame, orient='horizontal', command=self.snapshot_text.xview)
        self.snapshot_text.configure(yscrollcommand=snap_vsb.set, xscrollcommand=snap_hsb.set)
        self.snapshot_text.grid(row=0, column=0, sticky='nsew')
        snap_vsb.grid(row=0, column=1, sticky='ns')
        snap_hsb.grid(row=1, column=0, sticky='ew')
        snapshot_frame.rowconfigure(0, weight=1)
        snapshot_frame.columnconfigure(0, weight=1)

        ttk.Label(self, text=f"{len(drops)} drop(s) found — select a row above to view its snapshot",
                  foreground='gray').pack(pady=4)

        if drops:
            self.tv.selection_set('0')
            self._on_select()

    def _on_select(self, event=None):
        sel = self.tv.selection()
        if not sel:
            return
        drop = self.drops[int(sel[0])]
        self.snapshot_text.config(state='normal')
        self.snapshot_text.delete('1.0', tk.END)
        self.snapshot_text.insert('1.0', drop['snapshot'])
        self.snapshot_text.config(state='disabled')


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"🎮 Olmran Log Parser and Gear Set Creator v{VERSION}")
        self.geometry("1350x815")
        self.minsize(1150, 750)
        self.resizable(True, True)
        self._set_app_icon()

        self.files: list[dict] = []       # {path, name, size, type}
        self.parsed = {'chat': [], 'combat': [], 'loot': []}
        self.fields = {k: [dict(f) for f in v] for k, v in DEFAULT_FIELDS.items()}
        
        # Config file for persistent settings
        self.config_file = os.path.join(os.path.expanduser("~"), ".gaming_log_parser_config.json")
        
        # Load saved directories or use home as default
        self._load_config()

        self._build_ui()

        # Auto-load the bundled community equipment list on startup, if present
        self._auto_load_community_list()

        # Save config on close
        self.protocol("WM_DELETE_WINDOW", self._on_closing)
    
    def _load_config(self):
        """Load saved configuration from file"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                    self.last_open_dir = config.get('last_open_dir', os.path.expanduser("~"))
                    self.last_save_dir = config.get('last_save_dir', os.path.expanduser("~"))
                    self.armor_defaults = config.get('armor_defaults', {})
                    self.weapon_combo_defaults = config.get('weapon_combo_defaults', {})
                    # Raw data only (no tk.StringVar yet - the Saved Builds tab
                    # isn't built until later in __init__) - _build_saved_builds_tab
                    # turns this into self.saved_builds.
                    self._persisted_saved_builds = config.get('saved_builds', [])
                    self._persisted_saved_build_counter = config.get('saved_build_counter', 0)
            else:
                self.last_open_dir = os.path.expanduser("~")
                self.last_save_dir = os.path.expanduser("~")
                self.armor_defaults = {}
                self.weapon_combo_defaults = {}
                self._persisted_saved_builds = []
                self._persisted_saved_build_counter = 0
        except Exception:
            self.last_open_dir = os.path.expanduser("~")
            self.last_save_dir = os.path.expanduser("~")
            self.armor_defaults = {}
            self.weapon_combo_defaults = {}
            self._persisted_saved_builds = []
            self._persisted_saved_build_counter = 0

    def _save_config(self):
        """Save configuration to file"""
        try:
            config = {
                'last_open_dir': self.last_open_dir,
                'last_save_dir': self.last_save_dir,
                'armor_defaults': getattr(self, 'armor_defaults', {}),
                'weapon_combo_defaults': getattr(self, 'weapon_combo_defaults', {}),
                'saved_builds': [
                    {'name': save['name'].get(), 'headers': list(save['headers']),
                     'rows': [list(row) for row in save['rows']]}
                    for save in getattr(self, 'saved_builds', [])
                ],
                'saved_build_counter': getattr(self, 'saved_build_counter', 0),
            }
            with open(self.config_file, 'w') as f:
                json.dump(config, f)
        except Exception:
            pass  # Silently fail if can't save config
    
    def _on_closing(self):
        """Handle window closing"""
        self._save_config()
        self.destroy()

    # ── BUILD UI ──────────────────────────────────────────────
    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use('clam')
        style.configure('TNotebook.Tab', font=('Arial', 10, 'bold'), padding=[12, 6])
        style.configure('Accent.TButton', font=('Arial', 10, 'bold'))

        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True, padx=8, pady=8)

        self.tab_parse  = ttk.Frame(nb, padding=12)
        self.tab_fields = ttk.Frame(nb, padding=12)
        self.tab_export = ttk.Frame(nb, padding=12)
        self.tab_build = ttk.Frame(nb, padding=12)
        self.tab_results = ttk.Frame(nb, padding=12)
        self.tab_saved = ttk.Frame(nb, padding=12)
        self.tab_area_items = ttk.Frame(nb, padding=12)

        nb.add(self.tab_parse,  text='▶  Parse')
        nb.add(self.tab_fields, text='⚙  Fields')
        nb.add(self.tab_export, text='💾  Export')
        nb.add(self.tab_build, text='🔨  Build')
        nb.add(self.tab_results, text='📊  Results')  # Always visible
        nb.add(self.tab_saved, text='📌  Saved Builds')  # One tab holds every save
        nb.add(self.tab_area_items, text='🗺  Area Items')

        self.notebook = nb  # Store reference to notebook

        self._build_parse_tab()
        self._build_fields_tab()
        self._build_export_tab()
        self._build_build_tab()
        self._build_results_tab()
        self._build_saved_builds_tab()
        self._build_area_items_tab()

    # ── PARSE TAB (FILES + PARSING) ───────────────────────────
    def _build_parse_tab(self):
        t = self.tab_parse
        
        # ═══ FILES SECTION ═══
        ttk.Label(t, text="Load Log Files",
                  font=('Arial', 13, 'bold')).pack(anchor='w', pady=(0, 8))

        top = ttk.Frame(t)
        top.pack(fill='x')
        ttk.Button(top, text="➕ Add Files",   command=self._add_files).pack(side='left', padx=(0,6))
        ttk.Button(top, text="📁 Add Folder",  command=self._add_folder).pack(side='left', padx=(0,6))
        ttk.Button(top, text="❌ Remove",      command=self._remove_files).pack(side='left', padx=(0,6))
        ttk.Button(top, text="🗑 Clear All",   command=self._clear_files).pack(side='left')
        self.file_count_lbl = ttk.Label(top, text="", foreground='#444')
        self.file_count_lbl.pack(side='right')

        cols = ('File', 'Type', 'Size')
        frame = ttk.Frame(t)
        frame.pack(fill='both', expand=True, pady=8)

        self.file_tv = ttk.Treeview(frame, columns=cols, show='headings', height=8)
        self.file_tv.heading('File', text='Filename')
        self.file_tv.heading('Type', text='Auto-Detected Type')
        self.file_tv.heading('Size', text='Size')
        self.file_tv.column('File', width=380)
        self.file_tv.column('Type', width=150)
        self.file_tv.column('Size', width=90, anchor='e')

        vsb = ttk.Scrollbar(frame, command=self.file_tv.yview)
        self.file_tv.configure(yscrollcommand=vsb.set)
        self.file_tv.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        ttk.Label(t, text="💡 File types are auto-detected (CHAT = Chat logs, ACTION = Action logs)",
                 font=('Arial', 8, 'italic'), foreground='#666').pack(anchor='w', pady=(0,12))

        # ═══ SEARCH SECTION ═══
        # Search across every loaded file's actual lines - independent of Run
        # Parse, so you can e.g. find every time an item dropped across a large
        # batch of logs without parsing them all first.
        search_frame = ttk.LabelFrame(t, text="Search Logs", padding=10)
        search_frame.pack(fill='x', pady=(0,12))

        search_row = ttk.Frame(search_frame)
        search_row.pack(fill='x')
        ttk.Label(search_row, text="Find:", width=8).pack(side='left')
        self.search_query_var = tk.StringVar(value='')
        search_entry = ttk.Entry(search_row, textvariable=self.search_query_var, width=40)
        search_entry.pack(side='left', padx=4)
        search_entry.bind('<Return>', lambda e: self._search_logs())
        self.search_case_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(search_row, text="Case sensitive",
                       variable=self.search_case_var).pack(side='left', padx=8)
        self.search_drops_only_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(search_row, text="Drops only (with snapshot)",
                       variable=self.search_drops_only_var).pack(side='left', padx=8)
        ttk.Button(search_row, text="🔍 Search",
                  command=self._search_logs).pack(side='left', padx=4)

        ttk.Label(search_frame,
                 text="💡 \"Drops only\" catches every instance the item actually dropped (ignoring quality/material "
                      "prefixes like \"bright\" or \"glowing\", same as the Loot parser), and for each one shows a "
                      "snapshot of everything from the last timestamp through the drop line. Uncheck it to just "
                      "search every loaded file's raw text for any mention of the term.",
                 font=('Arial', 8, 'italic'), foreground='#666', wraplength=900,
                 justify='left').pack(anchor='w', pady=(4,0))

        # ═══ PARSING SECTION ═══
        parse_frame = ttk.LabelFrame(t, text="Parse Options", padding=10)
        parse_frame.pack(fill='x', pady=(0,12))
        
        ttk.Label(parse_frame, text="What to Parse:", 
                 font=('Arial', 9)).pack(anchor='w', pady=(0,4))
        
        opts_frame = ttk.Frame(parse_frame)
        opts_frame.pack(anchor='w', padx=20)
        
        self.do_chat = tk.BooleanVar(value=False)
        self.do_combat = tk.BooleanVar(value=False)
        self.do_loot   = tk.BooleanVar(value=False)
        
        self.chat_cb = ttk.Checkbutton(opts_frame, text="💬 Chat", 
                                       variable=self.do_chat)
        self.chat_cb.pack(side='left', padx=(0,20))
        
        self.combat_cb = ttk.Checkbutton(opts_frame, text="⚔️ Combat", 
                                         variable=self.do_combat)
        self.combat_cb.pack(side='left', padx=(0,20))
        
        self.loot_cb = ttk.Checkbutton(opts_frame, text="💎 Loot", 
                                       variable=self.do_loot)
        self.loot_cb.pack(side='left')

        run_frame = ttk.Frame(t)
        run_frame.pack(pady=8)
        ttk.Button(run_frame, text="▶  Run Parse",
                   command=self._run_parse, width=22, style='Accent.TButton').pack(side='left', padx=6)
        ttk.Button(run_frame, text="👁  Snapshot: Chat",
                   command=lambda: self._show_snapshot('chat')).pack(side='left', padx=6)
        ttk.Button(run_frame, text="👁  Snapshot: Combat",
                   command=lambda: self._show_snapshot('combat')).pack(side='left', padx=6)
        ttk.Button(run_frame, text="👁  Snapshot: Loot",
                   command=lambda: self._show_snapshot('loot')).pack(side='left', padx=6)

        self.status_var = tk.StringVar(value="Load files and click 'Run Parse' to begin")
        ttk.Label(t, textvariable=self.status_var,
                  font=('Arial', 10), foreground='#222').pack(anchor='w', pady=(12,0))

        # Mini stats
        self.stats_frame = ttk.LabelFrame(t, text="Parse Results", padding=10)
        self.stats_frame.pack(fill='x', pady=(12,0))
        
        self.stat_labels = {}
        for i, (k, icon) in enumerate([('chat','💬'),('combat','⚔️'),('loot','💎')]):
            f = ttk.Frame(self.stats_frame)
            f.grid(row=0, column=i, padx=20, pady=6)
            ttk.Label(f, text=f"{icon} {k.title()}", font=('Arial', 10, 'bold')).pack()
            lbl = ttk.Label(f, text="0 rows", foreground='#555')
            lbl.pack()
            self.stat_labels[k] = lbl

    # ── FIELDS TAB ────────────────────────────────────────────
    def _build_fields_tab(self):
        t = self.tab_fields
        ttk.Label(t, text="Customize Output Columns",
                  font=('Arial', 13, 'bold')).pack(anchor='w', pady=(0,6))
        ttk.Label(t, text="Each row below is one column in the exported spreadsheet.\n"
                           "Drag rows to reorder, or use Add/Edit/Remove. Column position sets the Excel column (1=A).",
                  foreground='#444').pack(anchor='w', pady=(0,8))

        ttk.Label(t, 
                 text="💡 Editing fields for types selected in Parse tab (Chat/Combat/Loot checkboxes)", 
                 font=('Arial', 9, 'italic'), foreground='#666').pack(anchor='w', pady=(0,8))

        # Create notebook for multiple field types
        self.fields_notebook = ttk.Notebook(t)
        self.fields_notebook.pack(fill='both', expand=True, pady=(0,6))
        
        # Create tabs for each type
        self.field_frames = {}
        self.field_tvs = {}
        
        for field_type, icon in [('chat', '💬'), ('combat', '⚔️'), ('loot', '💎')]:
            frame = ttk.Frame(self.fields_notebook, padding=12)
            self.fields_notebook.add(frame, text=f"{icon} {field_type.title()}")
            
            # Treeview for this type
            tree_frame = ttk.Frame(frame)
            tree_frame.pack(fill='both', expand=True, pady=(0,6))
            
            cols = ('Col #', 'Header Label', 'Data Source')
            tv = ttk.Treeview(tree_frame, columns=cols, show='headings', height=13)
            for c, w in zip(cols, [70, 220, 220]):
                tv.heading(c, text=c)
                tv.column(c, width=w)
            vsb = ttk.Scrollbar(tree_frame, command=tv.yview)
            tv.configure(yscrollcommand=vsb.set)
            tv.pack(side='left', fill='both', expand=True)
            vsb.pack(side='right', fill='y')
            
            self.field_frames[field_type] = frame
            self.field_tvs[field_type] = tv

        # Buttons (work on currently visible tab)
        btns = ttk.Frame(t)
        btns.pack(anchor='w')
        ttk.Button(btns, text="➕ Add",          command=self._add_field).pack(side='left', padx=4)
        ttk.Button(btns, text="✏️ Edit",         command=self._edit_field).pack(side='left', padx=4)
        ttk.Button(btns, text="❌ Remove",       command=self._remove_field).pack(side='left', padx=4)
        ttk.Button(btns, text="⬆ Move Up",      command=self._move_up).pack(side='left', padx=4)
        ttk.Button(btns, text="⬇ Move Down",    command=self._move_down).pack(side='left', padx=4)
        ttk.Button(btns, text="🔄 Reset Defaults", command=self._reset_fields).pack(side='left', padx=4)

        self._refresh_all_fields()

    # ── EXPORT TAB ────────────────────────────────────────────
    def _build_export_tab(self):
        t = self.tab_export
        ttk.Label(t, text="Export to Excel",
                  font=('Arial', 13, 'bold')).pack(anchor='w', pady=(0,10))

        info_frame = ttk.Frame(t)
        info_frame.pack(fill='x', pady=(0,12))
        ttk.Label(info_frame, 
                 text="💡 Export settings follow Parse tab selections (Chat/Combat/Loot checkboxes)", 
                 font=('Arial', 9, 'italic'), foreground='#666').pack(anchor='w')
        
        # Summary sheet option
        summary_frame = ttk.Frame(t)
        summary_frame.pack(fill='x', pady=(0,12))
        self.exp_summary = tk.BooleanVar(value=False)
        ttk.Checkbutton(summary_frame, text="📋 Include Summary Sheet", 
                       variable=self.exp_summary).pack(anchor='w')

        # Master file option
        master_frame = ttk.LabelFrame(t, text="Master Database (Loot Only)", padding=10)
        master_frame.pack(fill='x', pady=(0,12))
        
        checkbox_frame = ttk.Frame(master_frame)
        checkbox_frame.pack(anchor='w', pady=4)
        
        self.use_master = tk.BooleanVar(value=False)
        ttk.Checkbutton(checkbox_frame, text="📚 Append to Master Database", 
                       variable=self.use_master).pack(side='left', padx=(0, 20))
        
        self.exclude_fodder = tk.BooleanVar(value=True)
        ttk.Checkbutton(checkbox_frame, text="🚫 Exclude Fodder", 
                       variable=self.exclude_fodder).pack(side='left')
        
        ttk.Label(master_frame, text="Master file:", 
                 font=('Arial', 9)).pack(side='left', padx=(0,8))
        self.master_path_var = tk.StringVar(value='')
        ttk.Entry(master_frame, textvariable=self.master_path_var, 
                 width=35, state='readonly').pack(side='left', padx=4)
        ttk.Button(master_frame, text="Browse...", 
                  command=self._browse_master).pack(side='left', padx=2)
        ttk.Button(master_frame, text="Create New", 
                  command=self._create_new_master).pack(side='left', padx=2)
        
        ttk.Label(master_frame, 
                 text="💡 Master database builds a complete item list across all sessions",
                 font=('Arial', 8, 'italic'), foreground='#666').pack(anchor='w', pady=(4,0))

        fname_frame = ttk.Frame(t)
        fname_frame.pack(fill='x', pady=6)
        ttk.Label(fname_frame, text="Output filename:").pack(side='left')
        self.fname_var = tk.StringVar(value='game_log_export')
        ttk.Entry(fname_frame, textvariable=self.fname_var, width=32).pack(side='left', padx=8)
        ttk.Label(fname_frame, text=".xlsx").pack(side='left')
        
        # Export options
        export_opts_frame = ttk.Frame(t)
        export_opts_frame.pack(fill='x', pady=(0,8))
        
        self.separate_action_exports = tk.BooleanVar(value=False)
        ttk.Checkbutton(export_opts_frame, 
                       text="📂 Export Combat and Loot as separate files (when both selected)", 
                       variable=self.separate_action_exports).pack(anchor='w')

        ttk.Button(t, text="💾  Export to Excel",
                   command=self._export, width=26).pack(pady=8)

        self.export_status = ttk.Label(t, text="", foreground='#226622', font=('Arial', 10))
        self.export_status.pack(anchor='w')
    
    def _browse_master(self):
        """Browse for master database file"""
        path = filedialog.askopenfilename(
            title='Select Master Database File',
            initialdir=self.last_save_dir,
            filetypes=[('Excel', '*.xlsx'), ('All', '*.*')])
        if path:
            self.master_path_var.set(path)
            self.last_save_dir = os.path.dirname(path)
    
    def _create_new_master(self):
        """Create a new master database file"""
        # Suggest filename
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d')
        default_name = f"Master_Loot_Database_{timestamp}"
        
        path = filedialog.asksaveasfilename(
            title='Create New Master Database',
            initialdir=self.last_save_dir,
            initialfile=default_name + '.xlsx',
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx'), ('All', '*.*')])
        
        if not path:
            return
        
        try:
            # Create new Excel file with Loot sheet
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            ws = wb.create_sheet("Loot")
            
            # Write header row using loot fields
            from openpyxl.styles import Font, PatternFill, Alignment
            
            headers = []
            max_col = max(f['col'] for f in self.fields['loot'])
            headers = [''] * max_col
            for f in self.fields['loot']:
                headers[f['col'] - 1] = f['label']
            
            ws.append(headers)
            
            # Style header
            hdr_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            hdr_fill = PatternFill(fill_type='solid', start_color='4472C4', end_color='4472C4')
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(1, col_idx)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.freeze_panes = 'A2'
            
            # Save
            wb.save(path)
            
            # Set as master file
            self.master_path_var.set(path)
            self.use_master.set(True)
            self.last_save_dir = os.path.dirname(path)
            
            messagebox.showinfo("Master Created", 
                f"✓ New master database created!\n\n{path}\n\n"
                "Now parse your logs and export - new items will be added to this master file.")
                
        except Exception as e:
            messagebox.showerror("Create Error", f"Failed to create master file:\n{str(e)}")
    
    def _append_to_master(self):
        """Append new loot items to master database, avoiding duplicates"""
        master_path = self.master_path_var.get()
        if not os.path.exists(master_path):
            messagebox.showwarning("Master File Not Found", 
                f"Master database file not found:\n{master_path}\n\nSkipping master append.")
            return
        
        try:
            # Load existing master file
            wb = openpyxl.load_workbook(master_path)
            
            # Find or create Loot sheet
            if 'Loot' in wb.sheetnames:
                ws = wb['Loot']
            else:
                messagebox.showwarning("No Loot Sheet", 
                    "Master file has no 'Loot' sheet. Skipping master append.")
                return
            
            # Read existing items to avoid duplicates
            # Key: (Realm, Area, Mob, Item)
            existing_items = set()
            for row in range(2, ws.max_row + 1):  # Skip header
                realm = ws.cell(row, 1).value or ''
                area = ws.cell(row, 2).value or ''
                mob = ws.cell(row, 3).value or ''
                item = ws.cell(row, 4).value or ''
                existing_items.add((realm, area, mob, item))
            
            # Add new items that aren't already in master
            new_count = 0
            skipped_fodder = 0
            for loot_item in self.parsed['loot']:
                # Check if we should exclude fodder
                if self.exclude_fodder.get():
                    slot = loot_item.get('Slot', '').lower()
                    if slot == 'fodder':
                        skipped_fodder += 1
                        continue
                
                key = (loot_item['Realm'], loot_item['Area'], 
                      loot_item['Mob'], loot_item['Item'])
                if key not in existing_items:
                    # Append new row
                    row_num = ws.max_row + 1
                    for field in self.fields['loot']:
                        col = field['col']
                        value = loot_item.get(field['source_key'], '')
                        ws.cell(row_num, col, value)
                    new_count += 1
                    existing_items.add(key)  # Prevent duplicates within this session
            
            # Save master file
            wb.save(master_path)
            
            # Show results
            message = f"✓ Added {new_count} new items to master database"
            if skipped_fodder > 0:
                message += f"\n🚫 Excluded {skipped_fodder} fodder items"
            message += f"\n\n{master_path}"
            
            if new_count > 0:
                messagebox.showinfo("Master Updated", message)
            else:
                if skipped_fodder > 0:
                    messagebox.showinfo("Master Up-to-Date", 
                        f"No new items to add\n🚫 Excluded {skipped_fodder} fodder items")
                else:
                    messagebox.showinfo("Master Up-to-Date", 
                        "No new items to add (all items already in master)")
                    
        except Exception as e:
            messagebox.showerror("Master Append Error", 
                f"Failed to append to master:\n{str(e)}")


    # ── FILE HELPERS ──────────────────────────────────────────
    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title='Select log files',
            initialdir=self.last_open_dir,
            filetypes=[('Log/Text', '*.log *.txt'), ('All', '*.*')])
        
        if paths:
            # Save the directory for next time
            self.last_open_dir = os.path.dirname(paths[0])
        
        for p in paths:
            if not any(f['path'] == p for f in self.files):
                name = os.path.basename(p)
                # Default to loot for ACTION files, chat for CHAT files
                ftype = 'loot' if 'ACTION' in name.upper() else \
                        'chat' if 'CHAT' in name.upper() else 'loot'
                self.files.append({'path': p, 'name': name,
                                   'size': os.path.getsize(p), 'type': ftype})
        self._refresh_file_list()
        self._update_parse_options()

    def _add_folder(self):
        folder = filedialog.askdirectory(
            title='Select folder',
            initialdir=self.last_open_dir)
        if not folder:
            return
        
        # Save the directory for next time
        self.last_open_dir = folder
        
        for ext in ('*.log', '*.txt'):
            for p in Path(folder).glob(ext):
                sp = str(p)
                if not any(f['path'] == sp for f in self.files):
                    name = p.name
                    # Default to loot for ACTION files
                    ftype = 'loot' if 'ACTION' in name.upper() else \
                            'chat' if 'CHAT' in name.upper() else 'loot'
                    self.files.append({'path': sp, 'name': name,
                                       'size': os.path.getsize(sp), 'type': ftype})
        self._refresh_file_list()
        self._update_parse_options()

    def _remove_files(self):
        for iid in self.file_tv.selection():
            idx = self.file_tv.index(iid)
            if 0 <= idx < len(self.files):
                del self.files[idx]
        self._refresh_file_list()
        self._update_parse_options()

    def _clear_files(self):
        if self.files and messagebox.askyesno("Clear", "Remove all loaded files?"):
            self.files.clear()
            self._refresh_file_list()
            self._update_parse_options()

    def _update_parse_options(self):
        """Update parse checkboxes based on loaded files"""
        # Check what types of files are loaded
        has_chat = any(f['type'] == 'chat' for f in self.files)
        has_action = any(f['type'] in ('combat', 'loot') for f in self.files)
        
        if not self.files:
            # No files: uncheck and enable all
            self.do_chat.set(False)
            self.do_combat.set(False)
            self.do_loot.set(False)
            self.chat_cb.config(state='normal')
            self.combat_cb.config(state='normal')
            self.loot_cb.config(state='normal')
        elif has_chat and not has_action:
            # Only chat files: check chat, disable action options
            self.do_chat.set(True)
            self.do_combat.set(False)
            self.do_loot.set(False)
            self.chat_cb.config(state='normal')
            self.combat_cb.config(state='disabled')
            self.loot_cb.config(state='disabled')
        elif has_action and not has_chat:
            # Only action files: check loot, disable chat
            self.do_chat.set(False)
            self.do_combat.set(False)
            self.do_loot.set(True)
            self.chat_cb.config(state='disabled')
            self.combat_cb.config(state='normal')
            self.loot_cb.config(state='normal')
        else:
            # Both types: check both, enable all
            self.do_chat.set(True)
            self.do_combat.set(False)
            self.do_loot.set(True)
            self.chat_cb.config(state='normal')
            self.combat_cb.config(state='normal')
            self.loot_cb.config(state='normal')
    
    def _select_all_files(self):
        """Select all files in the file list"""
        for item in self.file_tv.get_children():
            self.file_tv.selection_add(item)
    
    def _apply_type(self):
        t = self.type_var.get()
        for iid in self.file_tv.selection():
            idx = self.file_tv.index(iid)
            if 0 <= idx < len(self.files):
                self.files[idx]['type'] = t
        self._refresh_file_list()

    def _refresh_file_list(self):
        self.file_tv.delete(*self.file_tv.get_children())
        for f in self.files:
            kb = f['size'] / 1024
            self.file_tv.insert('', 'end',
                values=(f['name'], f['type'], f"{kb:.1f} KB"))
        n = len(self.files)
        self.file_count_lbl.config(text=f"{n} file{'s' if n != 1 else ''} loaded")

    # ── PARSE HELPERS ─────────────────────────────────────────
    def _run_parse(self):
        if not self.files:
            messagebox.showwarning("No Files", "Load some log files first.")
            return

        self.parsed = {'chat': [], 'chat_files': {}, 'combat': [], 'loot': []}
        errors = []

        for f in self.files:
            ftype = f['type']
            try:
                # Auto-parse based on detected file type and checkboxes
                if ftype == 'chat' and self.do_chat.get():
                    chat_data = ChatParser.parse_file(f['path'])
                    self.parsed['chat'] += chat_data
                    # Store separately by filename for individual sheets
                    self.parsed['chat_files'][f['name']] = chat_data
                elif ftype == 'combat' and self.do_combat.get():
                    self.parsed['combat'] += CombatParser.parse_file(f['path'])
                elif ftype == 'loot' and self.do_loot.get():
                    self.parsed['loot'] += LootParser.parse_file(f['path'])
            except Exception as e:
                errors.append(f"{f['name']}: {e}")

        for k, lbl in self.stat_labels.items():
            if k in self.parsed:
                lbl.config(text=f"{len(self.parsed[k])} rows")

        msg = (f"✓  Chat: {len(self.parsed['chat'])} rows  |  "
               f"Combat: {len(self.parsed['combat'])} rows  |  "
               f"Loot: {len(self.parsed['loot'])} rows")
        if errors:
            msg += f"\n⚠ Errors: {'; '.join(errors)}"
        self.status_var.set(msg)

    def _show_snapshot(self, mode: str):
        rows = self.parsed.get(mode, [])
        if not rows:
            messagebox.showinfo("Snapshot",
                f"No {mode} data yet.\nRun Parse first (with {mode} files loaded).")
            return
        SnapshotViewer(self, mode.title(), rows, self.fields[mode])

    def _search_logs(self):
        """Search every loaded file for a term, e.g. an item name. In "Drops
        only" mode (default), finds every actual drop event of that item and
        shows a snapshot from the last timestamp through the drop; otherwise
        falls back to a plain raw-text search of every line."""
        query = self.search_query_var.get().strip()
        if not query:
            messagebox.showwarning("No Search Term", "Enter something to search for.")
            return
        if not self.files:
            messagebox.showwarning("No Files", "Load some log files first.")
            return

        if self.search_drops_only_var.get():
            drops, errors = self._find_item_drops(query)
            if not drops:
                msg = f"No drops of '{query}' were found in {len(self.files)} file(s)."
                if errors:
                    msg += f"\n\n{len(errors)} file(s) could not be read: " + "; ".join(errors)
                messagebox.showinfo("No Matches", msg)
                return
            DropSnapshotViewer(self, query, drops)
        else:
            results, errors = self._find_raw_lines(query)
            if not results:
                msg = f"No lines containing '{query}' were found in {len(self.files)} file(s)."
                if errors:
                    msg += f"\n\n{len(errors)} file(s) could not be read: " + "; ".join(errors)
                messagebox.showinfo("No Matches", msg)
                return
            LogSearchViewer(self, query, results)

        if errors:
            messagebox.showwarning("Some Files Skipped",
                f"{len(errors)} file(s) could not be read:\n" + "\n".join(errors))

    def _find_raw_lines(self, query):
        """Plain substring search of every loaded file's raw lines"""
        case_sensitive = self.search_case_var.get()
        needle = query if case_sensitive else query.lower()

        results = []
        errors = []
        for f in self.files:
            try:
                with open(f['path'], 'r', encoding='utf-8', errors='ignore') as fh:
                    for line_num, line in enumerate(fh, 1):
                        haystack = line if case_sensitive else line.lower()
                        if needle in haystack:
                            ts_match = re.match(r'^<(\d{2}:\d{2}:\d{2})>', line)
                            timestamp = ts_match.group(1) if ts_match else ''
                            results.append((f['name'], line_num, timestamp, line.strip()))
            except Exception as e:
                errors.append(f"{f['name']}: {e}")
        return results, errors

    def _find_item_drops(self, query):
        """Find every actual drop event of an item across all loaded files,
        matching the same way the Loot parser identifies drops (so quality
        and material prefixes like "bright" or "glowing" don't cause misses).
        Each result includes a snapshot of every line from the last timestamp
        seen through the drop line itself."""
        case_sensitive = self.search_case_var.get()
        needle = query if case_sensitive else query.lower()

        drops = []
        errors = []
        for f in self.files:
            try:
                with open(f['path'], 'r', encoding='utf-8', errors='ignore') as fh:
                    lines = fh.readlines()
            except Exception as e:
                errors.append(f"{f['name']}: {e}")
                continue

            current_ts = ''
            block_start = 0
            for idx, line in enumerate(lines):
                ts_m = LootParser.TS_RE.match(line)
                if ts_m:
                    current_ts = ts_m.group(1)
                    block_start = idx

                m = LootParser.DROP_RE.match(line.strip())
                if not m:
                    continue

                item_raw = m.group(3).strip()
                if re.match(r'a bag of (silver|coins?|gold)', item_raw, re.I):
                    continue

                item_clean = LootParser.clean_item_name(item_raw)
                haystack = item_clean if case_sensitive else item_clean.lower()
                if needle not in haystack:
                    continue

                mob_name = ((m.group(1) or '') + m.group(2)).strip()
                mob_name = re.sub(r'^(A|An|The)\s+', '', mob_name, flags=re.IGNORECASE).strip()

                snapshot = ''.join(lines[block_start:idx + 1]).rstrip('\n')
                drops.append({
                    'file': f['name'],
                    'timestamp': current_ts,
                    'item': item_clean,
                    'mob': mob_name,
                    'snapshot': snapshot,
                })

        return drops, errors

    # ── FIELD HELPERS ─────────────────────────────────────────
    def _get_current_field_mode(self):
        """Get the currently selected field type from notebook tab"""
        tab_index = self.fields_notebook.index(self.fields_notebook.select())
        return ['chat', 'combat', 'loot'][tab_index]
    
    def _refresh_all_fields(self):
        """Refresh all field treeviews"""
        for mode in ['chat', 'combat', 'loot']:
            tv = self.field_tvs[mode]
            tv.delete(*tv.get_children())
            for f in sorted(self.fields[mode], key=lambda x: x['col']):
                tv.insert('', 'end',
                         values=(f['col'], f['label'], f['source_key']))
    
    def _refresh_fields(self):
        """Refresh current field treeview"""
        mode = self._get_current_field_mode()
        tv = self.field_tvs[mode]
        tv.delete(*tv.get_children())
        for f in sorted(self.fields[mode], key=lambda x: x['col']):
            tv.insert('', 'end',
                     values=(f['col'], f['label'], f['source_key']))

    def _add_field(self):
        mode = self._get_current_field_mode()
        dlg = FieldEditorDialog(self, mode)
        self.wait_window(dlg)
        if dlg.result:
            self.fields[mode].append(dlg.result)
            self._refresh_fields()

    def _edit_field(self):
        mode = self._get_current_field_mode()
        tv = self.field_tvs[mode]
        sel = tv.selection()
        if not sel:
            return
        idx = tv.index(sel[0])
        flist = sorted(self.fields[mode], key=lambda x: x['col'])
        if idx >= len(flist):
            return
        dlg = FieldEditorDialog(self, mode, existing=flist[idx])
        self.wait_window(dlg)
        if dlg.result:
            orig = flist[idx]
            orig_idx = self.fields[mode].index(orig)
            self.fields[mode][orig_idx] = dlg.result
            self._refresh_fields()

    def _remove_field(self):
        mode = self._get_current_field_mode()
        tv = self.field_tvs[mode]
        sel = tv.selection()
        if not sel:
            return
        idx = tv.index(sel[0])
        flist = sorted(self.fields[mode], key=lambda x: x['col'])
        if idx < len(flist):
            self.fields[mode].remove(flist[idx])
            self._refresh_fields()

    def _move_up(self):
        mode = self._get_current_field_mode()
        tv = self.field_tvs[mode]
        sel = tv.selection()
        if not sel:
            return
        idx = tv.index(sel[0])
        flist = sorted(self.fields[mode], key=lambda x: x['col'])
        if idx > 0:
            flist[idx]['col'], flist[idx-1]['col'] = flist[idx-1]['col'], flist[idx]['col']
            self._refresh_fields()

    def _move_down(self):
        mode = self._get_current_field_mode()
        tv = self.field_tvs[mode]
        sel = tv.selection()
        if not sel:
            return
        idx = tv.index(sel[0])
        flist = sorted(self.fields[mode], key=lambda x: x['col'])
        if idx < len(flist) - 1:
            flist[idx]['col'], flist[idx+1]['col'] = flist[idx+1]['col'], flist[idx]['col']
            self._refresh_fields()

    def _reset_fields(self):
        mode = self._get_current_field_mode()
        self.fields[mode] = [dict(f) for f in DEFAULT_FIELDS[mode]]
        self._refresh_fields()

    # ── EXPORT ────────────────────────────────────────────────
    def _export(self):
        total = (len(self.parsed['chat']) + len(self.parsed['combat']) +
                 len(self.parsed['loot']))
        if total == 0:
            messagebox.showwarning("No Data", "Nothing parsed yet. Run Parse first.")
            return

        # Auto-generate filename based on what's being exported
        fname = self.fname_var.get().strip()
        if not fname:
            # Extract character name from ACTION log files
            char_name = None
            for f in self.files:
                name = f['name'].upper()
                if 'ACTION' in name:
                    # Pattern: ACTION-PDF-CHARACTERNAME-timestamp.log
                    parts = f['name'].split('-')
                    if len(parts) >= 3:
                        char_name = parts[2].title()  # Capitalize properly
                        break
            
            # Generate filename based on what's being parsed
            if self.do_combat.get() and len(self.parsed['combat']) > 0:
                # Combat export: CharacterName_Combat_timestamp
                from datetime import datetime
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                fname = f"{char_name}_Combat_{timestamp}" if char_name else f"Combat_{timestamp}"
            elif self.do_loot.get() and len(self.parsed['loot']) > 0:
                # Loot export: CharacterName_Loot_timestamp
                from datetime import datetime
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                fname = f"{char_name}_Loot_{timestamp}" if char_name else f"Loot_{timestamp}"
            elif self.do_chat.get() and len(self.parsed['chat']) > 0:
                # Chat export: CharacterName_Chat_timestamp
                from datetime import datetime
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                fname = f"{char_name}_Chat_{timestamp}" if char_name else f"Chat_{timestamp}"
            else:
                fname = 'game_log_export'
        
        path  = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            initialdir=self.last_save_dir,
            initialfile=fname + '.xlsx',
            filetypes=[('Excel', '*.xlsx'), ('All', '*.*')])
        if not path:
            return
        
        # Save the directory for next time
        self.last_save_dir = os.path.dirname(path)

        try:
            # Handle master database append if enabled
            if self.use_master.get() and self.master_path_var.get() and self.parsed['loot']:
                self._append_to_master()
            
            # Check if we should export Combat and Loot as separate files
            if (self.separate_action_exports.get() and 
                self.do_combat.get() and self.do_loot.get() and
                len(self.parsed['combat']) > 0 and len(self.parsed['loot']) > 0):
                # Export as two separate files
                self._export_separate_action_files(path)
                return
            
            # Normal single-file export
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            if self.exp_summary.get():
                ws = wb.create_sheet("Summary")
                self._write_summary(ws)

            # Export chat - create a separate sheet for each chat file
            if self.do_chat.get() and self.parsed['chat_files']:
                for filename, chat_data in self.parsed['chat_files'].items():
                    # Create sheet name from filename (remove extension, truncate if needed)
                    sheet_name = os.path.splitext(filename)[0]
                    # Excel sheet names max 31 chars
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:28] + "..."
                    ws = wb.create_sheet(sheet_name)
                    write_sheet(ws, chat_data, self.fields['chat'], 'chat')

            # Export combat if checkbox is checked and data exists
            if self.do_combat.get() and self.parsed['combat']:
                ws = wb.create_sheet("Combat")
                write_sheet(ws, self.parsed['combat'], self.fields['combat'], 'combat')

            # Export loot if checkbox is checked and data exists
            if self.do_loot.get() and self.parsed['loot']:
                ws = wb.create_sheet("Loot")
                write_sheet(ws, self.parsed['loot'], self.fields['loot'], 'loot')

            wb.save(path)
            self.export_status.config(
                text=f"✓ Saved to {os.path.basename(path)}")
            messagebox.showinfo("Exported",
                f"File saved!\n{path}\n\nSheets written:\n"
                f"  Chat:   {len(self.parsed['chat'])} rows\n"
                f"  Combat: {len(self.parsed['combat'])} rows\n"
                f"  Loot:   {len(self.parsed['loot'])} rows")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _export_separate_action_files(self, base_path):
        """Export Combat and Loot as two separate files"""
        from datetime import datetime
        
        # Get character name for filenames
        char_name = None
        for f in self.files:
            name = f['name'].upper()
            if 'ACTION' in name:
                parts = f['name'].split('-')
                if len(parts) >= 3:
                    char_name = parts[2].title()
                    break
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_dir = os.path.dirname(base_path)
        
        # Create Combat file
        combat_filename = f"{char_name}_Combat_{timestamp}.xlsx" if char_name else f"Combat_{timestamp}.xlsx"
        combat_path = os.path.join(base_dir, combat_filename)
        
        wb_combat = openpyxl.Workbook()
        wb_combat.remove(wb_combat.active)
        if self.exp_summary.get():
            ws = wb_combat.create_sheet("Summary")
            self._write_summary(ws)
        ws_combat = wb_combat.create_sheet("Combat")
        write_sheet(ws_combat, self.parsed['combat'], self.fields['combat'], 'combat')
        wb_combat.save(combat_path)
        
        # Create Loot file
        loot_filename = f"{char_name}_Loot_{timestamp}.xlsx" if char_name else f"Loot_{timestamp}.xlsx"
        loot_path = os.path.join(base_dir, loot_filename)
        
        wb_loot = openpyxl.Workbook()
        wb_loot.remove(wb_loot.active)
        if self.exp_summary.get():
            ws = wb_loot.create_sheet("Summary")
            self._write_summary(ws)
        ws_loot = wb_loot.create_sheet("Loot")
        write_sheet(ws_loot, self.parsed['loot'], self.fields['loot'], 'loot')
        wb_loot.save(loot_path)
        
        self.export_status.config(text=f"✓ Saved 2 files: {combat_filename}, {loot_filename}")
        messagebox.showinfo("Exported",
            f"Files saved as separate exports!\n\n"
            f"Combat: {combat_path}\n"
            f"  {len(self.parsed['combat'])} rows\n\n"
            f"Loot: {loot_path}\n"
            f"  {len(self.parsed['loot'])} rows")
    
    def _write_summary(self, ws):
        ws.title = "Summary"
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        hdr = Font(name='Arial', bold=True, size=11)
        ws['A1'] = "Gaming Log Parser — Export Summary"
        ws['A1'].font = Font(name='Arial', bold=True, size=13)
        ws['A3'] = "Generated:";  ws['B3'] = now
        ws['A4'] = "Files processed:"; ws['B4'] = len(self.files)
        ws['A5'] = "Chat rows:";    ws['B5'] = len(self.parsed['chat'])
        ws['A6'] = "Combat rows:";  ws['B6'] = len(self.parsed['combat'])
        ws['A7'] = "Loot rows:";    ws['B7'] = len(self.parsed['loot'])
        ws['A9'] = "Files"
        ws['A9'].font = hdr
        ws['B9'] = "Type"
        ws['B9'].font = hdr
        for i, f in enumerate(self.files, 10):
            ws.cell(row=i, column=1).value = f['name']
            ws.cell(row=i, column=2).value = f['type']
        ws.column_dimensions['A'].width = 44
        ws.column_dimensions['B'].width = 16
    
    # ── BUILD TAB ─────────────────────────────────────────────
    def _build_build_tab(self):
        # The Build tab is getting crowded (armor, weapon, realm, Defense
        # constraints...), so it's wrapped in a nested Notebook - everything
        # existing stays on the first sub-tab, unchanged, and a couple of
        # empty sub-tabs are reserved here so future controls have somewhere
        # to go without piling further onto one already-busy screen.
        #
        # Min/Max/Specific Level and the Find Optimal Build/Show All
        # Matches/Generate multiple build options controls are shared across
        # every sub-tab - they're built into shared_controls_frame further
        # below, packed after (so it stacks directly beneath) the
        # sub-notebook rather than inside any one sub-tab, so they stay
        # visible and usable no matter which sub-tab is selected. Neither
        # this notebook nor shared_controls_frame use expand=True/fill='both'
        # - each sub-tab already sizes itself to its own content, so letting
        # the notebook stretch to fill the whole tab would leave a big gap
        # before shared_controls_frame instead of it sitting right below.
        self.build_sub_notebook = ttk.Notebook(self.tab_build)
        self.build_sub_notebook.pack(fill='x')

        build_search_subtab = ttk.Frame(self.build_sub_notebook)
        self.build_sub_notebook.add(build_search_subtab, text='Basic Constraints')

        # Armor Type Constraints has grown enough (per-slot Defense and
        # Sigil controls alongside the armor type checkboxes) to warrant its
        # own sub-tab rather than taking up space on Search.
        self.build_armor_subtab = ttk.Frame(self.build_sub_notebook, padding=4)
        self.build_sub_notebook.add(self.build_armor_subtab, text='Armor Constraints')

        self.build_weapon_subtab = ttk.Frame(self.build_sub_notebook, padding=4)
        self.build_sub_notebook.add(self.build_weapon_subtab, text='Weapon Constraints')

        self.build_bank_subtab = ttk.Frame(self.build_sub_notebook, padding=8)
        self.build_sub_notebook.add(self.build_bank_subtab, text='Bank Build')

        # Armor type constraints - built now (into the Armor Constraints
        # sub-tab above) even though the rest of the Search sub-tab's
        # widgets are constructed further below; order of construction
        # doesn't need to match the tabs' left-to-right order. Wrapped in
        # its own bordered box, anchored to the top and sized to fit its
        # content, rather than left to sit in the sub-tab's full (much
        # taller) notebook viewport - keeps the border tight around what's
        # actually here instead of one big mostly-empty area. Wanted Sigils
        # (below) sits beside it in the empty space to the right, same
        # side-by-side pattern as Weapon Constraints/Melee Weapon Constraints.
        armor_top_row = ttk.Frame(self.build_armor_subtab)
        armor_top_row.pack(anchor='n', fill='x')

        armor_box = ttk.LabelFrame(armor_top_row, text="Armor Type Constraints", padding=8)
        armor_box.pack(side='left', anchor='n')

        armor_header = ttk.Frame(armor_box)
        armor_header.pack(fill='x', pady=(0,4))
        ttk.Button(armor_header, text="Set as Default",
                  command=self._save_armor_defaults).pack(side='left', padx=(0,4))
        ttk.Button(armor_header, text="Clear Default",
                  command=self._clear_armor_defaults).pack(side='left', padx=4)
        ttk.Button(armor_header, text="Clear All",
                  command=self._clear_all_armor).pack(side='left', padx=4)

        # Create checkbox storage for armor types
        self.armor_checks = {}

        # "All:" row and every per-slot row below are laid out with grid
        # in their own frame (armor_box itself already has pack-managed
        # children above - armor_header - and a single container can't mix
        # pack and grid for its direct children), sharing one set of
        # columns - 0: row label, 1: Max Lvl, 2-5: Cloth/Leather/Studded/
        # Plate, 6: Defense block, 7: Sigil block - so the Cloth/Leather/
        # Studded/Plate checkboxes (and everything after them) line up
        # vertically across every row even though "All:" has no Max Lvl
        # column of its own and only the per-slot rows have a Sigil column.
        armor_grid_frame = ttk.Frame(armor_box)
        armor_grid_frame.pack(fill='x')

        ttk.Label(armor_grid_frame, text="All:", width=12, font=('Arial', 9, 'bold')).grid(
            row=0, column=0, sticky='w', pady=2)

        self.armor_all_checks = {}
        for col, armor_type in enumerate(['cloth', 'leather', 'studded', 'plate']):
            var = tk.BooleanVar(value=False)
            self.armor_all_checks[armor_type] = var
            ttk.Checkbutton(armor_grid_frame, text=armor_type.title(),
                           variable=var,
                           command=lambda t=armor_type: self._update_all_armor(t)).grid(
                row=0, column=2 + col, sticky='w', padx=4)

        # Defense filter - opt-in constraint on the item list's Defense
        # column (much worse < worse < normal < better < much better), on
        # the same row as "All:" so it sits under the armor header buttons
        # above. In Find Optimal Build this is always a soft preference -
        # items in range are prioritized when possible, but every slot still
        # gets filled even if nothing in range is available. In Show All
        # Matches (a flat listing, no build to fill) it's a hard filter.
        defense_block = ttk.Frame(armor_grid_frame)
        defense_block.grid(row=0, column=6, sticky='w', padx=(20,0))

        self.use_defense_filter_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(defense_block, text="Defense:",
                       variable=self.use_defense_filter_var).pack(side='left', padx=(0,4))

        self.min_defense_var = tk.StringVar(value='normal')
        ttk.Combobox(defense_block, textvariable=self.min_defense_var, values=DEFENSE_LEVELS,
                    state='readonly', width=11).pack(side='left', padx=(0,4))

        ttk.Label(defense_block, text="to").pack(side='left', padx=(0,4))

        self.max_defense_var = tk.StringVar(value=DEFENSE_LEVELS[-1])
        ttk.Combobox(defense_block, textvariable=self.max_defense_var, values=DEFENSE_LEVELS,
                    state='readonly', width=11).pack(side='left', padx=(0,8))

        # Armor slots
        self.slot_defense_controls = {}
        self.slot_sigil_vars = {}
        # "Max Lvl" - a per-slot priority checkbox (not on the "All:" row).
        # Checked slots are greedily locked in with the highest-level item
        # that still carries a wanted/priority spell, before the normal
        # exact search runs for everything else - capped at 3 checked at
        # once, same grey-out-the-rest pattern as Melee Weapon Constraints'
        # Priority checkboxes (see _update_armor_maxlvl_cap).
        self.armor_maxlvl_vars = {}
        self.armor_maxlvl_checkbuttons = []
        for row, (slot, label) in enumerate([('head', 'Head'), ('cloak', 'Cloak'), ('body', 'Body'),
                           ('hands', 'Hands'), ('legs', 'Legs'), ('feet', 'Feet')], start=1):
            ttk.Label(armor_grid_frame, text=f"{label}:", width=12).grid(
                row=row, column=0, sticky='w', pady=2)

            maxlvl_var = tk.BooleanVar(value=False)
            self.armor_maxlvl_vars[slot] = maxlvl_var
            maxlvl_cb = ttk.Checkbutton(armor_grid_frame, text="Max Lvl", variable=maxlvl_var,
                                        command=self._update_armor_maxlvl_cap)
            maxlvl_cb.grid(row=row, column=1, sticky='w', padx=(0,10))
            self.armor_maxlvl_checkbuttons.append(maxlvl_cb)

            # Create checkboxes for each armor type
            self.armor_checks[slot] = {}
            for col, armor_type in enumerate(['cloth', 'leather', 'studded', 'plate']):
                var = tk.BooleanVar(value=False)
                self.armor_checks[slot][armor_type] = var
                ttk.Checkbutton(armor_grid_frame, text=armor_type.title(),
                               variable=var).grid(row=row, column=2 + col, sticky='w', padx=4)

            # Per-slot Defense - stacks with the global "All:" row Defense
            # controls above rather than replacing them: if both are active
            # for this slot, an item must satisfy both hard ranges, and earns
            # a priority point for each soft range it satisfies.
            slot_defense_block = ttk.Frame(armor_grid_frame)
            slot_defense_block.grid(row=row, column=6, sticky='w', padx=(20,0))

            use_var = tk.BooleanVar(value=False)
            min_var = tk.StringVar(value='normal')
            max_var = tk.StringVar(value=DEFENSE_LEVELS[-1])
            self.slot_defense_controls[slot] = {
                'use': use_var, 'min': min_var, 'max': max_var,
            }

            ttk.Checkbutton(slot_defense_block, text="Defense:",
                           variable=use_var).pack(side='left', padx=(0,4))

            ttk.Combobox(slot_defense_block, textvariable=min_var, values=DEFENSE_LEVELS,
                        state='readonly', width=11).pack(side='left', padx=(0,4))

            ttk.Label(slot_defense_block, text="to").pack(side='left', padx=(0,4))

            ttk.Combobox(slot_defense_block, textvariable=max_var, values=DEFENSE_LEVELS,
                        state='readonly', width=11).pack(side='left', padx=(0,8))

            # Per-slot Sigil - soft preference like Defense: if a Sigil is
            # picked, the search favors an item carrying that Sigil at the
            # highest available SigilLvl (1-5) for this slot, but never
            # excludes other items, so the slot still always gets filled.
            sigil_var = tk.StringVar(value='Any')
            self.slot_sigil_vars[slot] = sigil_var
            sigil_block = ttk.Frame(armor_grid_frame)
            sigil_block.grid(row=row, column=7, sticky='w', padx=(12,0))
            ttk.Label(sigil_block, text="Sigil:").pack(side='left', padx=(0,4))
            ttk.Combobox(sigil_block, textvariable=sigil_var, values=['Any'] + SIGIL_TYPES,
                        state='readonly', width=11).pack(side='left')

        # Wanted Sigils - same "add to a list, search tries to cover each
        # one" behavior as Wanted Spells (see wanted_block above), but for
        # the armor Sigil column instead of Spell: picking one here makes an
        # otherwise spell-less armor piece (many carry only a Sigil, no
        # Spell) a valid, actively-sought candidate for head/cloak/body/
        # hands/legs/feet - the per-slot "Sigil:" dropdowns above are only a
        # soft tie-break among items that already qualify some other way, so
        # they can't do this on their own. Uses the same SIGIL_TYPES list
        # already used for armor throughout (the per-slot Sigil dropdowns,
        # Shield Constraints, Melee Weapon Constraints).
        wanted_sigil_box = ttk.LabelFrame(armor_top_row, text="Wanted Sigils", padding=8)
        wanted_sigil_box.pack(side='left', anchor='n', padx=(12,0), fill='y')

        wanted_sigil_input_frame = ttk.Frame(wanted_sigil_box)
        wanted_sigil_input_frame.pack(fill='x')
        self.wanted_sigil_var = tk.StringVar(value='')
        ttk.Combobox(wanted_sigil_input_frame, textvariable=self.wanted_sigil_var, values=SIGIL_TYPES,
                    state='readonly', width=10).pack(side='left', padx=(0,4))
        ttk.Button(wanted_sigil_input_frame, text="Add to List",
                  command=self._add_wanted_sigil).pack(side='left')

        wanted_sigil_scroll_frame = ttk.Frame(wanted_sigil_box)
        wanted_sigil_scroll_frame.pack(fill='both', expand=True, pady=(6,4))

        self.wanted_sigils_data = []
        self.wanted_sigils_text = tk.Text(wanted_sigil_scroll_frame, height=4, width=22, wrap='word',
                                          cursor='arrow', state='disabled')
        wanted_sigil_scroll = ttk.Scrollbar(wanted_sigil_scroll_frame, orient='vertical',
                                            command=self.wanted_sigils_text.yview)
        self.wanted_sigils_text.configure(yscrollcommand=wanted_sigil_scroll.set)
        self.wanted_sigils_text.pack(side='left', fill='both', expand=True)
        wanted_sigil_scroll.pack(side='right', fill='y')

        ttk.Button(wanted_sigil_box, text="Clear All",
                  command=self._clear_wanted_sigils).pack(anchor='w')

        # Load saved armor defaults
        self._load_armor_defaults()

        # Weapon Constraints - its own sub-tab, same reasoning as Armor
        # Type Constraints above: built now, wrapped in a bordered box sized
        # to its own content rather than the sub-tab's full height. Sits
        # side by side with Melee Weapon Constraints in the empty space
        # to the right, rather than stacked full-width.
        weapon_constraints_row = ttk.Frame(self.build_weapon_subtab)
        weapon_constraints_row.pack(anchor='n', fill='x')

        weapon_box = ttk.LabelFrame(weapon_constraints_row, text="Weapon Constraints", padding=8)
        weapon_box.pack(side='left', anchor='n')

        # The old global Weapon Style (Melee/Direct (Caster)/Parry Staff/Any)
        # has been removed - Style is now per-combo (see Two-Handed and
        # 1h/Shield below), and more combos will get their own Style dropdown
        # over time rather than sharing one global control.

        # Weapon Types/Combo's - each pairs a specific weapon shape with its
        # own damage-type dropdown(s) (where it has one), rather than sharing
        # the one global Damage Type row below. Dual-Wield 1h needs two
        # physical weapon slots (main and off hand), each independently
        # typed; 1h/Shield and 2h/Shield each need one weapon slot plus a
        # shield, with their own type too. Plain "Weapon" and "Shield" were
        # removed here - every weapon combination now goes through one of
        # these named combos instead.
        combo_header = ttk.Label(weapon_box, text="Weapon Types/Combo's", font=('Arial', 9, 'bold'))
        combo_header.pack(anchor='w', pady=(6,2))

        dual_wield_frame = ttk.Frame(weapon_box)
        dual_wield_frame.pack(fill='x', pady=2)
        self.dual_wield_1h_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(dual_wield_frame, text="Dual-Wield 1h",
                       variable=self.dual_wield_1h_var, width=14).pack(side='left')
        ttk.Label(dual_wield_frame, text="Main:").pack(side='left', padx=(8,4))
        self.dual_wield_1h_main_var = tk.StringVar(value='Any')
        ttk.Combobox(dual_wield_frame, textvariable=self.dual_wield_1h_main_var,
                    values=WEAPON_DAMAGE_TYPES, state='readonly', width=10).pack(side='left')
        ttk.Label(dual_wield_frame, text="Off-Hand:").pack(side='left', padx=(8,4))
        self.dual_wield_1h_off_var = tk.StringVar(value='Any')
        ttk.Combobox(dual_wield_frame, textvariable=self.dual_wield_1h_off_var,
                    values=WEAPON_DAMAGE_TYPES, state='readonly', width=10).pack(side='left')

        combo_1h_shield_frame = ttk.Frame(weapon_box)
        combo_1h_shield_frame.pack(fill='x', pady=2)
        self.combo_1h_shield_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(combo_1h_shield_frame, text="1h/Shield",
                       variable=self.combo_1h_shield_var, width=14).pack(side='left')
        ttk.Label(combo_1h_shield_frame, text="Style:").pack(side='left', padx=(8,4))
        self.combo_1h_shield_style_var = tk.StringVar(value='Melee')
        ttk.Combobox(combo_1h_shield_frame, textvariable=self.combo_1h_shield_style_var,
                    values=['Melee', 'Direct'], state='readonly', width=8).pack(side='left')
        ttk.Label(combo_1h_shield_frame, text="Damage Type:").pack(side='left', padx=(8,4))
        self.combo_1h_shield_damage_var = tk.StringVar(value='Any')
        ttk.Combobox(combo_1h_shield_frame, textvariable=self.combo_1h_shield_damage_var,
                    values=WEAPON_DAMAGE_TYPES, state='readonly', width=10).pack(side='left')

        combo_2h_shield_frame = ttk.Frame(weapon_box)
        combo_2h_shield_frame.pack(fill='x', pady=2)
        self.combo_2h_shield_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(combo_2h_shield_frame, text="2h/Shield",
                       variable=self.combo_2h_shield_var, width=14).pack(side='left')
        ttk.Label(combo_2h_shield_frame, text="Damage Type:").pack(side='left', padx=(8,4))
        self.combo_2h_shield_damage_var = tk.StringVar(value='Any')
        ttk.Combobox(combo_2h_shield_frame, textvariable=self.combo_2h_shield_damage_var,
                    values=WEAPON_DAMAGE_TYPES, state='readonly', width=10).pack(side='left')

        # Fired 1h and Shield - "Fired" is its own weapon shape in the item
        # list (ranged/thrown), with no slash/thrust/crush sub-variant, so
        # unlike the combos above it needs no damage-type dropdown.
        combo_fired_1h_shield_frame = ttk.Frame(weapon_box)
        combo_fired_1h_shield_frame.pack(fill='x', pady=2)
        self.combo_fired_1h_shield_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(combo_fired_1h_shield_frame, text="Fired 1h/Shield",
                       variable=self.combo_fired_1h_shield_var).pack(side='left')

        # Two-Handed - moved in from the old Build Config row, now with its
        # own damage-type dropdown like the combos above, plus a Style
        # dropdown (Melee/Direct/Parry/Fired) - a first step toward per-combo
        # style selection; the global Weapon Style row above still governs
        # everything else for now, but is expected to go away eventually as
        # more combos get their own Style dropdown like this one. Fired
        # ("fired 2h" in the item list) has no damage-type sub-variant, so
        # its Damage Type dropdown greys out when Fired is selected.
        two_handed_frame = ttk.Frame(weapon_box)
        two_handed_frame.pack(fill='x', pady=2)
        self.two_handed_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(two_handed_frame, text="Two-Handed",
                       variable=self.two_handed_var, width=14).pack(side='left')
        ttk.Label(two_handed_frame, text="Style:").pack(side='left', padx=(8,4))
        self.two_handed_style_var = tk.StringVar(value='Melee')
        ttk.Combobox(two_handed_frame, textvariable=self.two_handed_style_var,
                    values=['Melee', 'Direct', 'Parry', 'Fired'], state='readonly', width=8).pack(side='left')
        ttk.Label(two_handed_frame, text="Damage Type:").pack(side='left', padx=(8,4))
        self.two_handed_damage_var = tk.StringVar(value='Any')
        self.two_handed_damage_combo = ttk.Combobox(two_handed_frame, textvariable=self.two_handed_damage_var,
                    values=WEAPON_DAMAGE_TYPES, state='readonly', width=10)
        self.two_handed_damage_combo.pack(side='left')
        self.two_handed_style_var.trace_add('write', lambda *args: self._update_two_handed_damage_state())

        # 1 Claw / 2 Claw - moved in from the old Build Config row; no
        # damage-type dropdown for either, unlike Two-Handed above. Each
        # claw slot gets its own Sigil preference instead: the 1st dropdown
        # covers claw_1 (used whether 1 Claw or 2 Claw is checked - it's
        # always the first claw slot filled), the 2nd covers claw_2 (only
        # meaningful when 2 Claw is checked). Same soft-preference mechanism
        # as every other Sigil dropdown - never excludes an item, just
        # favors one carrying it at the highest SigilLvl when tied otherwise.
        claw_frame = ttk.Frame(weapon_box)
        claw_frame.pack(fill='x', pady=2)
        self.claw_1_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(claw_frame, text="1 Claw", variable=self.claw_1_var).pack(side='left', padx=(0,4))
        self.claw_2_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(claw_frame, text="2 Claw", variable=self.claw_2_var).pack(side='left', padx=(0,12))

        ttk.Label(claw_frame, text="Sigil (1st):").pack(side='left', padx=(0,4))
        self.claw_1_sigil_var = tk.StringVar(value='Any')
        ttk.Combobox(claw_frame, textvariable=self.claw_1_sigil_var, values=['Any'] + SIGIL_TYPES,
                    state='readonly', width=10).pack(side='left', padx=(0,12))

        ttk.Label(claw_frame, text="Sigil (2nd):").pack(side='left', padx=(0,4))
        self.claw_2_sigil_var = tk.StringVar(value='Any')
        ttk.Combobox(claw_frame, textvariable=self.claw_2_sigil_var, values=['Any'] + SIGIL_TYPES,
                    state='readonly', width=10).pack(side='left')

        # The old global Damage Type checkboxes (Slashing/Thrusting/Crushing)
        # have been removed - each combo above now has its own Damage Type
        # dropdown instead (Claw intentionally has none).

        ttk.Label(weapon_box, text="Fired direct not factored in at this time",
                 font=('Arial', 8, 'italic'), foreground='#666').pack(anchor='w', pady=(6,0))

        # Right column next to Weapon Constraints, holding Melee Weapon
        # Constraints on top and Shield Constraints stacked below it.
        right_column = ttk.Frame(weapon_constraints_row)
        right_column.pack(side='left', anchor='n', padx=(20,0))

        # Melee Weapon Constraints - soft preferences that apply to any
        # weapon-style pick (Melee/Direct/Parry/Fired all use these, despite
        # the section's name). These never exclude a weapon outright - the
        # search always matches as many as it can and still fills the slot
        # with the best available item even if none match. "Priority" bumps
        # a constraint above the unchecked ones when scoring candidates;
        # capped at 3 checked at once - the remaining unchecked boxes grey
        # out once 3 are checked (see _update_melee_priority_cap).
        melee_box_label = ttk.Frame(right_column)
        ttk.Label(melee_box_label, text="Melee Weapon Constraints").pack(side='left')
        ttk.Label(melee_box_label, text="  (Applies to Direct, Parry and Fired weapons as well)",
                 font=('Arial', 8, 'italic'), foreground='#666').pack(side='left')
        melee_box = ttk.LabelFrame(right_column, labelwidget=melee_box_label, padding=8)
        melee_box.pack(anchor='n', fill='x')

        self.melee_damage_var = tk.StringVar(value='Any')
        self.melee_timer_var = tk.StringVar(value='Any')
        self.melee_fumble_var = tk.StringVar(value='Any')
        self.melee_accuracy_var = tk.StringVar(value='Any')
        self.melee_sigil_var = tk.StringVar(value='Any')

        self.melee_damage_priority_var = tk.BooleanVar(value=False)
        self.melee_timer_priority_var = tk.BooleanVar(value=False)
        self.melee_fumble_priority_var = tk.BooleanVar(value=False)
        self.melee_accuracy_priority_var = tk.BooleanVar(value=False)
        self.melee_sigil_priority_var = tk.BooleanVar(value=False)

        # Keyed collection used only to enforce the "at most 3 Priority
        # checked" cap - not otherwise part of the per-field wiring below.
        self.melee_priority_vars = [
            self.melee_damage_priority_var, self.melee_timer_priority_var,
            self.melee_fumble_priority_var, self.melee_accuracy_priority_var,
            self.melee_sigil_priority_var,
        ]
        self.melee_priority_checkbuttons = []

        for label, var, values, priority_var in [
                ('Damage:', self.melee_damage_var, MELEE_DAMAGE_LEVELS, self.melee_damage_priority_var),
                ('Timer:', self.melee_timer_var, MELEE_TIMER_VALUES, self.melee_timer_priority_var),
                ('Fumble:', self.melee_fumble_var, MELEE_FUMBLE_VALUES, self.melee_fumble_priority_var),
                ('Accuracy:', self.melee_accuracy_var, MELEE_ACCURACY_VALUES, self.melee_accuracy_priority_var),
                ('Sigil:', self.melee_sigil_var, MELEE_SIGIL_VALUES, self.melee_sigil_priority_var)]:
            row = ttk.Frame(melee_box)
            row.pack(fill='x', pady=2)
            ttk.Label(row, text=label, width=10).pack(side='left')
            ttk.Combobox(row, textvariable=var, values=values, state='readonly', width=14).pack(side='left')
            priority_cb = ttk.Checkbutton(row, text="Priority", variable=priority_var,
                                          command=self._update_melee_priority_cap)
            priority_cb.pack(side='left', padx=(8,0))
            self.melee_priority_checkbuttons.append(priority_cb)

        # Weight - a min/max range on the item list's Weight column, same
        # "applies to any weapon-style pick, including claws" scope as the
        # rest of this box. Soft preference by default (favors in-range
        # weapons, folded into _melee_constraint_score's normal_matches, but
        # never excludes one), same as Damage/Timer/Fumble/Accuracy/Sigil
        # above - checking "Hard Filter" instead excludes out-of-range
        # weapons outright, like Min/Max Level does for level.
        weight_row = ttk.Frame(melee_box)
        weight_row.pack(fill='x', pady=2)
        ttk.Label(weight_row, text="Weight:", width=10).pack(side='left')
        self.weapon_weight_min_var = tk.StringVar(value='')
        ttk.Entry(weight_row, textvariable=self.weapon_weight_min_var, width=6).pack(side='left')
        ttk.Label(weight_row, text=" to ").pack(side='left')
        self.weapon_weight_max_var = tk.StringVar(value='')
        ttk.Entry(weight_row, textvariable=self.weapon_weight_max_var, width=6).pack(side='left')
        self.weapon_weight_hard_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(weight_row, text="Hard Filter", variable=self.weapon_weight_hard_var).pack(
            side='left', padx=(8,0))

        # Shield Constraints - Defense works exactly like Armor Constraints'
        # per-slot Defense (soft preference, fed into the same
        # _defense_priority_score machinery, just as a single-value range
        # instead of a user-picked Min/Max); Sigil works like every other
        # Sigil dropdown (_sigil_priority_score). Neither excludes a shield
        # outright - same "match as many as possible" philosophy as above.
        # Shields are technically armor (hence the Cloth/Leather/Studded/
        # Plate checkboxes below, matching Armor Constraints) but are built
        # alongside a weapon, which is why they live here instead - unlike
        # Defense/Sigil, armor type IS a hard filter, same as it is in
        # Armor Constraints (one or more checked = must match one of them;
        # none checked = any type).
        shield_box = ttk.LabelFrame(right_column, text="Shield Constraints", padding=8)
        shield_box.pack(anchor='n', fill='x', pady=(10,0))

        self.shield_defense_var = tk.StringVar(value='Any')
        self.shield_sigil_var = tk.StringVar(value='Any')
        self.shield_armor_checks = {}

        # Cloth/Leather share the Sigil row, Studded/Plate share the Defense
        # row - rows packed tight (pady=1) against each other.
        sigil_row = ttk.Frame(shield_box)
        sigil_row.pack(fill='x', pady=1)
        ttk.Label(sigil_row, text="Sigil:", width=10).pack(side='left')
        ttk.Combobox(sigil_row, textvariable=self.shield_sigil_var, values=SHIELD_SIGIL_VALUES,
                    state='readonly', width=14).pack(side='left')
        # First checkbox in each row (Cloth/Studded) gets a fixed width so
        # the second one (Leather/Plate) starts at the same x position in
        # both rows, regardless of "Cloth" vs "Studded" being different
        # lengths - otherwise Leather and Plate wouldn't line up vertically.
        first_col_width = 9
        for armor_type in ('cloth', 'leather'):
            var = tk.BooleanVar(value=False)
            self.shield_armor_checks[armor_type] = var
            kwargs = {'width': first_col_width} if armor_type == 'cloth' else {}
            ttk.Checkbutton(sigil_row, text=armor_type.title(), variable=var, **kwargs).pack(
                side='left', padx=(12,0))

        defense_row = ttk.Frame(shield_box)
        defense_row.pack(fill='x', pady=1)
        ttk.Label(defense_row, text="Defense:", width=10).pack(side='left')
        ttk.Combobox(defense_row, textvariable=self.shield_defense_var, values=SHIELD_DEFENSE_LEVELS,
                    state='readonly', width=14).pack(side='left')
        for armor_type in ('studded', 'plate'):
            var = tk.BooleanVar(value=False)
            self.shield_armor_checks[armor_type] = var
            kwargs = {'width': first_col_width} if armor_type == 'studded' else {}
            ttk.Checkbutton(defense_row, text=armor_type.title(), variable=var, **kwargs).pack(
                side='left', padx=(12,0))

        # Load weapon defaults
        self._load_weapon_defaults()

        # Now that Armor Type Constraints has its own sub-tab, the Search
        # sub-tab's remaining content fits without needing to scroll, so it's
        # packed directly (anchored to the top) instead of stretching a
        # scrollable canvas across the sub-tab's full height and leaving a
        # big empty area below the last control.
        outer = build_search_subtab
        t = ttk.Frame(outer)
        t.pack(fill='x', anchor='n')

        ttk.Label(t, text="Build Creator",
                  font=('Arial', 13, 'bold')).pack(anchor='w', pady=(0,10))
        
        # Master file selection
        file_frame = ttk.LabelFrame(t, text="Master Database File", padding=10)
        file_frame.pack(fill='x', pady=(0,12))
        
        ttk.Label(file_frame, text="File:").pack(side='left', padx=(0,8))
        self.search_master_path = tk.StringVar(value='')
        ttk.Entry(file_frame, textvariable=self.search_master_path, 
                 width=40, state='readonly').pack(side='left', padx=4)
        ttk.Button(file_frame, text="Browse...", 
                  command=self._browse_search_master).pack(side='left', padx=2)
        ttk.Button(file_frame, text="Create New", 
                  command=self._create_new_search_master).pack(side='left', padx=2)
        ttk.Button(file_frame, text="Use Community List", 
                  command=self._use_community_list).pack(side='left', padx=2)
        ttk.Button(file_frame, text="Load", 
                  command=self._load_master_for_search).pack(side='left', padx=2)
        
        # Build constraints
        constraints_frame = ttk.LabelFrame(t, text="Build Constraints", padding=10)
        constraints_frame.pack(fill='x', pady=(0,12))
        
        # Desired spells (left) and Realm filter (right, in the space next to
        # the spell dropdowns) share a row so the realm checkboxes aren't stuck
        # way down below the weapon constraints in unused whitespace.
        spell_and_realm_frame = ttk.Frame(constraints_frame)
        spell_and_realm_frame.pack(fill='x')

        spell_block = ttk.Frame(spell_and_realm_frame)
        spell_block.pack(side='left', anchor='n')

        # Desired spells - one dropdown row per spell category, each paired with a tier dropdown
        self.category_spell_vars = {}
        self.category_tier_vars = {}
        self.category_spell_combos = {}
        self.category_tier_combos = {}

        for category, spells in SPELL_CATEGORIES.items():
            spell_input_frame = ttk.Frame(spell_block)
            spell_input_frame.pack(fill='x', pady=2)
            ttk.Label(spell_input_frame, text=f"{category}:", width=12).pack(side='left')

            spell_var = tk.StringVar(value='')
            self.category_spell_vars[category] = spell_var
            spell_combo = ttk.Combobox(spell_input_frame, textvariable=spell_var,
                                       values=spells, state='readonly', width=22)
            spell_combo.pack(side='left', padx=4)
            self.category_spell_combos[category] = spell_combo

            tier_var = tk.StringVar(value=SPELL_TIERS[0])
            self.category_tier_vars[category] = tier_var
            tier_combo = ttk.Combobox(spell_input_frame, textvariable=tier_var,
                        values=SPELL_TIERS, state='readonly', width=6)
            tier_combo.pack(side='left', padx=4)
            self.category_tier_combos[category] = tier_combo

            spell_var.trace_add('write', lambda *args, c=category: self._update_tier_options(c))

            ttk.Button(spell_input_frame, text="Add to List",
                      command=lambda c=category: self._add_categorized_spell(c)).pack(side='left', padx=4)

            if not spells:
                ttk.Label(spell_input_frame, text="(list coming soon)",
                         font=('Arial', 8, 'italic'), foreground='#888').pack(side='left', padx=4)

        # Manual entry - fallback for any spell not covered by the category
        # dropdowns above.
        manual_frame = ttk.Frame(spell_block)
        manual_frame.pack(fill='x', pady=2)
        ttk.Label(manual_frame, text="Manual:", width=12).pack(side='left')
        self.manual_spell_var = tk.StringVar(value='')
        ttk.Entry(manual_frame, textvariable=self.manual_spell_var, width=22).pack(side='left', padx=4)
        ttk.Button(manual_frame, text="Add to List",
                  command=self._add_manual_spell).pack(side='left', padx=4)

        # Wanted Spells (narrower, left) and Required Items (right, next to it)
        # share a row instead of each spanning the full width stacked vertically.
        wanted_and_required_frame = ttk.Frame(constraints_frame)
        wanted_and_required_frame.pack(fill='x', pady=4)

        # Wanted Spells - chips flow left-to-right and wrap to new lines. This
        # block expands to take the extra row width, making it the wider of
        # the two (Required Items stays a fixed, narrower width beside it).
        wanted_block = ttk.Frame(wanted_and_required_frame)
        wanted_block.pack(side='left', anchor='n', fill='both', expand=True)
        ttk.Label(wanted_block, text="Wanted Spells:").pack(anchor='w')

        spell_scroll_frame = ttk.Frame(wanted_block)
        spell_scroll_frame.pack(fill='both', expand=True)

        self.wanted_spells_data = []
        self.spell_chips_text = tk.Text(spell_scroll_frame, height=4, width=45, wrap='word',
                                        cursor='arrow', state='disabled')
        spell_scroll = ttk.Scrollbar(spell_scroll_frame, orient='vertical',
                                    command=self.spell_chips_text.yview)
        self.spell_chips_text.configure(yscrollcommand=spell_scroll.set)
        self.spell_chips_text.pack(side='left', fill='both', expand=True)
        spell_scroll.pack(side='right', fill='y')

        ttk.Button(wanted_block, text="Clear All",
                  command=self._clear_spell_list).pack(anchor='w', pady=(2,0))

        # Required Items - force specific gear into the build, calculating the
        # rest of the build around it. Tolerates spelling errors via fuzzy match.
        # Label + chip box come first (matching Wanted Spells' layout so both
        # chip areas line up in the same row); the entry to add a new one sits
        # below the chips instead of above them.
        required_block = ttk.Frame(wanted_and_required_frame)
        required_block.pack(side='left', anchor='n', padx=(20,0))

        ttk.Label(required_block, text="Required Items:").pack(anchor='w')
        required_scroll_frame = ttk.Frame(required_block)
        required_scroll_frame.pack(fill='both', expand=True)

        self.required_items = []
        self.required_items_text = tk.Text(required_scroll_frame, height=4, width=25, wrap='word',
                                           cursor='arrow', state='disabled')
        required_scroll = ttk.Scrollbar(required_scroll_frame, orient='vertical',
                                       command=self.required_items_text.yview)
        self.required_items_text.configure(yscrollcommand=required_scroll.set)
        self.required_items_text.pack(side='left', fill='both', expand=True)
        required_scroll.pack(side='right', fill='y')

        required_input_frame = ttk.Frame(required_block)
        required_input_frame.pack(fill='x', pady=(4,0))
        ttk.Label(required_input_frame, text="Require Item:").pack(side='left')
        self.required_item_var = tk.StringVar(value='')
        required_entry = ttk.Entry(required_input_frame, textvariable=self.required_item_var, width=26)
        required_entry.pack(side='left', padx=4)
        required_entry.bind('<Return>', lambda e: self._add_required_item())
        ttk.Button(required_input_frame, text="Add to Build",
                  command=self._add_required_item).pack(side='left', padx=4)
        ttk.Button(required_input_frame, text="Clear All",
                  command=self._clear_required_items).pack(side='left', padx=4)

        ttk.Label(required_block, text="(e.g. a specific weapon/armor piece - typos are OK, it'll offer close matches)",
                 font=('Arial', 8, 'italic'), foreground='#666').pack(anchor='w', pady=(2,0))

        # Min/Max/Specific Level moved out to shared_controls_frame below
        # (see the comment at the top of this method) - Min/Max Tier stays
        # here since it's specific to this tab's spell/tier matching.
        level_frame = ttk.Frame(constraints_frame)
        level_frame.pack(fill='x', pady=(8,4))

        # Min/Max Tier - bounds how far the search will fall back when the
        # exact tier requested for a spell isn't available (blank = unbounded).
        ttk.Label(level_frame, text="Min Tier:", width=12).pack(side='left')
        self.min_tier_var = tk.StringVar(value='')
        ttk.Combobox(level_frame, textvariable=self.min_tier_var, values=['', 'i', 'ii', 'iii'],
                    state='readonly', width=5).pack(side='left', padx=(0,8))

        ttk.Label(level_frame, text="Max Tier:").pack(side='left', padx=(0,4))
        self.max_tier_var = tk.StringVar(value='')
        ttk.Combobox(level_frame, textvariable=self.max_tier_var, values=['', 'i', 'ii', 'iii'],
                    state='readonly', width=5).pack(side='left', padx=(0,4))

        ttk.Label(level_frame, text="(leave blank for no restriction)",
                 font=('Arial', 8, 'italic'), foreground='#666').pack(side='left', padx=4)

        # Priority Spell - sits in the existing gap between the spell dropdowns
        # and "Only Found In", using space that's already there rather than
        # growing the row. Only offers spells currently in Wanted Spells or
        # Required Items, shown by their short base name (no .i/.ii/.iii tier
        # suffix) to save space. Multiple can be added; the viewing area below
        # is the same width as the dropdown and tall enough to reach down to
        # the Protects row of the spell category dropdowns beside it.
        priority_block = ttk.Frame(spell_and_realm_frame)
        priority_block.pack(side='left', anchor='n', padx=(20, 0))
        ttk.Label(priority_block, text="Priority Spell:").pack(anchor='w')

        priority_pick_frame = ttk.Frame(priority_block)
        priority_pick_frame.pack(anchor='w', pady=(2,0))
        self.priority_spell_var = tk.StringVar(value='(none)')
        self.priority_spell_combo = ttk.Combobox(priority_pick_frame, textvariable=self.priority_spell_var,
                                                 values=['(none)'], state='readonly', width=10)
        self.priority_spell_combo.pack(side='left')
        ttk.Button(priority_pick_frame, text="+", width=2,
                  command=self._add_priority_spell).pack(side='left', padx=(2,0))

        self.priority_spells_data = []
        self.priority_spells_text = tk.Text(priority_block, height=7, width=20, wrap='word',
                                            cursor='arrow', state='disabled')
        self.priority_spells_text.pack(anchor='w', pady=(4,0))

        # Priority Tier - a second, parallel priority box that pairs a spell
        # WITH a specific tier (unlike Priority Spell, which ignores tier
        # entirely). When a spell+tier pair is added here, the search targets
        # that tier for that spell specifically - even beating a higher tier
        # that's available - only falling back to other tiers if none of the
        # priority tiers are available for that spell.
        priority_tier_block = ttk.Frame(spell_and_realm_frame)
        priority_tier_block.pack(side='left', anchor='n', padx=(20, 0))
        ttk.Label(priority_tier_block, text="Priority Tier:").pack(anchor='w')

        priority_tier_pick_frame = ttk.Frame(priority_tier_block)
        priority_tier_pick_frame.pack(anchor='w', pady=(2,0))
        self.priority_tier_spell_var = tk.StringVar(value='(none)')
        self.priority_tier_spell_combo = ttk.Combobox(priority_tier_pick_frame, textvariable=self.priority_tier_spell_var,
                                                      values=['(none)'], state='readonly', width=10)
        self.priority_tier_spell_combo.pack(side='left')
        self.priority_tier_var = tk.StringVar(value='i')
        self.priority_tier_combo = ttk.Combobox(priority_tier_pick_frame, textvariable=self.priority_tier_var,
                                                values=['i', 'ii', 'iii'], state='readonly', width=5)
        self.priority_tier_combo.pack(side='left', padx=(2,0))
        # Narrow the tier options to whatever's valid for the selected spell,
        # same as the category dropdowns (e.g. Protects -> minor/normal/improved)
        self.priority_tier_spell_var.trace_add('write', lambda *args: self._update_priority_tier_options())
        ttk.Button(priority_tier_pick_frame, text="+", width=2,
                  command=self._add_priority_tier).pack(side='left', padx=(2,0))

        self.priority_tiers_data = []
        self.priority_tiers_text = tk.Text(priority_tier_block, height=7, width=20, wrap='word',
                                           cursor='arrow', state='disabled')
        self.priority_tiers_text.pack(anchor='w', pady=(4,0))

        # Realm filter (Only Found In) - placed in the empty space to the right
        # of the spell dropdowns, in the row frame set up alongside spell_block.
        realm_block = ttk.LabelFrame(spell_and_realm_frame, text="Only Found In", padding=8)
        realm_block.pack(side='left', anchor='n', padx=(20, 0))

        self.realm_filters = {}
        # Left two columns, in row-major order (2 per row): Evil/Glory Bea,
        # Good/Event, Chaos/Crafted. Glory Bea sits where Kaid used to be
        # (Kaid moved out into its own column to the right); Chaos took
        # Crafted's old spot, and Crafted moved below Event.
        realm_options = ['Evil', 'Glory Bea', 'Good', 'Event', 'Chaos', 'Crafted']
        cols = 2
        for i, realm in enumerate(realm_options):
            var = tk.BooleanVar(value=False)
            self.realm_filters[realm] = var
            ttk.Checkbutton(realm_block, text=realm, variable=var).grid(
                row=i // cols, column=i % cols, sticky='w', padx=4, pady=2)

        # Kaid gets its own column: "Kaid All" (renamed from the old plain
        # "Kaid" checkbox - same match key/behavior, matches every Kaid.*
        # realm) plus 4 new checkboxes for each specific Kaid color realm.
        # These two groups are mutually exclusive (grey out the other side),
        # matching an item against "kaid" broadly and against one specific
        # color realm at once doesn't make sense together.
        kaid_all_var = tk.BooleanVar(value=False)
        self.realm_filters['Kaid'] = kaid_all_var
        self.kaid_all_checkbutton = ttk.Checkbutton(realm_block, text="Kaid All",
                                                     variable=kaid_all_var,
                                                     command=self._update_kaid_exclusivity)
        self.kaid_all_checkbutton.grid(row=0, column=2, sticky='w', padx=(16,4), pady=2)

        self.kaid_color_checkbuttons = []
        for i, color in enumerate(['Kaid White', 'Kaid Green', 'Kaid Red', 'Kaid Purple'], start=1):
            var = tk.BooleanVar(value=False)
            self.realm_filters[color] = var
            cb = ttk.Checkbutton(realm_block, text=color, variable=var,
                                 command=self._update_kaid_exclusivity)
            cb.grid(row=i, column=2, sticky='w', padx=(16,4), pady=2)
            self.kaid_color_checkbuttons.append(cb)
        
        # Find Optimal Build/Show All Matches/Generate multiple build options
        # moved out to shared_controls_frame below (see the comment at the
        # top of this method).

        # Status
        self.search_status = ttk.Label(t, text="Load a master database file and add desired spells to begin", 
                                      foreground='#666')
        self.search_status.pack(anchor='w', pady=4, padx=4)
        
        # Store loaded data
        self.master_data = []
        self._bank_owned_keys = None

        # Bank Build - paste a bank/inventory listing and build the best
        # gear combo either strictly from those items, or from the whole
        # master database with owned items favored - see _find_bank_build.
        # Every other constraint (Wanted Spells, Armor/Weapon Constraints,
        # Min/Max Level, etc.) still applies exactly as normal; this only
        # adds ownership as an extra dimension on top.
        bank_tab = self.build_bank_subtab
        ttk.Label(bank_tab, text="Paste a bank/inventory listing below, then click "
                 "\"Find Best Bank Build\".", foreground='#666').pack(anchor='w', pady=(0,6))

        bank_text_frame = ttk.Frame(bank_tab)
        bank_text_frame.pack(fill='both', expand=True)
        self.bank_paste_text = tk.Text(bank_text_frame, height=14, wrap='none')
        bank_vsb = ttk.Scrollbar(bank_text_frame, orient='vertical', command=self.bank_paste_text.yview)
        bank_hsb = ttk.Scrollbar(bank_text_frame, orient='horizontal', command=self.bank_paste_text.xview)
        self.bank_paste_text.configure(yscrollcommand=bank_vsb.set, xscrollcommand=bank_hsb.set)
        self.bank_paste_text.grid(row=0, column=0, sticky='nsew')
        bank_vsb.grid(row=0, column=1, sticky='ns')
        bank_hsb.grid(row=1, column=0, sticky='ew')
        bank_text_frame.rowconfigure(0, weight=1)
        bank_text_frame.columnconfigure(0, weight=1)

        # Prioritize and Only are mutually exclusive modes for the same
        # underlying choice (see _find_bank_build) - checking either one
        # disables (and clears) the other, rather than letting both be
        # checked at once with ambiguous meaning.
        bank_checks_frame = ttk.Frame(bank_tab)
        bank_checks_frame.pack(fill='x', pady=(8,0))

        self.bank_prioritize_var = tk.BooleanVar(value=False)
        self.bank_prioritize_checkbox = ttk.Checkbutton(bank_checks_frame,
            text="Prioritize items I own (search everything, but favor owned items over ones you don't have)",
            variable=self.bank_prioritize_var, command=self._on_bank_prioritize_toggle)
        self.bank_prioritize_checkbox.pack(anchor='w')

        self.bank_only_var = tk.BooleanVar(value=True)
        self.bank_only_checkbox = ttk.Checkbutton(bank_checks_frame,
            text="Only Items I own (build only from the pasted list, nothing else)",
            variable=self.bank_only_var, command=self._on_bank_only_toggle)
        self.bank_only_checkbox.pack(anchor='w', pady=(2,0))
        # Only starts checked, so Prioritize starts disabled to match.
        self.bank_prioritize_checkbox.config(state='disabled')

        bank_controls_frame = ttk.Frame(bank_tab)
        bank_controls_frame.pack(fill='x', pady=(8,0))
        ttk.Button(bank_controls_frame, text="🏦 Find Best Bank Build",
                  command=self._find_bank_build).pack(side='left', padx=(0,4))
        ttk.Button(bank_controls_frame, text="Clear",
                  command=self._clear_bank_paste).pack(side='left', padx=4)

        self.bank_status = ttk.Label(bank_tab, text=(
            "\"Only\": only pasted items are considered - the best build achievable from what you own right now. "
            "\"Prioritize\": searches everything, just favoring items you own when otherwise close."),
            foreground='#666', wraplength=760, justify='left')
        self.bank_status.pack(anchor='w', pady=(6,0))

        # Shared controls - Min/Max/Specific Level and the search buttons -
        # packed under the sub-notebook rather than inside any one sub-tab,
        # so they stay visible and usable no matter which of Basic
        # Constraints/Armor Constraints/Weapon Constraints is selected.
        # Packed with the default side='top' (not 'bottom') so it stacks
        # directly beneath the notebook's actual content, rather than being
        # pinned to the bottom of the whole tab with a gap in between.
        shared_controls_frame = ttk.Frame(self.tab_build, padding=(0,8,0,0))
        shared_controls_frame.pack(fill='x')

        shared_level_frame = ttk.Frame(shared_controls_frame)
        shared_level_frame.pack(fill='x')

        # Centered within shared_level_frame - pack()'s default anchor is
        # 'center', so this inner frame (not fill='x') sits in the middle
        # of the row instead of hugging the left edge.
        level_fields_frame = ttk.Frame(shared_level_frame)
        level_fields_frame.pack()

        ttk.Label(level_fields_frame, text="Min Level:", width=12).pack(side='left')
        self.min_level_var = tk.StringVar(value='')
        self.min_level_entry = ttk.Entry(level_fields_frame, textvariable=self.min_level_var, width=8)
        self.min_level_entry.pack(side='left', padx=(0,8))
        self.min_level_var.trace_add('write', lambda *args: self._update_level_fields())

        ttk.Label(level_fields_frame, text="Max Level:").pack(side='left', padx=(8,4))
        self.max_level_var = tk.StringVar(value='')
        self.max_level_entry = ttk.Entry(level_fields_frame, textvariable=self.max_level_var, width=8)
        self.max_level_entry.pack(side='left', padx=(0,8))
        self.max_level_var.trace_add('write', lambda *args: self._update_level_fields())

        ttk.Label(level_fields_frame, text="Specific Level:").pack(side='left', padx=(8,4))
        self.specific_level_var = tk.StringVar(value='')
        self.specific_level_entry = ttk.Entry(level_fields_frame, textvariable=self.specific_level_var, width=8)
        self.specific_level_entry.pack(side='left', padx=(0,4))
        self.specific_level_var.trace_add('write', lambda *args: self._update_level_fields())

        # Specific Level fallback policy - only meaningful (and only
        # enabled) when Specific Level has a value. Governs what Find
        # Optimal Build does for a slot when nothing matches the wanted
        # spell at exactly that level (and, for spells requested at an
        # explicit tier, at exactly that tier too) - see _find_optimal_build
        # for how each policy relaxes level/tier and picks the highest
        # remaining match rather than just the first one found.
        specific_fallback_frame = ttk.Frame(shared_controls_frame)
        specific_fallback_frame.pack(pady=(4,0))

        ttk.Label(specific_fallback_frame, text="If no match at Specific Level:").pack(side='left', padx=(0,8))
        self.specific_level_fallback_var = tk.StringVar(value='none')
        self.specific_level_fallback_radios = []
        for label, value in [("Go down a tier", 'tier'), ("Go down in level", 'level'),
                             ("Both", 'both'), ("Don't populate slot", 'none')]:
            rb = ttk.Radiobutton(specific_fallback_frame, text=label, value=value,
                                 variable=self.specific_level_fallback_var)
            rb.pack(side='left', padx=4)
            self.specific_level_fallback_radios.append(rb)

        shared_button_frame = ttk.Frame(shared_controls_frame)
        shared_button_frame.pack(pady=8)

        ttk.Button(shared_button_frame, text="🎯 Find Optimal Build",
                  command=self._find_optimal_build, width=20).pack(side='left', padx=4)
        ttk.Button(shared_button_frame, text="📋 Show All Matches",
                  command=self._show_all_matches, width=20).pack(side='left', padx=4)

        self.generate_multi_builds_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(shared_button_frame, text="🎲 Generate multiple build options",
                       variable=self.generate_multi_builds_var).pack(side='left', padx=(12,4))

    # ── RESULTS TAB ───────────────────────────────────────────────
    def _build_results_tab(self):
        t = self.tab_results
        
        # Header with display mode and export button
        header_frame = ttk.Frame(t)
        header_frame.pack(fill='x', pady=(0,10))
        
        ttk.Label(header_frame, text="Build Search Results",
                 font=('Arial', 13, 'bold')).pack(side='left')
        
        # Display mode radio buttons (center)
        mode_frame = ttk.Frame(header_frame)
        mode_frame.pack(side='left', padx=40)
        ttk.Label(mode_frame, text="Display:").pack(side='left', padx=(0,8))
        self.results_display_mode = tk.StringVar(value='optimal')
        ttk.Radiobutton(mode_frame, text="🎯 Best Per Slot", 
                       variable=self.results_display_mode, value='optimal',
                       command=self._refresh_results_display).pack(side='left', padx=4)
        ttk.Radiobutton(mode_frame, text="📋 All Matches",
                       variable=self.results_display_mode, value='all',
                       command=self._refresh_results_display).pack(side='left', padx=4)

        # Build variant selector - only meaningful when "Generate multiple build
        # options" produced more than one alternate full build.
        ttk.Label(mode_frame, text="Build Variant:").pack(side='left', padx=(16,4))
        self.build_variant_var = tk.StringVar(value='Build 1')
        self.build_variant_combo = ttk.Combobox(mode_frame, textvariable=self.build_variant_var,
                                                values=['Build 1'], state='disabled', width=12)
        self.build_variant_combo.pack(side='left', padx=4)
        self.build_variant_combo.bind('<<ComboboxSelected>>', self._switch_build_variant)
        
        ttk.Button(header_frame, text="📤 Export As...",
                  command=self._export_build_results).pack(side='right')
        ttk.Button(header_frame, text="📌 Save Build",
                  command=self._save_current_results).pack(side='right', padx=(0,6))
        
        # Results treeview
        results_frame = ttk.LabelFrame(t, text="Suggested Build", padding=10)
        results_frame.pack(fill='both', expand=True)
        
        cols = ('Bank', 'Slot', 'Item', 'Type', 'Spell', 'Sigil', 'Level', 'Mob', 'Area', 'Div', 'Alt Options')
        col_widths = {'Slot': 55, 'Item': 200, 'Type': 60, 'Spell': 100, 'Sigil': 60, 'Level': 45,
                     'Mob': 120, 'Area': 120, 'Alt Options': 280}
        self.search_results_tv = ttk.Treeview(results_frame, columns=cols,
                                             show='headings', height=20)
        for col in cols:
            if col == 'Div':
                self.search_results_tv.heading(col, text='')
                self.search_results_tv.column(col, width=14, stretch=False, anchor='center')
            elif col == 'Bank':
                # Marks a row whose item came from a Bank Build paste (see
                # _build_dict_to_rows) - a plain emoji "icon" cell rather
                # than an image, same "icon-as-text" pattern already used
                # throughout this app's buttons/labels. Fixed-width like
                # Div, not autosized, since it only ever holds one glyph.
                self.search_results_tv.heading(col, text='📦', anchor='center')
                self.search_results_tv.column(col, width=28, stretch=False, anchor='center')
            else:
                heading_anchor = 'w' if col == 'Alt Options' else 'center'
                self.search_results_tv.heading(col, text=col, anchor=heading_anchor)
                # stretch=False so each column keeps the width _autosize_results_columns
                # computes - otherwise ttk redistributes/shrinks columns to fit the
                # visible area, undoing the sizing and making text unreadable.
                self.search_results_tv.column(col, width=col_widths[col], stretch=False)

        vsb = ttk.Scrollbar(results_frame, orient='vertical',
                           command=self.search_results_tv.yview)
        hsb = ttk.Scrollbar(results_frame, orient='horizontal',
                           command=self.search_results_tv.xview)
        self.search_results_tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.search_results_tv.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        results_frame.rowconfigure(0, weight=1)
        results_frame.columnconfigure(0, weight=1)
        
        # Filter controls at bottom
        filter_frame = ttk.Frame(t)
        filter_frame.pack(fill='x', pady=(8,0))
        
        # Remove by Area filter
        ttk.Label(filter_frame, text="Remove Area:").pack(side='left', padx=(0,4))
        self.remove_area_var = tk.StringVar(value='')
        ttk.Entry(filter_frame, textvariable=self.remove_area_var, width=20).pack(side='left', padx=4)
        ttk.Button(filter_frame, text="Remove Items", 
                  command=self._remove_items_by_area).pack(side='left', padx=4)
        ttk.Label(filter_frame, text="(e.g. 'Easter 2023', 'Olympics')", 
                 font=('Arial', 8, 'italic'), foreground='#666').pack(side='left', padx=4)
        
        # Store the last search results for toggling
        self.last_optimal_results = []
        self.last_all_results = []
        self.optimal_build = {}
        self.slot_alternates = {}
        self.build_variants = []
        self.items_by_slot = {}

    # ── SAVED BUILDS TAB ──────────────────────────────────────
    def _build_saved_builds_tab(self):
        """One permanent tab that holds every saved build as its own
        (renamable, removable) panel, rather than spawning a new tab per save."""
        outer = self.tab_saved
        ttk.Label(outer, text="Saved Builds",
                  font=('Arial', 13, 'bold')).pack(anchor='w', pady=(0,10))

        canvas = tk.Canvas(outer, highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        self.saved_builds_frame = ttk.Frame(canvas)
        saved_canvas_window = canvas.create_window((0, 0), window=self.saved_builds_frame, anchor='nw')

        def _on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox('all'))
        self.saved_builds_frame.bind('<Configure>', _on_frame_configure)

        def _on_canvas_configure(event):
            canvas.itemconfig(saved_canvas_window, width=event.width)
        canvas.bind('<Configure>', _on_canvas_configure)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')
        canvas.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>', _on_mousewheel))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))

        # list of {'name': tk.StringVar, 'headers': [...], 'rows': [...]} -
        # restored from the config file so saved builds survive closing and
        # reopening the program, not just held in memory for the session.
        self.saved_builds = [
            {'name': tk.StringVar(value=save.get('name', '')),
             'headers': save.get('headers', []),
             'rows': [tuple(row) for row in save.get('rows', [])]}
            for save in getattr(self, '_persisted_saved_builds', [])
        ]
        self.saved_build_counter = getattr(self, '_persisted_saved_build_counter', 0)
        self._render_saved_builds()

    def _save_current_results(self):
        """Add the currently displayed results as a new panel in the one
        Saved Builds tab."""
        if not self.search_results_tv.get_children():
            messagebox.showwarning("No Results", "No build results to save. Run a search first.")
            return

        headers = self.search_results_tv['columns']
        rows = []
        for iid in self.search_results_tv.get_children():
            values = self.search_results_tv.item(iid)['values']
            if str(values[0]).startswith('█'):
                # Divider row marking the end of the topmost build variant -
                # only that first (best) build gets saved, not every
                # stacked alternate "Generate multiple build options" produced.
                break
            rows.append(values)

        counter = getattr(self, 'saved_build_counter', 0) + 1
        self.saved_build_counter = counter
        self.saved_builds.append({
            'name': tk.StringVar(value=f"Save {counter}"),
            'headers': headers,
            'rows': rows,
        })
        self._render_saved_builds()
        self.notebook.select(self.tab_saved)
        self._save_config()

    def _remove_saved_build(self, index):
        if 0 <= index < len(self.saved_builds):
            del self.saved_builds[index]
            self._render_saved_builds()
            self._save_config()

    def _render_saved_builds(self):
        """Redraw every saved-build panel in order. Each panel has its own
        renamable name field, its own results table, and a Remove button."""
        for child in self.saved_builds_frame.winfo_children():
            child.destroy()

        if not self.saved_builds:
            ttk.Label(self.saved_builds_frame,
                     text="No builds saved yet - use \"Save Build\" from the Results tab.",
                     foreground='#666').pack(anchor='w', pady=20)
            return

        for index, save in enumerate(self.saved_builds):
            panel = ttk.LabelFrame(self.saved_builds_frame, padding=10)
            panel.pack(fill='x', pady=(0,12), padx=(0,4))

            header_frame = ttk.Frame(panel)
            header_frame.pack(fill='x', pady=(0,8))
            ttk.Label(header_frame, text="Name:").pack(side='left', padx=(0,4))
            ttk.Entry(header_frame, textvariable=save['name'], width=30).pack(side='left', padx=(0,8))
            ttk.Button(header_frame, text="✕ Remove",
                      command=lambda i=index: self._remove_saved_build(i)).pack(side='right')
            ttk.Button(header_frame, text="📤 Export As...",
                      command=lambda i=index: self._export_saved_build(i)).pack(side='right', padx=(0,6))

            tree_frame = ttk.Frame(panel)
            tree_frame.pack(fill='both', expand=True)

            cols = save['headers']
            tv = ttk.Treeview(tree_frame, columns=cols, show='headings', height=min(10, max(3, len(save['rows']))))
            for col in cols:
                src_heading = self.search_results_tv.heading(col)
                src_col = self.search_results_tv.column(col)
                if col == 'Div':
                    tv.heading(col, text='')
                else:
                    tv.heading(col, text=src_heading['text'], anchor=src_heading['anchor'])
                tv.column(col, width=src_col['width'], stretch=False)

            vsb2 = ttk.Scrollbar(tree_frame, orient='vertical', command=tv.yview)
            hsb2 = ttk.Scrollbar(tree_frame, orient='horizontal', command=tv.xview)
            tv.configure(yscrollcommand=vsb2.set, xscrollcommand=hsb2.set)
            tv.grid(row=0, column=0, sticky='nsew')
            vsb2.grid(row=0, column=1, sticky='ns')
            hsb2.grid(row=1, column=0, sticky='ew')
            tree_frame.rowconfigure(0, weight=1)
            tree_frame.columnconfigure(0, weight=1)

            for row in save['rows']:
                tv.insert('', 'end', values=row)

    # ── AREA ITEMS TAB ────────────────────────────────────────
    def _build_area_items_tab(self):
        """Pick an Area from the loaded master database and browse every
        item droppable there - a straight lookup, not a build search."""
        t = self.tab_area_items
        ttk.Label(t, text="Area Items", font=('Arial', 13, 'bold')).pack(anchor='w', pady=(0,10))

        picker_frame = ttk.Frame(t)
        picker_frame.pack(fill='x', pady=(0,8))
        ttk.Label(picker_frame, text="Area:").pack(side='left', padx=(0,4))
        self._all_area_names = []
        self.area_items_var = tk.StringVar(value='')
        self.area_items_entry = ttk.Entry(picker_frame, textvariable=self.area_items_var, width=40)
        self.area_items_entry.pack(side='left', padx=(0,8))
        self.area_items_entry.bind('<KeyRelease>', self._on_area_items_type)
        self.area_items_entry.bind('<Down>', self._area_suggest_focus_listbox)
        self.area_items_entry.bind('<Return>', self._on_area_items_enter)
        self.area_items_entry.bind('<Escape>', self._hide_area_suggestions)
        self.area_items_entry.bind('<FocusOut>', self._on_area_entry_focus_out)
        # Clicking/tabbing into the field shows the full alphabetical list
        # right away, same as opening an ordinary dropdown - not just once
        # typing starts.
        self.area_items_entry.bind('<FocusIn>', self._update_area_suggestions)
        self.area_items_entry.bind('<Button-1>', self._update_area_suggestions)

        self.area_items_status = ttk.Label(picker_frame, text="", foreground='#666')
        self.area_items_status.pack(side='left', padx=(8,0))

        # Type-ahead suggestion popup - a plain undecorated Toplevel holding
        # a Listbox, positioned right under the entry. Not ttk.Combobox's
        # own dropdown: that dropdown fights any attempt to keep it open
        # and the entry focused at the same time while typing, so a small
        # self-managed popup is the reliable way to show live suggestions.
        self._area_suggest_popup = None
        self._area_suggest_listbox = None

        tree_frame = ttk.Frame(t)
        tree_frame.pack(fill='both', expand=True)

        cols = ('Realm', 'Mob', 'Item', 'Slot', 'Type', 'Spell', 'Sigil', 'Level')
        col_widths = {'Realm': 70, 'Mob': 160, 'Item': 220, 'Slot': 55, 'Type': 110,
                     'Spell': 110, 'Sigil': 60, 'Level': 45}
        self.area_items_tv = ttk.Treeview(tree_frame, columns=cols, show='headings', height=20)
        for col in cols:
            self.area_items_tv.heading(col, text=col, anchor='center')
            self.area_items_tv.column(col, width=col_widths[col], stretch=False)

        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.area_items_tv.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.area_items_tv.xview)
        self.area_items_tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.area_items_tv.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

    def _refresh_area_items_dropdown(self):
        """Repopulate the Area Items autocomplete list from whatever's
        currently loaded in self.master_data - called whenever a master
        database finishes loading (including the silent startup auto-load)."""
        areas = sorted({(item.get('Area') or '').strip() for item in self.master_data} - {''})
        self._all_area_names = areas
        if self.area_items_var.get() not in areas:
            self.area_items_var.set('')
            self.area_items_tv.delete(*self.area_items_tv.get_children())
            self.area_items_status.config(text='')

    def _on_area_items_type(self, event):
        """Type-ahead: narrow the suggestion popup to Areas whose name
        contains what's been typed so far (case-insensitive), updated on
        every keystroke. Navigation keys are left to their own handlers so
        this doesn't fight them."""
        if event.keysym in ('Up', 'Down', 'Return', 'Escape', 'Tab'):
            return
        self._update_area_suggestions()

    def _update_area_suggestions(self, event=None):
        """Refresh the suggestion popup from whatever's currently typed -
        the full alphabetical Area list when the field is empty (e.g. on
        first click/focus, same as opening an ordinary dropdown), narrowed
        to matching Areas once something's typed. Shared by the typing,
        click, and focus-in handlers so all three show the same thing."""
        typed = self.area_items_var.get().strip().lower()
        matches = [a for a in self._all_area_names if typed in a.lower()] if typed else list(self._all_area_names)
        if matches:
            self._show_area_suggestions(matches)
        else:
            self._hide_area_suggestions()

    def _show_area_suggestions(self, matches):
        """Create (on first use) and fill the suggestion popup - a plain
        undecorated Toplevel holding a Listbox, positioned right under the
        Area entry. A real ttk.Combobox's built-in dropdown can't be kept
        open while the entry itself keeps typing focus (posting it steals
        focus, and restoring focus to the entry closes it again) - a
        self-managed popup sidesteps that fight entirely."""
        if self._area_suggest_popup is None:
            popup = tk.Toplevel(self)
            popup.withdraw()
            popup.overrideredirect(True)
            popup.attributes('-topmost', True)
            listbox = tk.Listbox(popup, height=8, activestyle='dotbox', exportselection=False)
            scrollbar = ttk.Scrollbar(popup, orient='vertical', command=listbox.yview)
            listbox.configure(yscrollcommand=scrollbar.set)
            listbox.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            listbox.bind('<Return>', self._on_area_suggest_pick)
            listbox.bind('<Double-Button-1>', self._on_area_suggest_pick)
            listbox.bind('<Escape>', self._hide_area_suggestions)
            self._area_suggest_popup = popup
            self._area_suggest_listbox = listbox

        listbox = self._area_suggest_listbox
        listbox.delete(0, tk.END)
        # Full alphabetical list when unfiltered (matches is already sorted -
        # see _refresh_area_items_dropdown), narrowed as the user types;
        # the popup itself scrolls, same as an ordinary dropdown would.
        for name in matches:
            listbox.insert(tk.END, name)

        x = self.area_items_entry.winfo_rootx()
        y = self.area_items_entry.winfo_rooty() + self.area_items_entry.winfo_height()
        width = max(self.area_items_entry.winfo_width(), 200)
        self._area_suggest_popup.geometry(f"{width}x160+{x}+{y}")
        self._area_suggest_popup.deiconify()

    def _hide_area_suggestions(self, event=None):
        if self._area_suggest_popup is not None:
            self._area_suggest_popup.withdraw()

    def _area_suggest_focus_listbox(self, event):
        """Down arrow in the entry moves focus into the suggestion list (if
        showing) so its usual arrow-key/Enter navigation takes over."""
        if self._area_suggest_popup is None or not self._area_suggest_popup.winfo_viewable():
            return
        listbox = self._area_suggest_listbox
        listbox.focus_set()
        if listbox.size() > 0:
            listbox.selection_set(0)
            listbox.activate(0)
        return 'break'

    def _on_area_suggest_pick(self, event):
        """Enter or double-click on a suggestion - use it and run the lookup."""
        listbox = self._area_suggest_listbox
        selection = listbox.curselection()
        if not selection:
            return
        value = listbox.get(selection[0])
        self.area_items_var.set(value)
        self._hide_area_suggestions()
        self.area_items_entry.focus_set()
        self._refresh_area_items_results()

    def _on_area_entry_focus_out(self, event):
        """Hide the suggestion popup once focus has settled somewhere that
        isn't the entry or the popup's own listbox - deferred slightly
        since FocusOut fires before the listbox actually gains focus when
        the user arrows down into it."""
        self.after(150, self._hide_area_suggestions_if_unfocused)

    def _hide_area_suggestions_if_unfocused(self):
        if self._area_suggest_popup is None:
            return
        focused = self.focus_get()
        if focused not in (self.area_items_entry, self._area_suggest_listbox):
            self._hide_area_suggestions()

    def _on_area_items_enter(self, event):
        """Enter in the entry itself (not the suggestion list) - if what's
        typed exactly matches an Area, run the lookup directly."""
        typed = self.area_items_var.get().strip()
        match = next((a for a in self._all_area_names if a.lower() == typed.lower()), None)
        if match:
            self.area_items_var.set(match)
            self._hide_area_suggestions()
            self._refresh_area_items_results()

    def _refresh_area_items_results(self):
        """Fill the Area Items table with every master-database item whose
        Area matches the one currently picked in the dropdown."""
        self.area_items_tv.delete(*self.area_items_tv.get_children())
        area = self.area_items_var.get().strip()
        if not area:
            self.area_items_status.config(text='')
            return

        matches = [item for item in self.master_data if (item.get('Area') or '').strip() == area]
        # Armor slots first (grouped together, not just one), then jewels,
        # then shields, then weapons. Within the armor group, Plate first,
        # then Studded, then Leather, then Cloth - then slot and Item name
        # so it doesn't read as randomly shuffled within any of those.
        slot_group = {'jewel': 1, 'shield': 2, 'weapon': 3}
        armor_type_order = {'plate': 0, 'studded': 1, 'leather': 2, 'cloth': 3}
        matches.sort(key=lambda item: (
            slot_group.get((item.get('Slot') or '').strip().lower(), 0),
            armor_type_order.get((item.get('Type') or '').strip().lower(), 4),
            (item.get('Slot') or '').strip().lower(),
            (item.get('Item') or '').strip().lower(),
        ))
        for item in matches:
            self.area_items_tv.insert('', 'end', values=(
                item.get('Realm', ''), item.get('Mob', ''), item.get('Item', ''),
                item.get('Slot', ''), item.get('Type', ''), item.get('Spell', ''),
                item.get('Sigil', ''), item.get('Level', ''),
            ))
        self.area_items_status.config(text=f"{len(matches)} item(s)")

    def _remove_items_by_area(self):
        """Remove items from results by area name"""
        area_filter = self.remove_area_var.get().strip().lower()
        if not area_filter:
            messagebox.showwarning("No Area", "Please enter an area name to remove")
            return

        area_idx = list(self.search_results_tv['columns']).index('Area')

        # Remove from both stored results
        self.last_optimal_results = [
            row for row in self.last_optimal_results
            if area_filter not in (row[area_idx] or '').lower()
        ]
        self.last_all_results = [
            row for row in self.last_all_results
            if area_filter not in (row[area_idx] or '').lower()
        ]
        
        # Refresh display
        self._refresh_results_display()
        
        # Clear the input
        self.remove_area_var.set('')
    
    def _apply_realm_filter(self):
        """Filter results by selected realms - applied during search"""
        # Realm filtering is now integrated into search functions
        # This function is just a placeholder for the checkbox command
        pass
    
    def _update_tier_options(self, category):
        """Restrict the tier dropdown based on the selected spell (e.g. Agility has no tier iii)"""
        spell = self.category_spell_vars[category].get().strip().lower()
        allowed = SPELL_TIER_RESTRICTIONS.get(spell, SPELL_TIERS)
        combo = self.category_tier_combos[category]
        combo['values'] = allowed
        if self.category_tier_vars[category].get() not in allowed:
            self.category_tier_vars[category].set(allowed[0])

    def _update_priority_tier_options(self):
        """Same as _update_tier_options, but for the Priority Tier box's own
        tier dropdown (e.g. picking a Protect there should offer
        minor/normal/improved, not the default i/ii/iii)."""
        spell = self.priority_tier_spell_var.get().strip().lower()
        allowed = [t for t in SPELL_TIER_RESTRICTIONS.get(spell, SPELL_TIERS) if t != '(any)']
        self.priority_tier_combo['values'] = allowed
        if self.priority_tier_var.get() not in allowed:
            self.priority_tier_var.set(allowed[0])

    def _add_categorized_spell(self, category):
        """Add the spell selected in a category dropdown (plus tier) to the wanted spells list"""
        spell = self.category_spell_vars[category].get().strip()
        if not spell:
            return

        combined = SPELL_VALUE_OVERRIDES.get(spell.lower(), spell.lower())
        tier = self.category_tier_vars[category].get()
        # Protects show friendlier labels (minor/normal/improved) but these
        # map directly onto the ordinary i/ii/iii suffix - no special format.
        tier = PROTECT_TIER_TO_SUFFIX.get(tier, tier)
        if tier and tier != '(any)':
            combined = f"{combined}.{tier}"

        self._append_wanted_spell(combined)

    def _add_manual_spell(self):
        """Free-text add - fallback for any spell not covered by the category dropdowns"""
        spell = self.manual_spell_var.get().strip()
        if not spell:
            return
        self._append_wanted_spell(spell.lower())
        self.manual_spell_var.set('')

    def _append_wanted_spell(self, spell_value):
        """Add a spell string to the wanted list (deduped) and re-render the chip display"""
        if spell_value not in [s.lower() for s in self.wanted_spells_data]:
            self.wanted_spells_data.append(spell_value)
            self._render_spell_chips()

    def _remove_wanted_spell(self, spell_value):
        """Remove one spell chip (called by a chip's own ✕ button)"""
        if spell_value in self.wanted_spells_data:
            self.wanted_spells_data.remove(spell_value)
            self._render_spell_chips()

    def _clear_spell_list(self):
        """Clear all spells from list"""
        self.wanted_spells_data = []
        self._render_spell_chips()

    def _render_spell_chips(self):
        """Redraw the Wanted Spells area as chips that flow horizontally and
        wrap to new lines, instead of one spell per vertical row."""
        text = self.spell_chips_text
        text.config(state='normal')
        text.delete('1.0', tk.END)
        for spell in self.wanted_spells_data:
            chip = ttk.Frame(text, relief='raised', borderwidth=1)
            ttk.Label(chip, text=spell, padding=(4, 1)).pack(side='left')
            remove_lbl = ttk.Label(chip, text='✕', padding=(4, 1),
                                   foreground='#a33', cursor='hand2')
            remove_lbl.pack(side='left')
            remove_lbl.bind('<Button-1>', lambda e, s=spell: self._remove_wanted_spell(s))
            text.window_create(tk.END, window=chip)
            text.insert(tk.END, ' ')
        text.config(state='disabled')
        self._refresh_priority_spell_options()

    def _add_wanted_sigil(self):
        """Add the sigil selected in Armor Constraints' dropdown to the wanted sigils list"""
        sigil = self.wanted_sigil_var.get().strip()
        if not sigil:
            return
        if sigil not in self.wanted_sigils_data:
            self.wanted_sigils_data.append(sigil)
            self._render_wanted_sigil_chips()

    def _remove_wanted_sigil(self, sigil):
        """Remove one sigil chip (called by a chip's own ✕ button)"""
        if sigil in self.wanted_sigils_data:
            self.wanted_sigils_data.remove(sigil)
            self._render_wanted_sigil_chips()

    def _clear_wanted_sigils(self):
        """Clear all sigils from the wanted sigils list"""
        self.wanted_sigils_data = []
        self._render_wanted_sigil_chips()

    def _render_wanted_sigil_chips(self):
        """Redraw the Wanted Sigils area as chips, same flow-and-wrap style as Wanted Spells."""
        text = self.wanted_sigils_text
        text.config(state='normal')
        text.delete('1.0', tk.END)
        for sigil in self.wanted_sigils_data:
            chip = ttk.Frame(text, relief='raised', borderwidth=1)
            ttk.Label(chip, text=sigil, padding=(4, 1)).pack(side='left')
            remove_lbl = ttk.Label(chip, text='✕', padding=(4, 1),
                                   foreground='#a33', cursor='hand2')
            remove_lbl.pack(side='left')
            remove_lbl.bind('<Button-1>', lambda e, s=sigil: self._remove_wanted_sigil(s))
            text.window_create(tk.END, window=chip)
            text.insert(tk.END, ' ')
        text.config(state='disabled')

    def _refresh_priority_spell_options(self):
        """Rebuild the Priority Spell and Priority Tier spell dropdowns from
        whatever's currently in Wanted Spells and Required Items, shown by
        base name (no tier suffix)"""
        bases = {_spell_base(s) for s in self.wanted_spells_data}
        bases.update(_spell_base(item.get('Spell') or '') for item in self.required_items)
        bases.discard('')
        values = ['(none)'] + sorted(bases)
        self.priority_spell_combo['values'] = values
        if self.priority_spell_var.get() not in values:
            self.priority_spell_var.set('(none)')
        self.priority_tier_spell_combo['values'] = values
        if self.priority_tier_spell_var.get() not in values:
            self.priority_tier_spell_var.set('(none)')

    def _add_priority_spell(self):
        """Add the spell currently picked in the dropdown to the priority list"""
        spell = self.priority_spell_var.get()
        if not spell or spell == '(none)':
            return
        if spell not in self.priority_spells_data:
            self.priority_spells_data.append(spell)
            self._render_priority_spell_chips()

    def _remove_priority_spell(self, spell):
        """Remove one priority spell chip (called by a chip's own ✕ button)"""
        if spell in self.priority_spells_data:
            self.priority_spells_data.remove(spell)
            self._render_priority_spell_chips()

    def _render_priority_spell_chips(self):
        """Redraw the priority spells viewing area as removable chips"""
        text = self.priority_spells_text
        text.config(state='normal')
        text.delete('1.0', tk.END)
        for spell in self.priority_spells_data:
            chip = ttk.Frame(text, relief='raised', borderwidth=1)
            ttk.Label(chip, text=spell, padding=(4, 1)).pack(side='left')
            remove_lbl = ttk.Label(chip, text='✕', padding=(4, 1),
                                   foreground='#a33', cursor='hand2')
            remove_lbl.pack(side='left')
            remove_lbl.bind('<Button-1>', lambda e, s=spell: self._remove_priority_spell(s))
            text.window_create(tk.END, window=chip)
            text.insert(tk.END, ' ')
        text.config(state='disabled')

    def _add_priority_tier(self):
        """Add the (spell, tier) pair currently picked to the priority tier list"""
        spell = self.priority_tier_spell_var.get()
        tier = self.priority_tier_var.get()
        if not spell or spell == '(none)' or not tier:
            return
        # Protects show friendlier labels (minor/normal/improved) but these
        # map directly onto the ordinary i/ii/iii tier rank used everywhere else.
        tier = PROTECT_TIER_TO_SUFFIX.get(tier, tier)
        pair = (spell, tier)
        if pair not in self.priority_tiers_data:
            self.priority_tiers_data.append(pair)
            self._render_priority_tier_chips()

    def _remove_priority_tier(self, pair):
        """Remove one priority tier chip (called by a chip's own ✕ button)"""
        if pair in self.priority_tiers_data:
            self.priority_tiers_data.remove(pair)
            self._render_priority_tier_chips()

    def _render_priority_tier_chips(self):
        """Redraw the priority tiers viewing area as removable chips"""
        text = self.priority_tiers_text
        text.config(state='normal')
        text.delete('1.0', tk.END)
        for pair in self.priority_tiers_data:
            spell, tier = pair
            chip = ttk.Frame(text, relief='raised', borderwidth=1)
            ttk.Label(chip, text=f"{spell} ({tier})", padding=(4, 1)).pack(side='left')
            remove_lbl = ttk.Label(chip, text='✕', padding=(4, 1),
                                   foreground='#a33', cursor='hand2')
            remove_lbl.pack(side='left')
            remove_lbl.bind('<Button-1>', lambda e, p=pair: self._remove_priority_tier(p))
            text.window_create(tk.END, window=chip)
            text.insert(tk.END, ' ')
        text.config(state='disabled')

    def _add_required_item(self):
        """Look up a specific item the user typed in the master database and
        force it into the build. Falls back to fuzzy matching (tolerating
        spelling errors) when there's no exact (case-insensitive) match."""
        query = self.required_item_var.get().strip()
        if not query:
            return
        if not self.master_data:
            messagebox.showwarning("No Data", "Load a master database first.")
            return

        exact = [it for it in self.master_data
                if (it.get('Item') or '').strip().lower() == query.lower()]
        if exact:
            self._confirm_and_add_required_item(exact[0])
            return

        all_names = sorted({(it.get('Item') or '').strip()
                            for it in self.master_data if it.get('Item')})
        matches = difflib.get_close_matches(query, all_names, n=5, cutoff=0.5)
        if not matches:
            messagebox.showinfo("No Match Found",
                f"No item found matching '{query}' (even allowing for spelling errors).")
            return

        if len(matches) == 1:
            if messagebox.askyesno("Did You Mean?",
                    f"No exact match for '{query}'.\n\nDid you mean '{matches[0]}'?"):
                item = next(it for it in self.master_data
                           if (it.get('Item') or '').strip() == matches[0])
                self._confirm_and_add_required_item(item)
            return

        dlg = ItemMatchDialog(self, query, matches)
        self.wait_window(dlg)
        if dlg.result:
            item = next(it for it in self.master_data
                       if (it.get('Item') or '').strip() == dlg.result)
            self._confirm_and_add_required_item(item)

    def _confirm_and_add_required_item(self, item):
        """Add a resolved item (exact or fuzzy-confirmed) to the required list"""
        name = (item.get('Item') or '').strip()
        if any((it.get('Item') or '').strip().lower() == name.lower() for it in self.required_items):
            messagebox.showinfo("Already Added", f"'{name}' is already in your required items list.")
            return
        self.required_items.append(item)
        self.required_item_var.set('')
        self._render_required_item_chips()

    def _remove_required_item(self, item):
        """Remove one required-item chip (called by a chip's own ✕ button)"""
        if item in self.required_items:
            self.required_items.remove(item)
            self._render_required_item_chips()

    def _clear_required_items(self):
        """Clear all required items"""
        self.required_items = []
        self._render_required_item_chips()

    def _render_required_item_chips(self):
        """Redraw the Required Items area as removable chips showing item + slot"""
        text = self.required_items_text
        text.config(state='normal')
        text.delete('1.0', tk.END)
        for item in self.required_items:
            label = f"{item.get('Item', '')} ({(item.get('Slot') or '').title()})"
            chip = ttk.Frame(text, relief='raised', borderwidth=1)
            ttk.Label(chip, text=label, padding=(4, 1)).pack(side='left')
            remove_lbl = ttk.Label(chip, text='✕', padding=(4, 1),
                                   foreground='#a33', cursor='hand2')
            remove_lbl.pack(side='left')
            remove_lbl.bind('<Button-1>', lambda e, it=item: self._remove_required_item(it))
            text.window_create(tk.END, window=chip)
            text.insert(tk.END, ' ')
        text.config(state='disabled')
        self._refresh_priority_spell_options()

    def _browse_search_master(self):
        """Browse for master database to search"""
        path = filedialog.askopenfilename(
            title='Select Master Database to Search',
            initialdir=self.last_save_dir,
            filetypes=[('Excel', '*.xlsx'), ('All', '*.*')])
        if path:
            self.search_master_path.set(path)
    
    def _set_app_icon(self):
        """Set the window/taskbar icon from the bundled swirl_icon.ico, if
        present - same "look next to the script, bundled via --add-data"
        lookup as _find_community_list_path. iconbitmap(default=...) also
        propagates to every Toplevel dialog (ItemMatchDialog, etc.), not
        just the main window. Wrapped in try/except since a missing or
        unreadable icon file shouldn't ever stop the app from starting."""
        possible_paths = [
            os.path.join(os.path.dirname(__file__), 'swirl_icon.ico'),
            'swirl_icon.ico',
            os.path.join(os.path.dirname(__file__), '..', 'swirl_icon.ico'),
        ]
        for path in possible_paths:
            if os.path.exists(path):
                try:
                    self.iconbitmap(default=os.path.abspath(path))
                except tk.TclError:
                    pass
                return

    def _find_community_list_path(self):
        """Look for the bundled community equipment list in common locations,
        returning its absolute path or None if it isn't found."""
        possible_paths = [
            # Same directory as the script
            os.path.join(os.path.dirname(__file__), 'Olmran_Community_Eq_and_Stats_List.xlsx'),
            # Current working directory
            'Olmran_Community_Eq_and_Stats_List.xlsx',
            # One level up
            os.path.join(os.path.dirname(__file__), '..', 'Olmran_Community_Eq_and_Stats_List.xlsx'),
        ]
        for path in possible_paths:
            if os.path.exists(path):
                return os.path.abspath(path)
        return None

    def _use_community_list(self):
        """Use the bundled community equipment list as master database"""
        community_file = self._find_community_list_path()
        if not community_file:
            messagebox.showerror("File Not Found",
                "Could not find 'Olmran_Community_Eq_and_Stats_List.xlsx'\n\n"
                "Please ensure the community equipment list is in the same folder as the program.")
            return

        self.search_master_path.set(community_file)
        self._load_master_for_search()

    def _auto_load_community_list(self):
        """Silently load the bundled community list at startup if present, so
        the Build tab is ready to search without an extra click. Unlike the
        manual button, this doesn't show an error if the file is missing -
        "Use Community List" is still there to retry normally."""
        community_file = self._find_community_list_path()
        if community_file:
            self.search_master_path.set(community_file)
            self._load_master_for_search(silent=True)
    
    def _create_new_search_master(self):
        """Create a new master database file from Build tab"""
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d')
        default_name = f"Master_Loot_Database_{timestamp}"
        
        path = filedialog.asksaveasfilename(
            title='Create New Master Database',
            initialdir=self.last_save_dir,
            initialfile=default_name + '.xlsx',
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx'), ('All', '*.*')])
        
        if not path:
            return
        
        try:
            # Create new Excel file with Loot sheet
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            ws = wb.create_sheet("Loot")
            
            # Write header row using loot fields
            from openpyxl.styles import Font, PatternFill, Alignment
            
            headers = []
            max_col = max(f['col'] for f in self.fields['loot'])
            headers = [''] * max_col
            for f in self.fields['loot']:
                headers[f['col'] - 1] = f['label']
            
            ws.append(headers)
            
            # Style header
            hdr_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            hdr_fill = PatternFill(fill_type='solid', start_color='4472C4', end_color='4472C4')
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(1, col_idx)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.freeze_panes = 'A2'
            
            # Save
            wb.save(path)
            
            # Set as master file and load it
            self.search_master_path.set(path)
            self.last_save_dir = os.path.dirname(path)
            
            messagebox.showinfo("Master Created", 
                f"✓ New master database created!\n\n{path}\n\n"
                "Click 'Load' to load it for searching.")
                
        except Exception as e:
            messagebox.showerror("Create Error", f"Failed to create master file:\n{str(e)}")
    
    def _load_master_for_search(self, silent=False):
        """Load master database into memory for searching. silent=True (used
        for the startup auto-load) skips the confirmation/error popups -
        the search status label still reflects what happened either way."""
        path = self.search_master_path.get()
        if not path or not os.path.exists(path):
            if not silent:
                messagebox.showwarning("No File", "Please select a master database file first")
            return

        try:
            wb = openpyxl.load_workbook(path, read_only=True)

            # Check for either 'Loot' or 'Equipment' sheet
            sheet_name = None
            if 'Loot' in wb.sheetnames:
                sheet_name = 'Loot'
            elif 'Equipment' in wb.sheetnames:
                sheet_name = 'Equipment'
            else:
                if not silent:
                    messagebox.showerror("No Equipment Sheet",
                        "Selected file has no 'Loot' or 'Equipment' sheet")
                return

            ws = wb[sheet_name]

            # Read all data
            self.master_data = []
            headers = [cell.value for cell in ws[1]]
            item_col_idx = headers.index('Item') if 'Item' in headers else None

            skipped_struck = 0
            for row in ws.iter_rows(min_row=2, values_only=False):
                # Struck-through items in the source sheet mark removed/invalid
                # entries - skip them so the search never treats them as real.
                if item_col_idx is not None:
                    item_cell = row[item_col_idx]
                    if item_cell.font and item_cell.font.strike:
                        skipped_struck += 1
                        continue

                item_dict = {}
                for idx, header in enumerate(headers):
                    if header and idx < len(row):
                        item_dict[header] = row[idx].value or ''
                if item_dict.get('Item'):  # Only add if has item name
                    self.master_data.append(item_dict)

            wb.close()

            status_text = f"✓ Loaded {len(self.master_data)} items from {os.path.basename(path)}"
            if skipped_struck:
                status_text += f" ({skipped_struck} struck-through skipped)"
            self.search_status.config(text=status_text)
            self._refresh_area_items_dropdown()
            if not silent:
                messagebox.showinfo("Loaded",
                    f"Master database loaded!\n{len(self.master_data)} items ready to search")

        except Exception as e:
            if not silent:
                messagebox.showerror("Load Error", f"Failed to load master:\n{str(e)}")
    
    def _clear_armor_defaults(self):
        """Clear saved defaults from config file"""
        self.armor_defaults = {}
        self.weapon_combo_defaults = {}
        self._save_config()
        messagebox.showinfo("Defaults Cleared", "Saved armor and weapon defaults have been cleared!")

    def _clear_all_armor(self):
        """Clear all armor and weapon checkboxes"""
        # Clear All row
        for armor_type in ['cloth', 'leather', 'studded', 'plate']:
            self.armor_all_checks[armor_type].set(False)

        # Clear all individual armor slots
        for slot in ['head', 'cloak', 'body', 'hands', 'legs', 'feet']:
            for armor_type in ['cloth', 'leather', 'studded', 'plate']:
                self.armor_checks[slot][armor_type].set(False)
            self.armor_maxlvl_vars[slot].set(False)
        self._update_armor_maxlvl_cap()

        # Clear Weapon Types/Combo's
        self.two_handed_var.set(False)
        self.two_handed_style_var.set('Melee')
        self.two_handed_damage_var.set('Any')
        self.claw_1_var.set(False)
        self.claw_2_var.set(False)
        self.claw_1_sigil_var.set('Any')
        self.claw_2_sigil_var.set('Any')
        self.dual_wield_1h_var.set(False)
        self.dual_wield_1h_main_var.set('Any')
        self.dual_wield_1h_off_var.set('Any')
        self.combo_1h_shield_var.set(False)
        self.combo_1h_shield_style_var.set('Melee')
        self.combo_1h_shield_damage_var.set('Any')
        self.combo_2h_shield_var.set(False)
        self.combo_2h_shield_damage_var.set('Any')
        self.combo_fired_1h_shield_var.set(False)

        # Clear Melee Weapon Constraints
        self.melee_damage_var.set('Any')
        self.melee_timer_var.set('Any')
        self.melee_fumble_var.set('Any')
        self.melee_accuracy_var.set('Any')
        self.melee_sigil_var.set('Any')
        self.melee_damage_priority_var.set(False)
        self.melee_timer_priority_var.set(False)
        self.melee_fumble_priority_var.set(False)
        self.melee_accuracy_priority_var.set(False)
        self.melee_sigil_priority_var.set(False)
        self.weapon_weight_min_var.set('')
        self.weapon_weight_max_var.set('')
        self.weapon_weight_hard_var.set(False)
        self._update_melee_priority_cap()

        # Clear Shield Constraints
        self.shield_defense_var.set('Any')
        self.shield_sigil_var.set('Any')
        for armor_type in ('cloth', 'leather', 'studded', 'plate'):
            self.shield_armor_checks[armor_type].set(False)

        # Clear all level filters
        self.min_level_var.set('')
        self.max_level_var.set('')
        self.specific_level_var.set('')
    
    def _update_two_handed_damage_state(self):
        """Fired 2h weapons have no damage-type sub-variant. While Style is
        Fired, the Damage Type dropdown is locked to a single "Fired" value
        (not one of the shared WEAPON_DAMAGE_TYPES, so it's never a pickable
        option for any other style) and disabled; switching to any other
        style restores the normal Any/Slashing/Thrusting/Crushing list."""
        if self.two_handed_style_var.get() == 'Fired':
            self.two_handed_damage_combo.config(values=['Fired'], state='disabled')
            self.two_handed_damage_var.set('Fired')
        else:
            self.two_handed_damage_combo.config(values=WEAPON_DAMAGE_TYPES, state='readonly')
            if self.two_handed_damage_var.get() == 'Fired':
                self.two_handed_damage_var.set('Any')

    def _update_melee_priority_cap(self):
        """At most 3 of the 5 Melee Weapon Constraints can be marked Priority
        at once - once 3 are checked, grey out the remaining unchecked boxes
        so a 4th can't be picked; un-grey them as soon as one is unchecked."""
        checked_count = sum(1 for v in self.melee_priority_vars if v.get())
        at_cap = checked_count >= 3
        for var, checkbutton in zip(self.melee_priority_vars, self.melee_priority_checkbuttons):
            if var.get():
                checkbutton.config(state='normal')
            else:
                checkbutton.config(state='disabled' if at_cap else 'normal')

    def _update_armor_maxlvl_cap(self):
        """At most 3 armor slots can be marked Max Lvl at once - same
        grey-out-the-rest pattern as _update_melee_priority_cap."""
        checked_count = sum(1 for v in self.armor_maxlvl_vars.values() if v.get())
        at_cap = checked_count >= 3
        for slot, checkbutton in zip(self.armor_maxlvl_vars.keys(), self.armor_maxlvl_checkbuttons):
            if self.armor_maxlvl_vars[slot].get():
                checkbutton.config(state='normal')
            else:
                checkbutton.config(state='disabled' if at_cap else 'normal')

    def _update_all_armor(self, armor_type):
        """When an 'All' checkbox is clicked, update all individual slots to match"""
        all_value = self.armor_all_checks[armor_type].get()
        for slot in ['head', 'cloak', 'body', 'hands', 'legs', 'feet']:
            self.armor_checks[slot][armor_type].set(all_value)
        # Note: After this, individual slots can be changed independently
        # The "All" checkbox doesn't stay synced - it's just a quick-set tool
    
    def _refresh_results_display(self):
        """Toggle between optimal and all matches view"""
        mode = self.results_display_mode.get()
        
        # Clear current display
        self.search_results_tv.delete(*self.search_results_tv.get_children())
        
        # Show appropriate results
        if mode == 'optimal':
            results = self.last_optimal_results
        else:
            results = self.last_all_results
        
        # Display results
        for row in results:
            self.search_results_tv.insert('', 'end', values=row)

        self._autosize_results_columns()

    def _autosize_results_columns(self):
        """Shrink/grow each results column to the minimum width that fits its
        header and currently displayed cell values - no wasted or cramped space."""
        tv = self.search_results_tv
        cols = tv['columns']
        font = tkfont.Font(font=ttk.Style().lookup('Treeview', 'font') or 'TkDefaultFont')
        padding = 24  # cell borders/inset allowance

        # Divider rows (between stacked build variants) fill every cell with a
        # long block-character string so it reads as a solid line - that's not
        # real content, and would otherwise force every column to the same
        # oversized width to fit it. Skip those rows when measuring.
        real_rows = [tv.item(iid)['values'] for iid in tv.get_children()
                    if not str(tv.item(iid)['values'][0]).startswith('█')]

        for idx, col in enumerate(cols):
            if col in ('Div', 'Bank'):
                continue  # fixed-width visual spacer/icon column, not content-driven
            max_width = font.measure(tv.heading(col)['text'])
            for values in real_rows:
                w = font.measure(str(values[idx]))
                if w > max_width:
                    max_width = w
            tv.column(col, width=max_width + padding)

    def _build_dict_to_rows(self, build_dict):
        """Turn a {slot: item} build mapping into result-table rows. Alt
        Options lists other items in the same slot that provide the same
        base spell (any item still within the active level/armor/weapon/realm
        constraints, not just exact scoring ties) - the spell is only shown
        for an alternate when its tier differs from the item actually picked,
        since restating the identical spell/tier is redundant."""
        slot_order = ['head', 'jewel_1', 'jewel_2', 'cloak', 'body', 'hands', 'legs', 'feet',
                     'weapon', 'weapon_off', 'shield', 'claw_1', 'claw_2']
        attempted_slots = getattr(self, 'attempted_slots', set())
        rows = []
        for slot in slot_order:
            if slot not in build_dict:
                # Shown instead of a "No Items For Some Spells" popup - a
                # slot the search actually tried to fill (per the current
                # combo checkboxes) but found nothing for still shows up
                # here, rather than silently vanishing from the table. A
                # slot that was never attempted at all (e.g. claw_2 when
                # only 1 Claw is checked) stays hidden as before.
                if slot in attempted_slots:
                    display_slot = ('jewel' if slot.startswith('jewel') else 'claw' if slot.startswith('claw')
                                    else 'off-hand' if slot == 'weapon_off' else slot)
                    rows.append((
                        '', display_slot.title(), 'No suitable item found',
                        '', '', '', '', '', '', '████████', '(none)'
                    ))
                continue
            item = build_dict[slot]
            # weapon_off (Dual-Wield 1h's off-hand slot) gets its own label -
            # unlike jewel/claw's interchangeable pair, main and off-hand
            # weapons can have different damage types, so distinguishing
            # them here is useful, not redundant.
            display_slot = ('jewel' if slot.startswith('jewel') else 'claw' if slot.startswith('claw')
                            else 'off-hand' if slot == 'weapon_off' else slot)
            lookup_slot = 'jewel' if slot.startswith('jewel') else 'claw' if slot.startswith('claw') else slot

            chosen_spell = (item.get('Spell') or '').lower()
            chosen_base = _spell_base(chosen_spell)
            chosen_tier = _item_tier_rank(chosen_spell)

            seen_names = {(item.get('Item') or '').strip().lower()}
            alts = []
            for c in self.items_by_slot.get(lookup_slot, []):
                if c is item:
                    continue
                c_spell = (c.get('Spell') or '').lower()
                if _spell_base(c_spell) != chosen_base:
                    continue
                c_name = (c.get('Item') or '').strip()
                if c_name.lower() in seen_names:
                    continue
                seen_names.add(c_name.lower())
                c_level = c.get('Level', '')
                try:
                    c_level_num = int(c_level)
                except (ValueError, TypeError):
                    c_level_num = 0
                if _item_tier_rank(c_spell) == chosen_tier:
                    text = f"{c_level} - {c_name}"
                else:
                    text = f"{c_level} - {c_name} ({c.get('Spell', '')})"
                alts.append((c_level_num, text))
            # Highest level first (leftmost), lowest last (rightmost).
            alts.sort(key=lambda pair: pair[0], reverse=True)
            alt_text = ', '.join(text for _, text in alts) if alts else '(none)'

            rows.append((
                '📦' if self._is_bank_owned(item) else '',
                display_slot.title(),
                item.get('Item', ''),
                item.get('Type', ''),
                item.get('Spell', ''),
                item.get('Sigil', ''),
                item.get('Level', ''),
                item.get('Mob', ''),
                item.get('Area', ''),
                '████████',
                alt_text
            ))
        return rows

    def _all_variants_rows(self):
        """Stack every build variant's rows into one list, in the current
        order of self.build_variants, with a thin black divider row between
        each build so they're visually separated in the same results table."""
        # Each cell's text is far longer than any column could ever be, so it
        # overflows and fills the full cell width edge-to-edge (Treeview just
        # clips excess rather than wrapping) - that reads as one continuous
        # black line across the row instead of separate blocks per column.
        num_cols = len(self.search_results_tv['columns'])
        divider = tuple('█' * 300 for _ in range(num_cols))

        rows = []
        for i, variant in enumerate(self.build_variants):
            if i > 0:
                rows.append(divider)
            rows.extend(self._build_dict_to_rows(variant))
        return rows

    def _switch_build_variant(self, event=None):
        """Move the selected build variant to the top of the stacked list"""
        label = self.build_variant_var.get()
        try:
            idx = int(label.split()[-1]) - 1
        except (ValueError, IndexError):
            idx = 0
        if not (0 <= idx < len(self.build_variants)):
            return

        selected = self.build_variants.pop(idx)
        self.build_variants.insert(0, selected)

        variant_labels = [f"Build {i + 1}" for i in range(len(self.build_variants))]
        self.build_variant_combo['values'] = variant_labels
        self.build_variant_var.set(variant_labels[0])

        self.last_optimal_results = self._all_variants_rows()
        self.results_display_mode.set('optimal')
        self._refresh_results_display()

    def _update_kaid_exclusivity(self):
        """Kaid All and the 4 Kaid color checkboxes grey out whichever side
        isn't in use - checking Kaid All disables the colors, checking any
        color disables Kaid All. Same disable-only pattern as level filters
        (see _update_level_fields) - the other side's value isn't cleared,
        just made non-interactive while it wouldn't apply."""
        kaid_all_checked = self.realm_filters['Kaid'].get()
        any_color_checked = any(self.realm_filters[c].get()
                                for c in ('Kaid White', 'Kaid Green', 'Kaid Red', 'Kaid Purple'))

        self.kaid_all_checkbutton.config(state='disabled' if any_color_checked else 'normal')
        for cb in self.kaid_color_checkbuttons:
            cb.config(state='disabled' if kaid_all_checked else 'normal')

    def _update_level_fields(self):
        """Enable/disable level fields based on which one is being used"""
        has_min = bool(self.min_level_var.get().strip())
        has_max = bool(self.max_level_var.get().strip())
        has_specific = bool(self.specific_level_var.get().strip())
        
        # If specific level is set, disable min/max
        if has_specific:
            self.min_level_entry.config(state='disabled')
            self.max_level_entry.config(state='disabled')
        # If min or max is set, disable specific
        elif has_min or has_max:
            self.specific_level_entry.config(state='disabled')
        # If nothing is set, enable all
        else:
            self.min_level_entry.config(state='normal')
            self.max_level_entry.config(state='normal')
            self.specific_level_entry.config(state='normal')

        # The fallback policy only means anything alongside Specific Level.
        for rb in self.specific_level_fallback_radios:
            rb.config(state='normal' if has_specific else 'disabled')
    
    def _save_armor_defaults(self):
        """Save current armor and weapon selections as defaults"""
        self.armor_defaults = {}
        for slot in ['head', 'cloak', 'body', 'hands', 'legs', 'feet']:
            self.armor_defaults[slot] = {}
            for armor_type in ['cloth', 'leather', 'studded', 'plate']:
                self.armor_defaults[slot][armor_type] = self.armor_checks[slot][armor_type].get()
            self.armor_defaults[slot]['maxlvl'] = self.armor_maxlvl_vars[slot].get()
        
        # Save Weapon Types/Combo's defaults
        self.weapon_combo_defaults = {
            'two_handed': self.two_handed_var.get(),
            'two_handed_style': self.two_handed_style_var.get(),
            'two_handed_damage': self.two_handed_damage_var.get(),
            'claw_1': self.claw_1_var.get(),
            'claw_2': self.claw_2_var.get(),
            'claw_1_sigil': self.claw_1_sigil_var.get(),
            'claw_2_sigil': self.claw_2_sigil_var.get(),
            'dual_wield_1h': self.dual_wield_1h_var.get(),
            'dual_wield_1h_main': self.dual_wield_1h_main_var.get(),
            'dual_wield_1h_off': self.dual_wield_1h_off_var.get(),
            'combo_1h_shield': self.combo_1h_shield_var.get(),
            'combo_1h_shield_style': self.combo_1h_shield_style_var.get(),
            'combo_1h_shield_damage': self.combo_1h_shield_damage_var.get(),
            'combo_2h_shield': self.combo_2h_shield_var.get(),
            'combo_2h_shield_damage': self.combo_2h_shield_damage_var.get(),
            'combo_fired_1h_shield': self.combo_fired_1h_shield_var.get(),
            'melee_damage': self.melee_damage_var.get(),
            'melee_timer': self.melee_timer_var.get(),
            'melee_fumble': self.melee_fumble_var.get(),
            'melee_accuracy': self.melee_accuracy_var.get(),
            'melee_sigil': self.melee_sigil_var.get(),
            'melee_damage_priority': self.melee_damage_priority_var.get(),
            'melee_timer_priority': self.melee_timer_priority_var.get(),
            'melee_fumble_priority': self.melee_fumble_priority_var.get(),
            'melee_accuracy_priority': self.melee_accuracy_priority_var.get(),
            'melee_sigil_priority': self.melee_sigil_priority_var.get(),
            'weapon_weight_min': self.weapon_weight_min_var.get(),
            'weapon_weight_max': self.weapon_weight_max_var.get(),
            'weapon_weight_hard': self.weapon_weight_hard_var.get(),
            'shield_defense': self.shield_defense_var.get(),
            'shield_sigil': self.shield_sigil_var.get(),
            'shield_armor_types': [t for t in ('cloth', 'leather', 'studded', 'plate')
                                   if self.shield_armor_checks[t].get()],
        }

        self._save_config()
        messagebox.showinfo("Defaults Saved", "Armor and weapon preferences saved as defaults!")
    
    def _load_armor_defaults(self):
        """Load saved armor defaults into checkboxes"""
        if hasattr(self, 'armor_defaults') and self.armor_defaults:
            for slot in ['head', 'cloak', 'body', 'hands', 'legs', 'feet']:
                if slot in self.armor_defaults:
                    for armor_type in ['cloth', 'leather', 'studded', 'plate']:
                        if armor_type in self.armor_defaults[slot]:
                            self.armor_checks[slot][armor_type].set(
                                self.armor_defaults[slot][armor_type])
                    self.armor_maxlvl_vars[slot].set(self.armor_defaults[slot].get('maxlvl', False))
            self._update_armor_maxlvl_cap()
    
    def _load_weapon_defaults(self):
        """Load saved weapon defaults into checkboxes"""
        # Load Weapon Types/Combo's defaults
        combos = getattr(self, 'weapon_combo_defaults', None)
        if combos:
            self.two_handed_var.set(combos.get('two_handed', False))
            self.two_handed_style_var.set(combos.get('two_handed_style', 'Melee'))
            self.two_handed_damage_var.set(combos.get('two_handed_damage', 'Any'))
            self.claw_1_var.set(combos.get('claw_1', False))
            self.claw_2_var.set(combos.get('claw_2', False))
            self.claw_1_sigil_var.set(combos.get('claw_1_sigil', 'Any'))
            self.claw_2_sigil_var.set(combos.get('claw_2_sigil', 'Any'))
            self.dual_wield_1h_var.set(combos.get('dual_wield_1h', False))
            self.dual_wield_1h_main_var.set(combos.get('dual_wield_1h_main', 'Any'))
            self.dual_wield_1h_off_var.set(combos.get('dual_wield_1h_off', 'Any'))
            self.combo_1h_shield_var.set(combos.get('combo_1h_shield', False))
            self.combo_1h_shield_style_var.set(combos.get('combo_1h_shield_style', 'Melee'))
            self.combo_1h_shield_damage_var.set(combos.get('combo_1h_shield_damage', 'Any'))
            self.combo_2h_shield_var.set(combos.get('combo_2h_shield', False))
            self.combo_2h_shield_damage_var.set(combos.get('combo_2h_shield_damage', 'Any'))
            self.combo_fired_1h_shield_var.set(combos.get('combo_fired_1h_shield', False))
            self.melee_damage_var.set(combos.get('melee_damage', 'Any'))
            self.melee_timer_var.set(combos.get('melee_timer', 'Any'))
            self.melee_fumble_var.set(combos.get('melee_fumble', 'Any'))
            self.melee_accuracy_var.set(combos.get('melee_accuracy', 'Any'))
            self.melee_sigil_var.set(combos.get('melee_sigil', 'Any'))
            self.melee_damage_priority_var.set(combos.get('melee_damage_priority', False))
            self.melee_timer_priority_var.set(combos.get('melee_timer_priority', False))
            self.melee_fumble_priority_var.set(combos.get('melee_fumble_priority', False))
            self.melee_accuracy_priority_var.set(combos.get('melee_accuracy_priority', False))
            self.melee_sigil_priority_var.set(combos.get('melee_sigil_priority', False))
            self.weapon_weight_min_var.set(combos.get('weapon_weight_min', ''))
            self.weapon_weight_max_var.set(combos.get('weapon_weight_max', ''))
            self.weapon_weight_hard_var.set(combos.get('weapon_weight_hard', False))
            self.shield_defense_var.set(combos.get('shield_defense', 'Any'))
            self.shield_sigil_var.set(combos.get('shield_sigil', 'Any'))
            shield_armor_types = combos.get('shield_armor_types', [])
            for armor_type in ('cloth', 'leather', 'studded', 'plate'):
                self.shield_armor_checks[armor_type].set(armor_type in shield_armor_types)

        # .set() doesn't fire the checkboxes' own command callback, so sync
        # the grey-out state explicitly after a bulk load.
        self._update_melee_priority_cap()

    def _get_export_data(self):
        """Collect the currently displayed results as (headers, rows), with
        the visual Div spacer column and any divider rows (between stacked
        build variants) stripped out - shared by every "Save As" format."""
        headers = ['Bank', 'Slot', 'Item', 'Type', 'Spell', 'Sigil', 'Level', 'Mob', 'Area', 'Alt Options']
        rows = []
        for item_id in self.search_results_tv.get_children():
            values = list(self.search_results_tv.item(item_id)['values'])
            if str(values[0]).startswith('█'):
                continue  # divider row between stacked build variants
            del values[9]  # drop the Div spacer column
            rows.append(values)
        return headers, rows

    def _export_build_results(self):
        """Save the currently displayed build search results as a spreadsheet,
        HTML page, image, or formatted text document."""
        if not self.search_results_tv.get_children():
            messagebox.showwarning("No Results", "No build results to export. Run a search first.")
            return
        headers, rows = self._get_export_data()
        self._export_data_as(headers, rows, default_name="Build_Results", build_name="Build Results")

    def _export_saved_build(self, index):
        """Save one specific Saved Builds panel as a spreadsheet, HTML page,
        image, or formatted text document."""
        if not (0 <= index < len(self.saved_builds)):
            return
        save = self.saved_builds[index]
        headers = [h for h in save['headers'] if h != 'Div']
        div_idx = list(save['headers']).index('Div')
        rows = [[v for i, v in enumerate(row) if i != div_idx] for row in save['rows']]
        build_name = save['name'].get().strip() or "Build Results"
        self._export_data_as(headers, rows, default_name=build_name, build_name=build_name)

    def _export_data_as(self, headers, rows, default_name, build_name=None):
        """Shared "ask for a file, save in whichever format was picked" logic
        used by both the live Results tab and each Saved Builds panel.
        build_name is shown as a title at the top of the exported document
        itself (not just baked into the suggested filename below), so an
        exported file is still identifiable after being renamed, printed, or
        forwarded on its own."""
        build_name = build_name or default_name
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        path = filedialog.asksaveasfilename(
            title='Export Build Results',
            initialdir=self.last_save_dir,
            initialfile=f"{default_name}_{timestamp}.xlsx",
            defaultextension='.xlsx',
            filetypes=[
                ('Excel Spreadsheet', '*.xlsx'),
                ('HTML Page', '*.html'),
                ('Image (PNG)', '*.png'),
                ('Text Document', '*.txt'),
                ('All Files', '*.*'),
            ])

        if not path:
            return

        ext = os.path.splitext(path)[1].lower()

        try:
            if ext == '.xlsx':
                self._save_results_xlsx(path, headers, rows, build_name)
            elif ext in ('.html', '.htm'):
                self._save_results_html(path, headers, rows, build_name)
            elif ext == '.png':
                self._save_results_image(path, headers, rows, build_name)
            elif ext == '.txt':
                self._save_results_text(path, headers, rows, build_name)
            else:
                messagebox.showerror("Unknown Format",
                    f"Unrecognized file extension '{ext}'. Use .xlsx, .html, .png, or .txt.")
                return

            self.last_save_dir = os.path.dirname(path)
            messagebox.showinfo("Exported",
                f"Build results exported!\n\n{len(rows)} items exported to:\n{path}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export results:\n{str(e)}")

    def _save_results_xlsx(self, path, headers, rows, build_name):
        """Save as an Excel spreadsheet, with build_name as a title row above the headers"""
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (re.sub(r'[\[\]:*?/\\]', '_', build_name).strip()[:31]) or "Build Results"

        ws.append([build_name])
        title_cell = ws.cell(1, 1)
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        if len(headers) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        ws.append(headers)
        hdr_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        hdr_fill = PatternFill(fill_type='solid', start_color='4472C4', end_color='4472C4')
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(2, col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in rows:
            ws.append(row)

        # Indexed by column number rather than col[0].column_letter - the
        # title row's merged cells (columns 2+) are MergedCell placeholders
        # there, which don't carry a column_letter of their own.
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = max((len(str(c.value)) for c in col if c.value is not None), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        ws.freeze_panes = 'A3'
        wb.save(path)

    def _save_results_html(self, path, headers, rows, build_name):
        """Save as a styled standalone HTML page, with build_name as the page title/heading"""
        import html as html_module

        escaped_name = html_module.escape(str(build_name))
        parts = [
            '<!DOCTYPE html><html><head><meta charset="utf-8">',
            f'<title>{escaped_name}</title><style>',
            'body{font-family:Arial,sans-serif;background:#f4f4f4;padding:24px;}',
            'table{border-collapse:collapse;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.15);}',
            'th,td{border:1px solid #ccc;padding:6px 12px;text-align:left;font-size:13px;white-space:nowrap;}',
            'th{background:#4472C4;color:#fff;}',
            'tr:nth-child(even){background:#f7f7f7;}',
            '</style></head><body>',
            f'<h2>{escaped_name}</h2><table><tr>',
        ]
        parts.append(''.join(f'<th>{html_module.escape(str(h))}</th>' for h in headers))
        parts.append('</tr>')
        for row in rows:
            parts.append('<tr>' + ''.join(f'<td>{html_module.escape(str(v))}</td>' for v in row) + '</tr>')
        parts.append('</table></body></html>')

        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(parts))

    def _save_results_image(self, path, headers, rows, build_name):
        """Save as a rendered PNG table image (drawn directly, not a screen
        capture, so it doesn't depend on the window being visible/unobstructed),
        with build_name as a title band above the table"""
        from PIL import Image, ImageDraw, ImageFont

        try:
            font = ImageFont.truetype("arial.ttf", 14)
            font_bold = ImageFont.truetype("arialbd.ttf", 14)
            font_title = ImageFont.truetype("arialbd.ttf", 20)
        except Exception:
            font = ImageFont.load_default()
            font_bold = font
            font_title = font

        padding = 10
        row_height = 26
        title_height = 40
        measurer = ImageDraw.Draw(Image.new('RGB', (10, 10)))
        col_widths = []
        for i, header in enumerate(headers):
            width = measurer.textlength(str(header), font=font_bold)
            for row in rows:
                width = max(width, measurer.textlength(str(row[i]), font=font))
            col_widths.append(int(width) + padding * 2)

        total_width = max(sum(col_widths), int(measurer.textlength(str(build_name), font=font_title)) + padding * 2)
        total_height = title_height + row_height * (len(rows) + 1)

        img = Image.new('RGB', (total_width, total_height), 'white')
        draw = ImageDraw.Draw(img)

        draw.text((padding, (title_height - 20) // 2), str(build_name), fill='black', font=font_title)
        draw.line([(0, title_height), (total_width, title_height)], fill=(200, 200, 200))

        draw.rectangle([0, title_height, total_width, title_height + row_height], fill=(68, 114, 196))
        x = 0
        for i, header in enumerate(headers):
            draw.text((x + padding, title_height + (row_height - 14) // 2), str(header), fill='white', font=font_bold)
            x += col_widths[i]

        y = title_height + row_height
        for r_idx, row in enumerate(rows):
            if r_idx % 2 == 1:
                draw.rectangle([0, y, total_width, y + row_height], fill=(245, 245, 245))
            x = 0
            for i, value in enumerate(row):
                draw.text((x + padding, y + (row_height - 14) // 2), str(value), fill='black', font=font)
                x += col_widths[i]
            y += row_height

        x = 0
        for w in col_widths:
            x += w
            draw.line([(x, title_height), (x, total_height)], fill=(200, 200, 200))
        y = title_height
        for _ in range(len(rows) + 2):
            draw.line([(0, y), (total_width, y)], fill=(200, 200, 200))
            y += row_height

        img.save(path)

    def _save_results_text(self, path, headers, rows, build_name):
        """Save as a plain text document with aligned, fixed-width columns
        so the formatting/structure survives in any text editor, with
        build_name as the first line."""
        col_widths = []
        for i, header in enumerate(headers):
            width = len(str(header))
            for row in rows:
                width = max(width, len(str(row[i])))
            col_widths.append(width)

        def format_row(values):
            return '  '.join(str(v).ljust(col_widths[i]) for i, v in enumerate(values))

        lines = [str(build_name), ""]
        header_line = format_row(headers)
        lines.append(header_line)
        lines.append('-' * len(header_line))
        for row in rows:
            lines.append(format_row(row))

        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

    def _assign_required_items(self, build, covered_bases, wanted_bases):
        """Seed the build dict with user-required items, one per available
        slot (jewel/claw items fill whichever of their two slots is still
        open). Returns (crafted_count, used_claw_items) so the rest of the
        search's Crafted-cap and no-duplicate-claw rules stay consistent."""
        crafted_count = 0
        used_claw_items = []

        for item in self.required_items:
            item_slot = (item.get('Slot') or '').strip().lower()
            item_type = (item.get('Type') or '').strip().lower()
            # Claws are stored as Slot=weapon/Type=claw in the source data -
            # there is no Slot='claw' value anywhere in the bundled list.
            is_claw_item = item_slot == 'weapon' and 'claw' in item_type
            if item_slot == 'jewel':
                target = 'jewel_1' if 'jewel_1' not in build else ('jewel_2' if 'jewel_2' not in build else None)
            elif is_claw_item:
                target = 'claw_1' if 'claw_1' not in build else ('claw_2' if 'claw_2' not in build else None)
            else:
                target = item_slot if item_slot and item_slot not in build else None

            if not target:
                continue  # no free matching slot left for this item

            build[target] = item

            item_spell = (item.get('Spell') or '').lower()
            for base in wanted_bases:
                if base in item_spell:
                    covered_bases.add(base)

            if 'crafted' in (item.get('Realm') or '').strip().lower():
                crafted_count += 1
            if is_claw_item:
                used_claw_items.append(item)

        return crafted_count, used_claw_items

    def _clear_bank_paste(self):
        """Clear the Bank Build paste box and its status line"""
        self.bank_paste_text.delete('1.0', tk.END)
        self.bank_status.config(text=(
            "\"Only\": only pasted items are considered - the best build achievable from what you own right now. "
            "\"Prioritize\": searches everything, just favoring items you own when otherwise close."))

    def _on_bank_only_toggle(self):
        """Only Items I own and Prioritize items I own are mutually
        exclusive - checking one clears and disables the other."""
        if self.bank_only_var.get():
            self.bank_prioritize_var.set(False)
            self.bank_prioritize_checkbox.config(state='disabled')
        else:
            self.bank_prioritize_checkbox.config(state='normal')

    def _on_bank_prioritize_toggle(self):
        """See _on_bank_only_toggle - same mutual exclusion, other direction."""
        if self.bank_prioritize_var.get():
            self.bank_only_var.set(False)
            self.bank_only_checkbox.config(state='disabled')
        else:
            self.bank_only_checkbox.config(state='normal')

    def _find_bank_build(self):
        """Parse the pasted bank/inventory listing, then run Find Optimal
        Build either restricted to just those items ("Prioritize" off) or
        against the full master database with owned items favored
        ("Prioritize" on) - see W_BANK_OWNED/_is_bank_owned in
        _find_optimal_build for how the latter is scored."""
        if not self.master_data:
            messagebox.showwarning("No Data", "Please load a master database first")
            return

        text = self.bank_paste_text.get('1.0', tk.END)
        owned_keys, recognized = _parse_bank_paste_text(text)
        if not owned_keys:
            messagebox.showwarning("No Items",
                "No equippable items were recognized in the pasted text. Expected lines like:\n"
                "12.) a glowing wispweave hood of fallen snow [60|Head|CP2|evadeenhance2]")
            return

        matched_items = [item for item in self.master_data if _bank_item_key(item) in owned_keys]
        matched_key_count = len({_bank_item_key(item) for item in matched_items})
        unmatched_count = len(owned_keys) - matched_key_count

        # Set for both modes, not just Prioritize - _is_bank_owned (and so
        # the Results tab's Bank column icon) needs this regardless of
        # which mode actually ran, including Only mode where every result
        # is trivially owned by definition.
        self._bank_owned_keys = owned_keys
        try:
            if self.bank_prioritize_var.get():
                self._find_optimal_build()
                mode_text = "searched everything, prioritizing owned items"
            else:
                if not matched_items:
                    messagebox.showwarning("No Matches",
                        "None of the pasted items were recognized against the loaded master database.")
                    return
                original_master_data = self.master_data
                self.master_data = matched_items
                try:
                    self._find_optimal_build()
                finally:
                    self.master_data = original_master_data
                mode_text = "restricted to owned items only"
        finally:
            self._bank_owned_keys = None

        self.bank_status.config(
            text=f"Parsed {recognized} equippable item(s) from the paste - {matched_key_count} recognized in the "
                 f"master database ({unmatched_count} not found/unmatched) - {mode_text}.")

    def _is_bank_owned(self, item):
        """True if this item matches something from the currently active
        Bank Build paste - self._bank_owned_keys is only set (non-None)
        while a Bank Build "Prioritize" search is actually running (see
        _find_bank_build). Also used by _build_dict_to_rows to mark a
        result row with the Bank column's icon."""
        return bool(self._bank_owned_keys) and _bank_item_key(item) in self._bank_owned_keys

    def _find_optimal_build(self):
        """Find optimal item combination that covers ALL wanted spells"""
        # Set this immediately, before any validation checks below can
        # return early - otherwise clicking this button after a previous
        # "Show All Matches" left the Results tab stuck showing All Matches
        # results even though Best Per Slot is what was actually clicked.
        self.results_display_mode.set('optimal')

        if not self.master_data:
            messagebox.showwarning("No Data", "Please load a master database first")
            return
        
        # Get wanted spells - Priority Spells are always searched for too, not
        # just preferred among items that already match a wanted spell, so
        # they're folded in here before any item filtering happens. None of
        # these (nor Required Items/Wanted Sigils) are actually required to
        # run a search - weapon/shield slots always fall back to the best
        # available item regardless (see is_always_fill_slot below), so a
        # search with nothing wanted at all still works for e.g. just
        # weapon-hunting via Weapon/Melee Weapon Constraints.
        wanted_spells = list(self.wanted_spells_data)
        priority_spells = list(self.priority_spells_data)
        wanted_sigils = list(self.wanted_sigils_data)

        # Priority Tiers - each entry pairs a specific spell with a specific
        # tier. For that spell, the search targets that tier (still preferring
        # the highest of any priority tiers set for it if more than one),
        # rather than always chasing the single highest tier available. Other
        # spells are unaffected. Grouped per base spell for quick lookup below.
        priority_tier_ranks_by_base = {}
        for spell, tier in self.priority_tiers_data:
            priority_tier_ranks_by_base.setdefault(spell, set()).add(_TIER_RANK[tier])

        # A Wanted Spell chip added with an explicit tier (not "(any)") is
        # itself an implicit tier target, same as manually adding it to
        # Priority Tier - picking "Dexterity ii" specifically should search
        # for tier ii, not silently upgrade to the highest tier available
        # just because a tier iii dexterity item exists somewhere. A chip
        # added as "(any)" (no tier suffix) still has no target, so those
        # keep today's "prefer the highest tier found" behavior. Skipped
        # for a base that already has an explicit Priority Tier entry -
        # that's a deliberate, more specific request (Priority Tier can
        # target a tier *lower* than what a same-spell chip asked for,
        # e.g. for cost/rarity reasons - it shouldn't get diluted/outvoted
        # by the chip's own tier just because that tier happens to be higher).
        for wanted in wanted_spells:
            base = _spell_base(wanted)
            if base in priority_tier_ranks_by_base:
                continue
            explicit_tier = _spell_tier_rank(wanted)
            if explicit_tier > 0:
                priority_tier_ranks_by_base.setdefault(base, set()).add(explicit_tier)

        # Base spells (tier stripped) for the early candidate filter below -
        # a lower/higher tier than exactly requested should still qualify an
        # item as a candidate, not just an exact tier match.
        wanted_spell_bases = {_spell_base(w) for w in wanted_spells}
        wanted_spell_bases.update(priority_spells)

        # Wanted Sigils - lowercased for the early candidate filter below,
        # same idea as wanted_spell_bases but matched against an item's own
        # Sigil value instead of its Spell, and only for the armor slots
        # Wanted Sigils applies to (see ARMOR_SIGIL_SLOTS).
        wanted_sigils_lower = {s.strip().lower() for s in wanted_sigils}

        # Bank Build's "Prioritize items I own" - self._bank_owned_keys is
        # set by _find_bank_build right before calling this, cleared right
        # after (see _is_bank_owned). None (the normal case, every other
        # search) means owned_match is always 0 below.

        # Clear previous results
        self.search_results_tv.delete(*self.search_results_tv.get_children())

        # Get level constraints
        min_level = None
        max_level = None
        specific_level = None

        # Check for specific level first (takes precedence)
        specific_level_str = self.specific_level_var.get().strip()
        if specific_level_str:
            try:
                specific_level = int(specific_level_str)
            except ValueError:
                messagebox.showwarning("Invalid Level", "Specific level must be a number")
                return
        else:
            # If no specific level, check min/max
            min_level_str = self.min_level_var.get().strip()
            if min_level_str:
                try:
                    min_level = int(min_level_str)
                except ValueError:
                    messagebox.showwarning("Invalid Level", "Minimum level must be a number")
                    return

            max_level_str = self.max_level_var.get().strip()
            if max_level_str:
                try:
                    max_level = int(max_level_str)
                except ValueError:
                    messagebox.showwarning("Invalid Level", "Maximum level must be a number")
                    return

            # Validate min <= max
            if min_level is not None and max_level is not None and min_level > max_level:
                messagebox.showwarning("Invalid Range", "Minimum level cannot be greater than maximum level")
                return

        # Min/Max Tier - bounds how far a spell match can fall back to a
        # different tier than exactly requested (blank side = unbounded).
        min_tier_rank = _TIER_RANK.get(self.min_tier_var.get()) if self.min_tier_var.get() else None
        max_tier_rank = _TIER_RANK.get(self.max_tier_var.get()) if self.max_tier_var.get() else None
        if min_tier_rank is not None and max_tier_rank is not None and min_tier_rank > max_tier_rank:
            messagebox.showwarning("Invalid Range", "Minimum tier cannot be greater than maximum tier")
            return

        # Weight range - Melee Weapon Constraints' min/max on the item
        # list's Weight column (blank side = unbounded). Soft preference by
        # default; "Hard Filter" checked makes it exclude out-of-range
        # weapons outright instead, same as Min/Max Level does for level.
        weapon_weight_min = None
        weapon_weight_max = None
        weapon_weight_min_str = self.weapon_weight_min_var.get().strip()
        if weapon_weight_min_str:
            try:
                weapon_weight_min = int(weapon_weight_min_str)
            except ValueError:
                messagebox.showwarning("Invalid Weight", "Minimum weight must be a number")
                return
        weapon_weight_max_str = self.weapon_weight_max_var.get().strip()
        if weapon_weight_max_str:
            try:
                weapon_weight_max = int(weapon_weight_max_str)
            except ValueError:
                messagebox.showwarning("Invalid Weight", "Maximum weight must be a number")
                return
        if weapon_weight_min is not None and weapon_weight_max is not None and weapon_weight_min > weapon_weight_max:
            messagebox.showwarning("Invalid Range", "Minimum weight cannot be greater than maximum weight")
            return
        weapon_weight_hard = self.weapon_weight_hard_var.get()

        def _weapon_weight_in_range(item):
            """True if item's Weight falls within the configured range, or
            there's no range set at all. An item with a blank/non-numeric
            Weight fails this (same convention as Level filtering elsewhere
            in this search: missing data can't be confirmed in range, so it
            doesn't pass a hard filter) - only matters when Hard Filter is
            actually checked, since this is never called otherwise."""
            if weapon_weight_min is None and weapon_weight_max is None:
                return True
            try:
                item_weight = int(item.get('Weight') or '')
            except (ValueError, TypeError):
                return False
            if weapon_weight_min is not None and item_weight < weapon_weight_min:
                return False
            if weapon_weight_max is not None and item_weight > weapon_weight_max:
                return False
            return True

        def _weapon_weight_soft_match(item):
            """True only if item has a real, in-range Weight - unlike
            _weapon_weight_in_range (used for the hard filter, where a
            blank/non-numeric Weight passes through untouched), a blank
            Weight earns no soft-preference bonus here since there's
            nothing to actually match."""
            if weapon_weight_min is None and weapon_weight_max is None:
                return False
            try:
                item_weight = int(item.get('Weight') or '')
            except (ValueError, TypeError):
                return False
            if weapon_weight_min is not None and item_weight < weapon_weight_min:
                return False
            if weapon_weight_max is not None and item_weight > weapon_weight_max:
                return False
            return True

        # Defense range - always a soft preference here (never excludes an
        # item, so every slot still gets filled even when nothing in range
        # is available): checking "Defense:" prioritizes in-range items over
        # out-of-range ones when the search would otherwise be indifferent.
        defense_priority_active = self.use_defense_filter_var.get()
        defense_min_rank = None
        defense_max_rank = None
        if defense_priority_active:
            defense_min_rank = DEFENSE_RANK[self.min_defense_var.get()]
            defense_max_rank = DEFENSE_RANK[self.max_defense_var.get()]
            if defense_min_rank > defense_max_rank:
                messagebox.showwarning("Invalid Range", "Minimum Defense cannot be greater than maximum Defense")
                return

        # Per-slot Defense - stacks with the global control above rather than
        # replacing it: an item earns a priority point for each range (global
        # and/or this slot's own) it satisfies.
        slot_defense_priority_active = {}
        slot_defense_min_rank = {}
        slot_defense_max_rank = {}
        for slot, controls in self.slot_defense_controls.items():
            priority_active = controls['use'].get()
            slot_defense_priority_active[slot] = priority_active
            if priority_active:
                min_rank = DEFENSE_RANK[controls['min'].get()]
                max_rank = DEFENSE_RANK[controls['max'].get()]
                if min_rank > max_rank:
                    messagebox.showwarning("Invalid Range",
                        f"Minimum Defense cannot be greater than maximum Defense ({slot.title()})")
                    return
                slot_defense_min_rank[slot] = min_rank
                slot_defense_max_rank[slot] = max_rank

        # Shield Constraints' Defense - a single-value pick, fed into the
        # same per-slot machinery above as an exact one-value range.
        shield_defense_choice = self.shield_defense_var.get()
        if shield_defense_choice != 'Any':
            shield_defense_rank = DEFENSE_RANK[shield_defense_choice]
            slot_defense_priority_active['shield'] = True
            slot_defense_min_rank['shield'] = shield_defense_rank
            slot_defense_max_rank['shield'] = shield_defense_rank
        else:
            slot_defense_priority_active['shield'] = False

        def _defense_priority_score(item, lookup_slot):
            """Sum of priority points this item earns toward Defense
            targets - up to 2 (global + this slot's own), never a hard
            requirement, so it never keeps a slot from being filled."""
            score = 0
            item_defense_rank = None
            if defense_priority_active or slot_defense_priority_active.get(lookup_slot):
                item_defense_rank = DEFENSE_RANK.get((item.get('Defense') or '').strip().lower())
            if defense_priority_active and item_defense_rank is not None and defense_min_rank <= item_defense_rank <= defense_max_rank:
                score += 1
            if (slot_defense_priority_active.get(lookup_slot) and item_defense_rank is not None
                    and slot_defense_min_rank[lookup_slot] <= item_defense_rank <= slot_defense_max_rank[lookup_slot]):
                score += 1
            return score

        # Per-slot Sigil - soft preference, same philosophy as Defense: pick
        # a Sigil type for a slot and the search favors whichever candidate
        # for that slot carries it at the highest SigilLvl, but an item
        # without it is still eligible, so the slot never goes unfilled.
        slot_sigil_choice = {
            slot: var.get().strip().lower()
            for slot, var in self.slot_sigil_vars.items()
            if var.get() != 'Any'
        }
        # Shield Constraints' Sigil - same mechanism, just one more slot.
        if self.shield_sigil_var.get() != 'Any':
            slot_sigil_choice['shield'] = self.shield_sigil_var.get().strip().lower()

        def _sigil_priority_score(item, lookup_slot):
            """(sigil_match, sigil_level) for this item against the Sigil
            chosen for lookup_slot - (0, 0) if none chosen or it doesn't
            match, never a hard requirement."""
            wanted_sigil = slot_sigil_choice.get(lookup_slot)
            if not wanted_sigil:
                return (0, 0)
            item_sigil = (item.get('Sigil') or '').strip().lower()
            if item_sigil != wanted_sigil:
                return (0, 0)
            try:
                sigil_level = int(item.get('SigilLvl') or 0)
            except (ValueError, TypeError):
                sigil_level = 0
            return (1, sigil_level)

        # Melee Weapon Constraints - soft preferences that only apply to
        # Melee-styled weapons (non-direct, non-staff item type). Never a
        # hard requirement: the search always matches as many of these as it
        # can, but still fills the slot with the best available item even if
        # none match. Priority-checked fields (capped at 3) outscore the rest.
        melee_constraint_fields = [
            (self.melee_damage_var.get(), self.melee_damage_priority_var.get(), 'Damage'),
            (self.melee_timer_var.get(), self.melee_timer_priority_var.get(), 'Timer'),
            (self.melee_fumble_var.get(), self.melee_fumble_priority_var.get(), 'Fumble'),
            (self.melee_accuracy_var.get(), self.melee_accuracy_priority_var.get(), 'Accuracy'),
            (self.melee_sigil_var.get(), self.melee_sigil_priority_var.get(), 'Sigil'),
        ]

        def _melee_constraint_score(item, item_type):
            """(priority_matches, normal_matches) toward the Melee Weapon
            Constraints dropdowns - despite the section's name, these apply
            to any weapon style (Melee/Direct/Parry/Fired), not just plain
            Melee ones; Direct and Parry Staff items carry these same stat
            columns (Damage/Timer/Fumble/Accuracy/Sigil) too. (0, 0) if
            nothing is set beyond 'Any'."""
            priority_matches = 0
            normal_matches = 0
            for value, is_priority, column in melee_constraint_fields:
                if value == 'Any':
                    continue
                item_value = str(item.get(column) or '').strip()
                if column == 'Sigil' and value == 'None':
                    matched = not item_value
                else:
                    matched = item_value.lower() == value.lower()
                if matched:
                    if is_priority:
                        priority_matches += 1
                    else:
                        normal_matches += 1
            # Weight - no Priority checkbox of its own (see the Hard Filter
            # toggle instead), so it's always counted as a normal (non-
            # priority) match, same tier as an unchecked Damage/Timer/
            # Fumble/Accuracy/Sigil field above.
            if _weapon_weight_soft_match(item):
                normal_matches += 1
            return (priority_matches, normal_matches)

        # Get armor type constraints from checkboxes
        armor_constraints = {}
        for slot in ['head', 'cloak', 'body', 'hands', 'legs', 'feet']:
            # Get checked armor types for this slot
            checked_types = []
            for armor_type in ['cloth', 'leather', 'studded', 'plate']:
                if self.armor_checks[slot][armor_type].get():
                    checked_types.append(armor_type)
            armor_constraints[slot] = checked_types  # List of allowed types (empty = any)
        # Shield Constraints' armor type checkboxes - shields are technically
        # armor, so this reuses the exact same hard-filter mechanism as the
        # Armor Constraints tab's per-slot checkboxes, just sourced from
        # Weapon Constraints' own Shield Constraints checkboxes instead.
        armor_constraints['shield'] = [t for t in ('cloth', 'leather', 'studded', 'plate')
                                       if self.shield_armor_checks[t].get()]

        # Weapon Types/Combo's - each has its own Style and/or Damage Type
        # dropdown(s) (Claw has neither); the old global Weapon Style radios
        # and Damage Type checkboxes have been removed in favor of these.
        wants_two_handed = self.two_handed_var.get()
        two_handed_style = self.two_handed_style_var.get()
        two_handed_damage = self.two_handed_damage_var.get()
        wants_claw_1 = self.claw_1_var.get()
        wants_claw_2 = self.claw_2_var.get()
        wants_dual_wield_1h = self.dual_wield_1h_var.get()
        dual_wield_1h_main = self.dual_wield_1h_main_var.get()
        dual_wield_1h_off = self.dual_wield_1h_off_var.get()
        wants_1h_shield = self.combo_1h_shield_var.get()
        combo_1h_shield_style = self.combo_1h_shield_style_var.get()
        combo_1h_shield_damage = self.combo_1h_shield_damage_var.get()
        wants_2h_shield = self.combo_2h_shield_var.get()
        combo_2h_shield_damage = self.combo_2h_shield_damage_var.get()
        # Fired 1h and Shield - "fired" is its own weapon shape in the item
        # list with no slash/thrust/crush sub-variant, so no dropdown needed.
        wants_fired_1h_shield = self.combo_fired_1h_shield_var.get()

        # Most real weapons carry no Spell at all, so a checked combo has to
        # be able to pull one in without relying on the spell-match gate
        # below - otherwise the "always fill this slot" checkboxes would
        # never actually fill anything. Weapon-combo items still compete on
        # spell coverage first when they do have one; a spell-less item is
        # just the fallback pick when nothing better is available (see the
        # exact-search scoring, which already ranks coverage above all else).
        any_weapon_combo_active = (wants_two_handed or wants_dual_wield_1h or wants_1h_shield
                                    or wants_2h_shield or wants_fired_1h_shield)
        any_shield_combo_active = wants_1h_shield or wants_2h_shield or wants_fired_1h_shield
        # Claws also carry no Spell in practice - same reasoning applies.
        any_claw_combo_active = wants_claw_1 or wants_claw_2

        # Filter items by constraints and group by slot
        items_by_slot = {}
        all_slots = ['head', 'jewel', 'jewel', 'cloak', 'body', 'hands', 'legs', 'feet', 'weapon', 'shield']

        for item in self.master_data:
            item_slot = (item.get('Slot') or '').lower()
            item_spell = (item.get('Spell') or '').lower()
            item_type = (item.get('Type') or '').lower()
            item_sigil = (item.get('Sigil') or '').strip().lower()
            item_realm = (item.get('Realm') or '').strip()

            # Claws are stored in the source data as Slot=weapon/Type=claw,
            # not a distinct Slot='claw' value - there is no such value
            # anywhere in the bundled equipment list, so any check against
            # item_slot == 'claw' can never match a real row.
            is_claw_item = item_slot == 'weapon' and 'claw' in item_type

            is_combo_mandated_slot = ((item_slot == 'weapon' and not is_claw_item and any_weapon_combo_active)
                                       or (item_slot == 'shield' and any_shield_combo_active)
                                       or (is_claw_item and any_claw_combo_active))
            # Weapon and shield are always in slots_to_fill (see below) even
            # with no combo checked at all, so they must always be allowed to
            # fall back to a non-wanted-spell (or spell-less) item too - not
            # just when a combo happens to be active - otherwise, once
            # nothing carrying the wanted spell is left, the candidate pool
            # for these slots would be empty and the exact search's own
            # zero-new-coverage fallback (further below) would have nothing
            # to pick from except items already claimed by another slot,
            # which is what caused the same spell/tier to appear to be
            # "duplicated" into a second slot for no real benefit.
            is_always_fill_slot = (item_slot == 'weapon' and not is_claw_item) or item_slot == 'shield'

            # Crafted items never show up unless the Crafted checkbox is
            # explicitly checked - even if the item's Realm string also
            # happens to contain another checked realm (e.g. "Crafted -
            # Evil" checked when only "Evil" is on), and even when no realm
            # checkboxes are checked at all (Crafted is not part of the
            # implicit "search everything" default). The at-most-one-
            # Crafted-item-per-build cap below still applies once Crafted
            # is checked.
            if 'crafted' in item_realm.lower() and not self.realm_filters['Crafted'].get():
                continue

            # Apply realm filter if any selected
            selected_realms = [realm for realm, var in self.realm_filters.items()
                              if var.get()]
            if selected_realms:
                realm_match = False
                for selected in selected_realms:
                    if selected.lower() in item_realm.lower():
                        realm_match = True
                        break
                if not realm_match:
                    continue

            # Weight - only meaningful for weapons (includes claws, which
            # are stored as Slot=weapon too). Hard Filter excludes anything
            # outside the range outright; otherwise it's left as a soft
            # preference, scored later in _melee_constraint_score.
            if item_slot == 'weapon' and weapon_weight_hard and not _weapon_weight_in_range(item):
                continue

            # Wanted Sigils - many armor pieces carry a Sigil but no Spell at
            # all, so without this an item that only satisfies a Wanted
            # Sigil would never even become a candidate below. Scoped to the
            # armor slots Wanted Sigils applies to (see ARMOR_SIGIL_SLOTS).
            has_wanted_sigil = (bool(wanted_sigils_lower) and item_slot in ARMOR_SIGIL_SLOTS
                                and item_sigil in wanted_sigils_lower)

            # Skip items without spells - unless a weapon/shield combo wants
            # this slot filled regardless (most real weapons have no spell),
            # or the item is otherwise wanted for its Sigil alone.
            if not item_spell and not has_wanted_sigil and not is_combo_mandated_slot and not is_always_fill_slot:
                continue

            # Check if item's spell matches any wanted or priority spell's
            # base - any tier counts here, so an item at a different tier
            # than exactly requested still becomes an eligible candidate.
            has_wanted_spell = False
            for wanted in wanted_spell_bases:
                if wanted in item_spell:
                    has_wanted_spell = True
                    break

            if not has_wanted_spell and not has_wanted_sigil and not is_combo_mandated_slot and not is_always_fill_slot:
                continue

            # Apply level constraints
            if specific_level is not None or min_level is not None or max_level is not None:
                item_level_str = item.get('Level', '')
                if item_level_str:
                    try:
                        item_level = int(item_level_str)

                        # Check specific level - a fallback policy other than
                        # "Don't populate slot" widens this to "at or below"
                        # so lower-level/tier candidates are still in the
                        # pool for the per-base qualification check further
                        # below to consider; never above, since fallback
                        # only ever goes down.
                        if specific_level is not None:
                            if self.specific_level_fallback_var.get() == 'none':
                                if item_level != specific_level:
                                    continue
                            elif item_level > specific_level:
                                continue
                        # Check min/max range
                        else:
                            if min_level is not None and item_level < min_level:
                                continue
                            if max_level is not None and item_level > max_level:
                                continue
                    except ValueError:
                        # Skip items with invalid level data
                        continue
                else:
                    # Skip items without level data when filtering by level
                    continue
            
            # Defense is never a hard filter here - it's scored later as a
            # soft preference (see _defense_priority_score) so a slot is
            # never left empty just because nothing hits the Defense range.

            # Apply armor type constraints (checkboxes)
            if item_slot in armor_constraints:
                allowed_types = armor_constraints[item_slot]
                # If any checkboxes are checked, item must match one of them
                if allowed_types:
                    type_match = False
                    for allowed_type in allowed_types:
                        if allowed_type.lower() in item_type:
                            type_match = True
                            break
                    if not type_match:
                        continue
            
            # Apply weapon/shield/claw constraints. The old global Weapon
            # Style radios and Damage Type checkboxes are gone - Style and
            # Damage Type are now per-combo dropdowns (see each branch below).
            if item_slot == 'weapon' or item_slot == 'shield':
                slot_accepted = True
                offhand_weapon_accepted = False  # only ever set for physical (non-claw) weapon items

                if is_claw_item:
                    # Claws (Slot=weapon/Type=claw) have no Style or Damage
                    # Type dropdown - 1 Claw fills one claw slot, 2 Claw
                    # fills both. Handled entirely separately from the
                    # physical-weapon sub-matching below, which a Type=claw
                    # item could never satisfy anyway (its Type string is
                    # just "claw" - no 1h/2h/offhand/fired substring).
                    slot_accepted = wants_claw_1 or wants_claw_2
                else:
                    # If no combo is checked, accept everything
                    has_any_config = (wants_two_handed or wants_claw_1 or wants_claw_2
                                       or wants_dual_wield_1h or wants_1h_shield
                                       or wants_2h_shield or wants_fired_1h_shield)

                    if has_any_config:
                        slot_accepted = False

                        if item_slot == 'weapon':
                            is_offhand = 'offhand' in item_type
                            is_1h = '1h' in item_type
                            is_2h = '2h' in item_type
                            is_fired = 'fired' in item_type

                            # Two-handed weapons, gated by their own Style
                            # (Melee/Direct/Parry/Fired) and Damage Type dropdowns
                            if wants_two_handed and _two_handed_matches(item_type, item_spell, two_handed_style, two_handed_damage):
                                slot_accepted = True
                            # Dual-Wield 1h - main hand: 1h, non-offhand,
                            # matching the Main dropdown's damage type
                            if (wants_dual_wield_1h and is_1h and not is_offhand
                                    and _weapon_damage_matches(item_type, dual_wield_1h_main)):
                                slot_accepted = True
                            # 1h/Shield: 1h, non-offhand, matching its own
                            # Style (Melee/Direct - no Parry option here) and
                            # Damage Type dropdowns. Direct additionally requires
                            # some spell to be present (see _direct_weapon_eligible).
                            if (wants_1h_shield and is_1h and not is_offhand
                                    and _weapon_style_matches(item_type, combo_1h_shield_style)
                                    and (combo_1h_shield_style != 'Direct' or _direct_weapon_eligible(item_spell))
                                    and _weapon_damage_matches(item_type, combo_1h_shield_damage)):
                                slot_accepted = True
                            # 2h/Shield: 2h, matching its own damage type
                            if (wants_2h_shield and is_2h
                                    and _weapon_damage_matches(item_type, combo_2h_shield_damage)):
                                slot_accepted = True
                            # Fired 1h and Shield: "fired" is its own weapon
                            # shape, no separate damage type to match.
                            if wants_fired_1h_shield and is_fired and is_1h and not is_offhand:
                                slot_accepted = True

                            # Dual-Wield 1h - off hand: offhand-tagged 1h item
                            # matching the Off-Hand dropdown's damage type.
                            # Independent of the main-hand outcome above, so
                            # this same item can be considered for the
                            # separate weapon_off slot even when it wasn't
                            # accepted as a main-hand candidate.
                            if (wants_dual_wield_1h and is_1h and is_offhand
                                    and _weapon_damage_matches(item_type, dual_wield_1h_off)):
                                offhand_weapon_accepted = True

                        elif item_slot == 'shield':
                            # Shields - only implied by a weapon+shield combo
                            # now (plain "Shield" was removed)
                            if wants_1h_shield or wants_2h_shield or wants_fired_1h_shield:
                                slot_accepted = True

                if not slot_accepted and not offhand_weapon_accepted:
                    continue

            # Add to slot group. The 'weapon' slot splits into three candidate
            # pools - main-hand (items_by_slot['weapon']), off-hand (only
            # relevant when Dual-Wield 1h is checked, items_by_slot['weapon_off']),
            # and claws (items_by_slot['claw'], Type=claw items only) - since
            # the same physical item can never satisfy more than one of
            # these, an item lands in at most one bucket.
            if is_claw_item:
                if slot_accepted:
                    items_by_slot.setdefault('claw', []).append(item)
            elif item_slot == 'weapon':
                if slot_accepted:
                    items_by_slot.setdefault('weapon', []).append(item)
                if offhand_weapon_accepted:
                    items_by_slot.setdefault('weapon_off', []).append(item)
            else:
                items_by_slot.setdefault(item_slot, []).append(item)

        # Kept for Alt Options: any other item sharing a slot's spell, within
        # the same level/armor/weapon/realm constraints already applied above.
        self.items_by_slot = items_by_slot

        # OPTIMAL BUILD ALGORITHM: Greedy approach to cover maximum spells
        # Wanted spells are grouped by base name (stripping .i/.ii/.iii) so that
        # requesting the same spell at two tiers is one requirement, not two -
        # a single item satisfying either tier fills it, instead of hunting for
        # a second item elsewhere in the build.
        wanted_bases = {}
        for wanted in wanted_spells:
            wanted_bases.setdefault(_spell_base(wanted), []).append(wanted)

        # Priority Spells are always searched for, not just preferred among
        # items that already match a wanted spell - merge them in as their own
        # base so an item providing only a priority spell (nothing else wanted)
        # still qualifies as a candidate and actively gets sought out.
        for p in priority_spells:
            wanted_bases.setdefault(p, []).append(p)

        # Specific Level fallback - for each base spell, the set of tiers
        # explicitly requested (a bare/"(any)" chip or a Priority Spell
        # contributes no tier, leaving this empty - tier is already
        # unconstrained for those regardless of the fallback policy below).
        base_target_tier_ranks = {
            base: {_spell_tier_rank(w) for w in wanteds if _spell_tier_rank(w) > 0}
            for base, wanteds in wanted_bases.items()
        }
        specific_level_fallback = self.specific_level_fallback_var.get() if specific_level is not None else 'none'

        def _specific_level_qualifies(base, item_level_num, item_tier):
            """Whether an item carrying `base` at `item_level_num`/`item_tier`
            counts toward that wanted base under Specific Level - True
            unconditionally when Specific Level isn't set. Each fallback
            policy relaxes exactly one or both of level/tier away from an
            exact match; "Don't populate slot" (or no Specific Level fallback
            requested) requires both exactly. Levels/tiers are only ever
            relaxed downward, never up - matching the widened raw level
            filter above, which never lets an item above Specific Level
            through in the first place."""
            if specific_level is None:
                return True
            target_ranks = base_target_tier_ranks.get(base)
            level_exact = (item_level_num == specific_level)
            tier_exact = (not target_ranks) or (item_tier in target_ranks)

            if specific_level_fallback == 'none':
                return level_exact and tier_exact
            if specific_level_fallback == 'tier':
                if not level_exact:
                    return False
                return tier_exact or (bool(target_ranks) and 0 < item_tier <= max(target_ranks))
            if specific_level_fallback == 'level':
                if not tier_exact:
                    return False
                return item_level_num <= specific_level
            # 'both' - level and tier can each independently fall short of
            # exact, as long as neither goes above what was requested.
            tier_ok = tier_exact or (bool(target_ranks) and 0 < item_tier <= max(target_ranks))
            return item_level_num <= specific_level and tier_ok

        build = {}
        covered_bases = set()
        self.slot_alternates = {}

        # Force any Required Items directly into the build first - they aren't
        # subject to the other filters (armor/weapon/level/realm), since the
        # user explicitly asked for them. The rest of the build is calculated
        # around them: their spells count as covered and their slots are skipped
        # below, same as anything the greedy search would have picked itself.
        crafted_count, used_claw_items = self._assign_required_items(build, covered_bases, wanted_bases)

        # Max Lvl (Armor Constraints) - up to 3 armor slots can be marked to
        # greedily grab the highest-level candidate for that slot BEFORE the
        # exact search below runs, rather than letting the exact search pick
        # whichever item best optimizes overall spell coverage. Candidates
        # here are already restricted to items carrying a wanted/priority
        # spell and matching that slot's armor type checkboxes (Cloth/
        # Leather/Studded/Plate), since items_by_slot was built with those
        # filters already applied - Max Lvl only changes which of those
        # already-eligible items wins, never what's eligible in the first
        # place. Ties (same highest level) are broken the same way the exact
        # search would: Priority Spell match, then Priority Tier match, then
        # Sigil match/level, then Defense preference, then spell tier.
        for slot, var in self.armor_maxlvl_vars.items():
            if not var.get() or slot in build:
                continue
            best_item = None
            best_key = None
            for item in items_by_slot.get(slot, []):
                is_crafted = 'crafted' in (item.get('Realm') or '').strip().lower()
                if is_crafted and crafted_count >= MAX_CRAFTED_ITEMS:
                    continue
                try:
                    item_level_num = int(item.get('Level') or 0)
                except (ValueError, TypeError):
                    item_level_num = 0
                item_spell = (item.get('Spell') or '').lower()
                item_tier = _item_tier_rank(item_spell)
                item_base = _spell_base(item_spell)
                priority_matches = sum(1 for p in priority_spells if p in item_spell)
                item_priority_tiers = priority_tier_ranks_by_base.get(item_base)
                tier_priority_match = 1 if item_priority_tiers and item_tier in item_priority_tiers else 0
                sigil_match, sigil_level = _sigil_priority_score(item, slot)
                defense_priority_match = _defense_priority_score(item, slot)
                key = (item_level_num, priority_matches, tier_priority_match,
                       sigil_match, sigil_level, defense_priority_match, item_tier)
                if best_key is None or key > best_key:
                    best_key = key
                    best_item = item

            if best_item is not None:
                build[slot] = best_item
                matched_bases = [base for base in wanted_bases if base in (best_item.get('Spell') or '').lower()]
                covered_bases.update(matched_bases)
                if 'crafted' in (best_item.get('Realm') or '').strip().lower():
                    crafted_count += 1

        # Slots to fill (including 2 jewel slots). Claw slots are only attempted
        # when 1 Claw / 2 Claw is checked - 2 Claw fills both claw slots (dual-wield).
        # weapon_off (a second physical weapon, for Dual-Wield 1h) is only
        # attempted when that combo is checked - it shares its candidate pool
        # with the item list's Slot=weapon items (see items_by_slot['weapon_off']
        # below), but is otherwise treated as an ordinary slot by the exact
        # search, so it's optimized for spell coverage right alongside everything else.
        # Claws are their own one-handed weapon - a build using 1 or 2 Claw
        # doesn't also equip a separate physical weapon/shield, so those two
        # slots are left out entirely rather than always-filled alongside them.
        # Likewise, weapon/shield are only attempted at all if at least one
        # Weapon Types/Combo's checkbox that implies them is checked - with
        # nothing checked there's no reason to force a weapon or shield into
        # the build, so both are left unpopulated rather than always-filled.
        slots_to_fill = ['head', 'jewel_1', 'jewel_2', 'cloak', 'body', 'hands', 'legs', 'feet']
        if not (wants_claw_1 or wants_claw_2):
            if any_weapon_combo_active:
                slots_to_fill.append('weapon')
            if any_shield_combo_active:
                slots_to_fill.append('shield')
        if wants_dual_wield_1h:
            slots_to_fill.append('weapon_off')
        if wants_claw_2:
            slots_to_fill += ['claw_1', 'claw_2']
        elif wants_claw_1:
            slots_to_fill += ['claw_1']


        # Claw slots allow redundant/duplicate spell coverage (dual-wielding
        # needs a physical item in each hand even if it repeats a spell) so
        # they don't fit the "each base covered by exactly one slot" model
        # below - they're filled the old greedy way, after the exact search
        # settles everything else.
        exact_slots = [s for s in slots_to_fill if not s.startswith('claw') and s not in build]
        claw_slots = [s for s in slots_to_fill if s.startswith('claw') and s not in build]

        # EXACT OPTIMAL BUILD SEARCH: a fixed slot-processing order means a
        # greedy pass can lock a spell into a lower tier in an early slot even
        # though a higher tier for it sits unclaimed in a later slot occupied
        # by something swappable (e.g. strength stuck at .ii on a jewel while
        # a .iii item exists on boots, because feet hadn't been considered
        # yet when jewel grabbed its best-available match). Wanted/priority
        # bases are tracked as a bitmask, one bit each, so the search state -
        # and therefore what gets memoized - stays small enough to explore
        # every real combination instead of committing slot-by-slot.
        base_list = list(wanted_bases.keys())
        base_bit = {base: (1 << i) for i, base in enumerate(base_list)}
        initial_covered_bitmask = 0
        for base in covered_bases:
            initial_covered_bitmask |= base_bit.get(base, 0)

        # Score weights: each tier's full possible range is dwarfed by one
        # unit of the tier above it, so summing these as plain integers along
        # a search path is equivalent to comparing (bases_covered,
        # priority_matches, tier_priority_hits, sigil_match, sigil_level,
        # melee_priority_constraint_matches, melee_constraint_matches,
        # defense_priority_hits, tier, level) lexicographically, while
        # staying simple/fast integer math. Each step is 10^6 above the
        # last - generously larger than any realistic per-slot total below it.
        W_LEVEL = 1
        # Jewels have no armor type, and no per-slot Defense/Sigil options
        # yet (Sigil for jewels isn't implemented), so unlike an armor
        # slot, a jewel's own level carries no real significance - it's
        # just whichever one happens to carry the wanted spell. Weighted
        # above W_LEVEL but below W_TIER so the search still prefers a
        # jewel over an armor-slot item for the exact same wanted spell
        # (trying jewels before spreading requirements onto armor slots,
        # per the user's request), without letting that preference override
        # a genuine tier difference (a higher spell tier elsewhere still wins).
        W_JEWEL_PREFERENCE = 10**3
        W_TIER = 10**6
        W_DEFENSE_PRIORITY = 10**12
        W_MELEE_MATCH = 10**18
        W_MELEE_PRIORITY_MATCH = 10**24
        W_SIGIL_LEVEL = 10**30
        W_SIGIL_MATCH = 10**36
        # Wanted Sigils - deliberately below W_TIER_PRIORITY/W_PRIORITY/
        # W_COVERAGE so a Wanted Spell always wins a slot over a Wanted
        # Sigil when both are available (sigils are a secondary search),
        # but above the passive per-slot Sigil preference dropdown since
        # this is a more deliberate ask than that soft tie-break.
        W_WANTED_SIGIL = 10**39
        # Bank Build's "Prioritize items I own" - deliberately below
        # W_TIER_PRIORITY/W_PRIORITY/W_COVERAGE (an actual Wanted Spell/
        # Priority Tier need still wins over mere ownership), but above
        # W_WANTED_SIGIL since owning an item is a firmer asset than a
        # soft Sigil wish. 0 for every candidate outside Bank Build.
        W_BANK_OWNED = 10**40
        W_TIER_PRIORITY = 10**42
        W_PRIORITY = 10**48
        W_COVERAGE = 10**54

        # Pre-score every item once per lookup slot, collapsing items that are
        # identical in everything the search cares about (bases matched,
        # priority/tier-priority/defense-priority hits, tier, crafted-ness)
        # down to their single highest-level representative - keeps the
        # search fast without changing which overall score is reachable.
        # Alternates for "Generate multiple build options" are recomputed
        # from the full unpruned list further below, so nothing is lost.
        candidates_by_slot = {}
        for lookup_slot, items in items_by_slot.items():
            if lookup_slot == 'claw':
                continue
            seen = {}
            for item in items:
                item_spell = (item.get('Spell') or '').lower()
                item_type = (item.get('Type') or '').lower()
                item_sigil = (item.get('Sigil') or '').strip().lower()
                item_tier = _item_tier_rank(item_spell)
                if item_tier > 0:
                    if min_tier_rank is not None and item_tier < min_tier_rank:
                        continue
                    if max_tier_rank is not None and item_tier > max_tier_rank:
                        continue

                try:
                    item_level_num = int(item.get('Level') or 0)
                except (ValueError, TypeError):
                    item_level_num = 0

                item_bitmask = 0
                for base in base_list:
                    if base in item_spell and _specific_level_qualifies(base, item_level_num, item_tier):
                        item_bitmask |= base_bit[base]

                # Wanted Sigils - unlike Wanted Spells, sigils don't stack
                # toward one shared "covered" requirement (see wanted_bases
                # above, which is spell-only): every armor slot that can
                # carry a wanted Sigil is independently worth taking, so
                # this is a per-slot score bonus (added to step_score below
                # via W_WANTED_SIGIL) rather than a bit in item_bitmask -
                # duplicates across slots are fine, even encouraged. Kept
                # low enough in the weight ordering that a Wanted Spell
                # always wins the slot when both are available.
                wanted_sigil_match = 1 if (lookup_slot in ARMOR_SIGIL_SLOTS
                                            and item_sigil in wanted_sigils_lower) else 0

                # Bank Build's "Prioritize items I own" - always 0 outside
                # that flow (bank_owned_keys is None), so this is a no-op
                # for every other search.
                owned_match = 1 if self._is_bank_owned(item) else 0

                # Zero-coverage items are normally useless (they can never
                # win over "leave the slot empty"), except weapon/shield
                # slots a combo has mandated be filled regardless of spell,
                # or an armor item that at least carries a Wanted Sigil -
                # those still need a fallback candidate so the slot doesn't
                # end up empty just because nothing on it grants a spell.
                if not item_bitmask and not wanted_sigil_match and lookup_slot not in ('weapon', 'weapon_off', 'shield'):
                    continue

                priority_matches = sum(1 for p in priority_spells if p in item_spell)
                item_base = _spell_base(item_spell)
                item_priority_tiers = priority_tier_ranks_by_base.get(item_base)
                tier_priority_match = 1 if item_priority_tiers and item_tier in item_priority_tiers else 0
                is_crafted = 'crafted' in (item.get('Realm') or '').strip().lower()

                # "Prioritize" Defense: a soft bonus (global + per-slot) for
                # landing in the Min/Max Defense range(s) - never excludes
                # the item, unlike the hard "Defense:" filters applied above.
                defense_priority_match = _defense_priority_score(item, lookup_slot)

                # Sigil: a soft bonus for carrying the Sigil chosen for this
                # slot, tie-broken by SigilLvl - never excludes the item.
                sigil_match, sigil_level = _sigil_priority_score(item, lookup_slot)

                # Melee Weapon Constraints: only meaningful for weapon slots.
                melee_priority_match, melee_match = (
                    _melee_constraint_score(item, item_type) if lookup_slot in ('weapon', 'weapon_off') else (0, 0))

                sig = (item_bitmask, priority_matches, tier_priority_match, sigil_match,
                       sigil_level, wanted_sigil_match, owned_match, melee_priority_match, melee_match,
                       defense_priority_match, item_tier, is_crafted)
                prev = seen.get(sig)
                if prev is None or item_level_num > prev[1]:
                    seen[sig] = (item, item_level_num)

            candidates_by_slot[lookup_slot] = [
                (item, sig[0], sig[1], sig[2], sig[3], sig[4], sig[5], sig[6], sig[7], sig[8], sig[9],
                 sig[10], lvl, sig[11])
                for sig, (item, lvl) in seen.items()
            ]

        memo = {}

        def solve(idx, covered, crafted_n):
            if idx == len(exact_slots):
                return (0, [])
            key = (idx, covered, crafted_n)
            cached = memo.get(key)
            if cached is not None:
                return cached

            slot = exact_slots[idx]
            lookup_slot = 'jewel' if slot.startswith('jewel') else slot

            best_score, best_rest = solve(idx + 1, covered, crafted_n)
            best_choice = None

            for (item, item_bitmask, priority_matches, tier_priority_match, sigil_match, sigil_level,
                 wanted_sigil_match, owned_match, melee_priority_match, melee_match, defense_priority_match,
                 item_tier, item_level_num, is_crafted) in candidates_by_slot.get(lookup_slot, []):
                new_bases = item_bitmask & ~covered
                # Same fallback as above: a combo-mandated weapon/shield slot
                # can still take a zero-new-coverage item (any positive score
                # from level/tier/etc. beats leaving the slot empty), and so
                # can an armor item carrying a Wanted Sigil - every other
                # slot still requires contributing something new.
                if not new_bases and not wanted_sigil_match and lookup_slot not in ('weapon', 'weapon_off', 'shield'):
                    continue
                if is_crafted and crafted_n >= MAX_CRAFTED_ITEMS:
                    continue

                # An item whose spell is already fully covered elsewhere (new_bases
                # == 0 - only reachable here via the weapon/weapon_off/shield
                # fallback above) isn't actually fulfilling that spell, so its
                # tier/priority standing is meaningless noise - without zeroing
                # these out, a fallback pick could win purely because its
                # (unused) spell happens to be a high tier, making the same
                # spell/tier appear to be "duplicated" into a second slot for
                # no real benefit. Sigil/Melee/Defense preferences and level
                # stay live since those never depended on spell coverage.
                effective_priority_matches = priority_matches if new_bases else 0
                effective_tier_priority_match = tier_priority_match if new_bases else 0
                effective_item_tier = item_tier if new_bases else 0
                # Last-resort tie-break, folded into the level term (doubled
                # so a genuine 1-level difference still dominates): among
                # fallback items that are otherwise equal, prefer one whose
                # spell isn't a duplicate of an already-covered base at all
                # over one that happens to carry a now-redundant copy of a
                # wanted spell - purely cosmetic (neither contributes
                # coverage) but avoids the same spell/tier visibly showing up
                # in two slots when a spell-unrelated alternative is just as good.
                is_redundant_spell = 1 if (not new_bases and item_bitmask) else 0

                step_score = (bin(new_bases).count('1') * W_COVERAGE
                              + effective_priority_matches * W_PRIORITY
                              + effective_tier_priority_match * W_TIER_PRIORITY
                              + wanted_sigil_match * W_WANTED_SIGIL
                              + owned_match * W_BANK_OWNED
                              + sigil_match * W_SIGIL_MATCH
                              + sigil_level * W_SIGIL_LEVEL
                              + melee_priority_match * W_MELEE_PRIORITY_MATCH
                              + melee_match * W_MELEE_MATCH
                              + defense_priority_match * W_DEFENSE_PRIORITY
                              + effective_item_tier * W_TIER
                              + (W_JEWEL_PREFERENCE if lookup_slot == 'jewel' else 0)
                              + item_level_num * (W_LEVEL * 2)
                              + (0 if is_redundant_spell else W_LEVEL))

                rest_score, rest_path = solve(idx + 1, covered | new_bases,
                                               crafted_n + (1 if is_crafted else 0))
                total = step_score + rest_score
                if total > best_score:
                    best_score = total
                    best_choice = item
                    best_rest = rest_path

            result = (best_score,
                      [(slot, best_choice)] + best_rest) if best_choice is not None else (
                      best_score, [(slot, None)] + best_rest)
            memo[key] = result
            return result

        _, assignment = solve(0, initial_covered_bitmask, crafted_count)

        base_covered_by_slot = {}
        for slot, item in assignment:
            if item is None:
                continue
            build[slot] = item
            item_spell = (item.get('Spell') or '').lower()
            for base in base_list:
                if base in item_spell and base not in covered_bases:
                    covered_bases.add(base)
                    base_covered_by_slot[base] = slot
            if 'crafted' in (item.get('Realm') or '').strip().lower():
                crafted_count += 1

        # Post-hoc alternates for the exact-search slots: any other item that
        # would have contributed the exact same NEW bases at the exact same
        # priority/tier-priority/tier/level score, had it been chosen instead
        # of the one the search picked - genuinely interchangeable, so
        # swapping it into a build variant can't duplicate or drop a spell
        # elsewhere in the build.
        final_covered_bitmask = 0
        for base in covered_bases:
            final_covered_bitmask |= base_bit.get(base, 0)

        for slot, item in assignment:
            if item is None:
                continue
            lookup_slot = 'jewel' if slot.startswith('jewel') else slot
            this_slot_bitmask = 0
            for base, owner in base_covered_by_slot.items():
                if owner == slot:
                    this_slot_bitmask |= base_bit[base]
            covered_without_slot = final_covered_bitmask & ~this_slot_bitmask

            item_spell = (item.get('Spell') or '').lower()
            item_type = (item.get('Type') or '').lower()
            item_tier = _item_tier_rank(item_spell)
            item_is_crafted = 'crafted' in (item.get('Realm') or '').strip().lower()
            priority_matches = sum(1 for p in priority_spells if p in item_spell)
            item_base = _spell_base(item_spell)
            item_priority_tiers = priority_tier_ranks_by_base.get(item_base)
            tier_priority_match = 1 if item_priority_tiers and item_tier in item_priority_tiers else 0
            item_defense_priority_match = _defense_priority_score(item, lookup_slot)
            item_sigil_match, item_sigil_level = _sigil_priority_score(item, lookup_slot)
            item_wanted_sigil_match = 1 if (lookup_slot in ARMOR_SIGIL_SLOTS
                                             and (item.get('Sigil') or '').strip().lower() in wanted_sigils_lower) else 0
            item_owned_match = 1 if self._is_bank_owned(item) else 0
            item_melee_priority_match, item_melee_match = (
                _melee_constraint_score(item, item_type) if lookup_slot in ('weapon', 'weapon_off') else (0, 0))
            try:
                item_level_num = int(item.get('Level') or 0)
            except (ValueError, TypeError):
                item_level_num = 0
            chosen_key = (priority_matches, tier_priority_match, item_sigil_match, item_sigil_level,
                          item_wanted_sigil_match, item_owned_match, item_melee_priority_match, item_melee_match,
                          item_defense_priority_match, item_tier, item_level_num)

            alternates = []
            for other in items_by_slot.get(lookup_slot, []):
                if other is item:
                    continue
                other_spell = (other.get('Spell') or '').lower()
                other_tier = _item_tier_rank(other_spell)
                if other_tier > 0:
                    if min_tier_rank is not None and other_tier < min_tier_rank:
                        continue
                    if max_tier_rank is not None and other_tier > max_tier_rank:
                        continue

                other_is_crafted = 'crafted' in (other.get('Realm') or '').strip().lower()
                if other_is_crafted and not item_is_crafted and crafted_count >= MAX_CRAFTED_ITEMS:
                    continue

                other_type = (other.get('Type') or '').lower()
                new_bases = 0
                for base in base_list:
                    if base in other_spell:
                        new_bases |= base_bit[base]
                new_bases &= ~covered_without_slot
                if new_bases != this_slot_bitmask:
                    continue

                other_priority_matches = sum(1 for p in priority_spells if p in other_spell)
                other_base = _spell_base(other_spell)
                other_priority_tiers = priority_tier_ranks_by_base.get(other_base)
                other_tier_priority_match = 1 if other_priority_tiers and other_tier in other_priority_tiers else 0
                other_defense_priority_match = _defense_priority_score(other, lookup_slot)
                other_sigil_match, other_sigil_level = _sigil_priority_score(other, lookup_slot)
                other_wanted_sigil_match = 1 if (lookup_slot in ARMOR_SIGIL_SLOTS
                                                  and (other.get('Sigil') or '').strip().lower() in wanted_sigils_lower) else 0
                other_owned_match = 1 if self._is_bank_owned(other) else 0
                other_melee_priority_match, other_melee_match = (
                    _melee_constraint_score(other, other_type) if lookup_slot in ('weapon', 'weapon_off') else (0, 0))
                try:
                    other_level_num = int(other.get('Level') or 0)
                except (ValueError, TypeError):
                    other_level_num = 0
                other_key = (other_priority_matches, other_tier_priority_match, other_sigil_match, other_sigil_level,
                            other_wanted_sigil_match, other_owned_match, other_melee_priority_match, other_melee_match,
                            other_defense_priority_match, other_tier, other_level_num)
                if other_key != chosen_key:
                    continue

                alternates.append(other)

            if alternates:
                self.slot_alternates[slot] = alternates

        # Claw slots: unchanged greedy fill (see comment above exact_slots) -
        # duplicate spell coverage across both hands is fine here, so they're
        # picked after the rest of the build's coverage is already settled.
        # Claws carry no Spell in practice, same as most weapons - reaching
        # this loop at all already means a claw combo is checked, so a
        # zero-coverage item is still a valid fallback pick (matched_bases
        # empty just means it contributes nothing new, not that it's invalid).
        #
        # Sigil preference is per exact slot (claw_1/claw_2), not shared like
        # everything else above - claw_1 is always the first claw filled
        # (whether 1 Claw or 2 Claw is checked), claw_2 only exists when 2
        # Claw is checked, and each can want a different Sigil.
        claw_sigil_choice = {
            'claw_1': self.claw_1_sigil_var.get().strip().lower(),
            'claw_2': self.claw_2_sigil_var.get().strip().lower(),
        }
        for slot in claw_slots:
            lookup_slot = 'claw'
            if lookup_slot not in items_by_slot:
                continue

            wanted_claw_sigil = claw_sigil_choice.get(slot)
            if wanted_claw_sigil == 'any':
                wanted_claw_sigil = ''

            best_item = None
            best_matched_bases = []
            best_key = (-1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1)
            tied_items = []

            for item in items_by_slot[lookup_slot]:
                if any(item is used for used in used_claw_items):
                    continue

                item_realm = (item.get('Realm') or '').strip().lower()
                if 'crafted' in item_realm and crafted_count >= MAX_CRAFTED_ITEMS:
                    continue

                item_spell = (item.get('Spell') or '').lower()
                item_type = (item.get('Type') or '').lower()
                item_tier = _item_tier_rank(item_spell)
                if item_tier > 0:
                    if min_tier_rank is not None and item_tier < min_tier_rank:
                        continue
                    if max_tier_rank is not None and item_tier > max_tier_rank:
                        continue

                try:
                    item_level_num = int(item.get('Level') or 0)
                except (ValueError, TypeError):
                    item_level_num = 0

                matched_bases = [base for base in wanted_bases if base in item_spell]

                priority_matches = sum(1 for p in priority_spells if p in item_spell)
                item_base = _spell_base(item_spell)
                item_priority_tiers = priority_tier_ranks_by_base.get(item_base)
                tier_priority_match = 1 if item_priority_tiers and item_tier in item_priority_tiers else 0
                defense_priority_match = _defense_priority_score(item, lookup_slot)
                melee_priority_match, melee_match = _melee_constraint_score(item, item_type)

                claw_sigil_match = 0
                claw_sigil_level = 0
                if wanted_claw_sigil and (item.get('Sigil') or '').strip().lower() == wanted_claw_sigil:
                    claw_sigil_match = 1
                    try:
                        claw_sigil_level = int(item.get('SigilLvl') or 0)
                    except (ValueError, TypeError):
                        claw_sigil_level = 0

                owned_match = 1 if self._is_bank_owned(item) else 0

                key = (priority_matches, len(matched_bases), tier_priority_match,
                      claw_sigil_match, claw_sigil_level, owned_match,
                      melee_priority_match, melee_match, defense_priority_match,
                      item_tier, item_level_num)
                if key > best_key:
                    best_key = key
                    best_item = item
                    best_matched_bases = matched_bases
                    tied_items = [item]
                elif key == best_key and set(matched_bases) == set(best_matched_bases):
                    tied_items.append(item)

            if best_item:
                build[slot] = best_item
                covered_bases.update(best_matched_bases)
                if 'crafted' in (best_item.get('Realm') or '').strip().lower():
                    crafted_count += 1
                used_claw_items.append(best_item)
                alternates = [it for it in tied_items if it is not best_item]
                if alternates:
                    self.slot_alternates[slot] = alternates

        # Store the primary optimal build, then (if requested) generate additional
        # full build variants by swapping one tied alternate in at a time - lets
        # testers compare a few equally-good builds instead of just one.
        self.optimal_build = dict(build)
        self.build_variants = [dict(build)]
        if self.generate_multi_builds_var.get():
            for slot, alternates in self.slot_alternates.items():
                for alt in alternates:
                    if len(self.build_variants) >= MAX_BUILD_VARIANTS:
                        break
                    variant = dict(build)
                    variant[slot] = alt
                    self.build_variants.append(variant)
                if len(self.build_variants) >= MAX_BUILD_VARIANTS:
                    break

        variant_labels = [f"Build {i + 1}" for i in range(len(self.build_variants))]
        self.build_variant_combo['values'] = variant_labels
        self.build_variant_var.set(variant_labels[0])
        self.build_variant_combo.config(state='readonly' if len(variant_labels) > 1 else 'disabled')

        # Switch to the Results tab BEFORE inserting rows - a ttk.Treeview
        # populated while its Notebook tab isn't yet the active one can fail
        # to render until something else forces a redraw (e.g. picking a
        # different build variant), leaving the tab looking empty at first.
        self.results_display_mode.set('optimal')
        self.notebook.select(self.tab_results)

        # Slots worth flagging as "No suitable item found" (via
        # _build_dict_to_rows) if they stay unfilled. weapon/shield/claw are
        # always eligible (always-fill slots, or only in slots_to_fill at
        # all when their combo is checked) - an empty one there is always
        # worth a note. Armor/jewel slots are only eligible when at least
        # one wanted/priority spell couldn't be covered anywhere in the
        # whole build - an empty armor/jewel slot while everything wanted
        # is already covered elsewhere just means it wasn't needed, not
        # that the search failed to find anything for it.
        uncovered = sorted(set(wanted_bases) - covered_bases)
        if uncovered:
            self.attempted_slots = set(slots_to_fill)
        else:
            self.attempted_slots = {s for s in slots_to_fill
                                    if s in ('weapon', 'weapon_off', 'shield') or s.startswith('claw')}

        # Store and display every build variant stacked together, separated by
        # a thin black divider row, with Build 1 on top
        self.last_optimal_results = self._all_variants_rows()
        for row in self.last_optimal_results:
            self.search_results_tv.insert('', 'end', values=row)
        self._autosize_results_columns()

        # Show coverage (counted per distinct spell, not per tier requested)
        coverage = len(covered_bases)
        total = len(wanted_bases)
        status = (f"Optimal build covers {coverage}/{total} wanted spells | "
                   f"Spells covered: {', '.join(sorted(covered_bases))}")
        if self.slot_alternates:
            status += f" | {len(self.slot_alternates)} slot(s) have equally-good alternate items"
        if len(self.build_variants) > 1:
            status += f" | {len(self.build_variants)} build variants available"

        # Any wanted spell no item could satisfy at all, at any tier, under
        # the current armor/weapon/level/tier/realm constraints - every
        # other slot was filled, so this really means "nothing exists for
        # this exact configuration" rather than "not needed" (uncovered
        # was already computed above, to decide which empty slots get a
        # "No suitable item found" row). Shown in the status line instead
        # of an interrupting popup.
        if uncovered:
            status += f" | No items found for: {', '.join(uncovered)}"
        self.search_status.config(text=status)

        # Warn about any Priority Tier that couldn't actually be honored - the
        # spell itself may still be covered (via a fallback tier), just not at
        # the exact tier that was targeted for it specifically.
        unmatched_priority_tiers = []
        for spell, tier in self.priority_tiers_data:
            if spell not in covered_bases:
                continue  # already covered by the warning above
            actual_item = next((it for it in build.values()
                                if _spell_base((it.get('Spell') or '').lower()) == spell), None)
            if actual_item is None:
                continue
            actual_rank = _item_tier_rank((actual_item.get('Spell') or '').lower())
            if actual_rank != _TIER_RANK[tier]:
                actual_label = next((r for r, n in _TIER_RANK.items() if n == actual_rank), 'untiered')
                unmatched_priority_tiers.append(
                    f"{spell} ({tier}) cannot be used - tier {actual_label} used instead")

        if unmatched_priority_tiers:
            messagebox.showwarning("No Acceptable Gear At Priority Tier",
                "\n".join(unmatched_priority_tiers))

    def _show_all_matches(self):
        """Show ALL items matching criteria (not optimized for spell coverage)"""
        # Same reasoning as the matching line in _find_optimal_build - set
        # this before any validation checks below can return early, so the
        # Results tab's mode always reflects the button actually clicked.
        self.results_display_mode.set('all')

        if not self.master_data:
            messagebox.showwarning("No Data", "Please load a master database first")
            return

        # Get wanted spells
        wanted_spells = list(self.wanted_spells_data)
        if not wanted_spells:
            messagebox.showwarning("No Spells", "Please add at least one wanted spell")
            return

        # Clear previous results
        self.search_results_tv.delete(*self.search_results_tv.get_children())

        # Build variants only apply to "Find Optimal Build" results
        self.build_variants = [{}]
        self.build_variant_combo['values'] = ['Build 1']
        self.build_variant_var.set('Build 1')
        self.build_variant_combo.config(state='disabled')
        
        # Get level constraints
        min_level = None
        max_level = None
        specific_level = None
        
        # Check for specific level first (takes precedence)
        specific_level_str = self.specific_level_var.get().strip()
        if specific_level_str:
            try:
                specific_level = int(specific_level_str)
            except ValueError:
                messagebox.showwarning("Invalid Level", "Specific level must be a number")
                return
        else:
            # If no specific level, check min/max
            min_level_str = self.min_level_var.get().strip()
            if min_level_str:
                try:
                    min_level = int(min_level_str)
                except ValueError:
                    messagebox.showwarning("Invalid Level", "Minimum level must be a number")
                    return
            
            max_level_str = self.max_level_var.get().strip()
            if max_level_str:
                try:
                    max_level = int(max_level_str)
                except ValueError:
                    messagebox.showwarning("Invalid Level", "Maximum level must be a number")
                    return
            
            # Validate min <= max
            if min_level is not None and max_level is not None and min_level > max_level:
                messagebox.showwarning("Invalid Range", "Minimum level cannot be greater than maximum level")
                return

        # Defense filter (opt-in via checkbox) - bounds items to a range on
        # the item list's Defense column (much worse..much better).
        defense_min_rank = None
        defense_max_rank = None
        if self.use_defense_filter_var.get():
            defense_min_rank = DEFENSE_RANK[self.min_defense_var.get()]
            defense_max_rank = DEFENSE_RANK[self.max_defense_var.get()]
            if defense_min_rank > defense_max_rank:
                messagebox.showwarning("Invalid Range", "Minimum Defense cannot be greater than maximum Defense")
                return

        # Per-slot Defense filter (opt-in per armor slot) - stacks with the
        # global one above.
        slot_defense_min_rank = {}
        slot_defense_max_rank = {}
        for slot, controls in self.slot_defense_controls.items():
            if controls['use'].get():
                min_rank = DEFENSE_RANK[controls['min'].get()]
                max_rank = DEFENSE_RANK[controls['max'].get()]
                if min_rank > max_rank:
                    messagebox.showwarning("Invalid Range",
                        f"Minimum Defense cannot be greater than maximum Defense ({slot.title()})")
                    return
                slot_defense_min_rank[slot] = min_rank
                slot_defense_max_rank[slot] = max_rank

        # Get armor type constraints from checkboxes
        armor_constraints = {}
        for slot in ['head', 'cloak', 'body', 'hands', 'legs', 'feet']:
            # Get checked armor types for this slot
            checked_types = []
            for armor_type in ['cloth', 'leather', 'studded', 'plate']:
                if self.armor_checks[slot][armor_type].get():
                    checked_types.append(armor_type)
            armor_constraints[slot] = checked_types  # List of allowed types (empty = any)
        # Shield Constraints' armor type checkboxes - shields are technically
        # armor, so this reuses the exact same hard-filter mechanism as the
        # Armor Constraints tab's per-slot checkboxes, just sourced from
        # Weapon Constraints' own Shield Constraints checkboxes instead.
        armor_constraints['shield'] = [t for t in ('cloth', 'leather', 'studded', 'plate')
                                       if self.shield_armor_checks[t].get()]

        # Weapon Types/Combo's - each has its own Style and/or Damage Type
        # dropdown(s) (Claw has neither); the old global Weapon Style radios
        # and Damage Type checkboxes have been removed in favor of these.
        wants_two_handed = self.two_handed_var.get()
        two_handed_style = self.two_handed_style_var.get()
        two_handed_damage = self.two_handed_damage_var.get()
        wants_claw_1 = self.claw_1_var.get()
        wants_claw_2 = self.claw_2_var.get()
        wants_dual_wield_1h = self.dual_wield_1h_var.get()
        dual_wield_1h_main = self.dual_wield_1h_main_var.get()
        dual_wield_1h_off = self.dual_wield_1h_off_var.get()
        wants_1h_shield = self.combo_1h_shield_var.get()
        combo_1h_shield_style = self.combo_1h_shield_style_var.get()
        combo_1h_shield_damage = self.combo_1h_shield_damage_var.get()
        wants_2h_shield = self.combo_2h_shield_var.get()
        combo_2h_shield_damage = self.combo_2h_shield_damage_var.get()
        # Fired 1h and Shield - "fired" is its own weapon shape in the item
        # list with no slash/thrust/crush sub-variant, so no dropdown needed.
        wants_fired_1h_shield = self.combo_fired_1h_shield_var.get()

        # Most real weapons carry no Spell at all - when a combo wants a
        # weapon/shield slot filled regardless, this listing should still
        # surface those spell-less items rather than hiding them entirely.
        any_weapon_combo_active = (wants_two_handed or wants_dual_wield_1h or wants_1h_shield
                                    or wants_2h_shield or wants_fired_1h_shield)
        any_shield_combo_active = wants_1h_shield or wants_2h_shield or wants_fired_1h_shield
        # Claws also carry no Spell in practice - same reasoning applies.
        any_claw_combo_active = wants_claw_1 or wants_claw_2

        # Filter items by constraints and group by slot
        items_by_slot = {}
        all_slots = ['head', 'jewel', 'jewel', 'cloak', 'body', 'hands', 'legs', 'feet', 'weapon', 'shield']

        for item in self.master_data:
            item_slot = (item.get('Slot') or '').lower()
            item_spell = (item.get('Spell') or '').lower()
            item_type = (item.get('Type') or '').lower()

            # Claws are stored in the source data as Slot=weapon/Type=claw,
            # not a distinct Slot='claw' value - see _find_optimal_build.
            is_claw_item = item_slot == 'weapon' and 'claw' in item_type

            is_combo_mandated_slot = ((item_slot == 'weapon' and not is_claw_item and any_weapon_combo_active)
                                       or (item_slot == 'shield' and any_shield_combo_active)
                                       or (is_claw_item and any_claw_combo_active))

            # Skip items without spells - unless a weapon/shield combo wants
            # this slot filled regardless (most real weapons have no spell).
            if not item_spell and not is_combo_mandated_slot:
                continue

            # Check if item's spell matches any wanted spell (partial match)
            has_wanted_spell = False
            for wanted in wanted_spells:
                if wanted in item_spell:
                    has_wanted_spell = True
                    break

            item_realm = (item.get('Realm') or '').strip()

            # Crafted items never show up unless the Crafted checkbox is
            # explicitly checked - see the matching comment in
            # _find_optimal_build for the full reasoning.
            if 'crafted' in item_realm.lower() and not self.realm_filters['Crafted'].get():
                continue

            # Apply realm filter if any selected
            selected_realms = [realm for realm, var in self.realm_filters.items()
                              if var.get()]
            if selected_realms:
                realm_match = False
                for selected in selected_realms:
                    if selected.lower() in item_realm.lower():
                        realm_match = True
                        break
                if not realm_match:
                    continue

            if not has_wanted_spell and not is_combo_mandated_slot:
                continue

            # Apply level constraints
            if specific_level is not None or min_level is not None or max_level is not None:
                item_level_str = item.get('Level', '')
                if item_level_str:
                    try:
                        item_level = int(item_level_str)
                        
                        # Check specific level
                        if specific_level is not None:
                            if item_level != specific_level:
                                continue
                        # Check min/max range
                        else:
                            if min_level is not None and item_level < min_level:
                                continue
                            if max_level is not None and item_level > max_level:
                                continue
                    except ValueError:
                        # Skip items with invalid level data
                        continue
                else:
                    # Skip items without level data when filtering by level
                    continue
            
            # Apply Defense filter (opt-in) - item must have a recognized
            # Defense value within the selected min/max range.
            if defense_min_rank is not None:
                item_defense_rank = DEFENSE_RANK.get((item.get('Defense') or '').strip().lower())
                if item_defense_rank is None or not (defense_min_rank <= item_defense_rank <= defense_max_rank):
                    continue

            # Apply per-slot Defense filter (opt-in per armor slot) - stacks
            # with the global one above.
            if item_slot in slot_defense_min_rank:
                item_defense_rank = DEFENSE_RANK.get((item.get('Defense') or '').strip().lower())
                if (item_defense_rank is None
                        or not (slot_defense_min_rank[item_slot] <= item_defense_rank <= slot_defense_max_rank[item_slot])):
                    continue

            # Apply armor type constraints (checkboxes)
            if item_slot in armor_constraints:
                allowed_types = armor_constraints[item_slot]
                # If any checkboxes are checked, item must match one of them
                if allowed_types:
                    type_match = False
                    for allowed_type in allowed_types:
                        if allowed_type.lower() in item_type:
                            type_match = True
                            break
                    if not type_match:
                        continue
            
            # Apply weapon/shield/claw constraints. The old global Weapon
            # Style radios and Damage Type checkboxes are gone - Style and
            # Damage Type are now per-combo dropdowns (see each branch below).
            if item_slot == 'weapon' or item_slot == 'shield':
                if is_claw_item:
                    # Claws (Slot=weapon/Type=claw) have no Style or Damage
                    # Type dropdown - 1 Claw / 2 Claw is the only gate.
                    if not (wants_claw_1 or wants_claw_2):
                        continue
                else:
                    # If no combo is checked, accept everything
                    has_any_config = (wants_two_handed or wants_claw_1 or wants_claw_2
                                       or wants_dual_wield_1h or wants_1h_shield
                                       or wants_2h_shield or wants_fired_1h_shield)

                    if has_any_config:
                        slot_accepted = False

                        if item_slot == 'weapon':
                            is_offhand = 'offhand' in item_type
                            is_1h = '1h' in item_type
                            is_2h = '2h' in item_type
                            is_fired = 'fired' in item_type

                            # Two-handed weapons, gated by their own Style
                            # (Melee/Direct/Parry/Fired) and Damage Type dropdowns
                            if wants_two_handed and _two_handed_matches(item_type, item_spell, two_handed_style, two_handed_damage):
                                slot_accepted = True
                            # Dual-Wield 1h - this is a flat listing (no per-
                            # slot picking), so main-hand and off-hand
                            # candidates are just shown together: non-offhand
                            # matching Main, or offhand-tagged matching Off-Hand.
                            if (wants_dual_wield_1h and is_1h and not is_offhand
                                    and _weapon_damage_matches(item_type, dual_wield_1h_main)):
                                slot_accepted = True
                            if (wants_dual_wield_1h and is_1h and is_offhand
                                    and _weapon_damage_matches(item_type, dual_wield_1h_off)):
                                slot_accepted = True
                            # 1h/Shield - Style (Melee/Direct, no Parry) + Damage Type;
                            # Direct additionally requires some spell to be present
                            if (wants_1h_shield and is_1h and not is_offhand
                                    and _weapon_style_matches(item_type, combo_1h_shield_style)
                                    and (combo_1h_shield_style != 'Direct' or _direct_weapon_eligible(item_spell))
                                    and _weapon_damage_matches(item_type, combo_1h_shield_damage)):
                                slot_accepted = True
                            # 2h/Shield
                            if (wants_2h_shield and is_2h
                                    and _weapon_damage_matches(item_type, combo_2h_shield_damage)):
                                slot_accepted = True
                            # Fired 1h and Shield
                            if wants_fired_1h_shield and is_fired and is_1h and not is_offhand:
                                slot_accepted = True

                        elif item_slot == 'shield':
                            # Shields - only implied by a weapon+shield combo
                            # now (plain "Shield" was removed)
                            if wants_1h_shield or wants_2h_shield or wants_fired_1h_shield:
                                slot_accepted = True

                        if not slot_accepted:
                            continue

            # Add to slot group. Claws (Slot=weapon/Type=claw) get their own
            # 'claw' bucket instead of piling into 'weapon' with everything else.
            bucket_slot = 'claw' if is_claw_item else item_slot
            if bucket_slot not in items_by_slot:
                items_by_slot[bucket_slot] = []
            items_by_slot[bucket_slot].append(item)

        # Display ALL matching items (not just one per slot)
        # Group by slot for better organization
        all_matches = []
        
        for slot_key, items in items_by_slot.items():
            for item in items:
                all_matches.append((slot_key, item))
        
        # Sort by slot order for display
        slot_priority = {'head': 1, 'jewel': 2, 'cloak': 3, 'body': 4, 'hands': 5, 
                        'legs': 6, 'feet': 7, 'weapon': 8, 'shield': 9, 'claw': 10}
        all_matches.sort(key=lambda x: slot_priority.get(x[0], 99))
        
        # Track which spells are covered
        covered_spells = set()

        # Switch to the Results tab BEFORE inserting rows - a ttk.Treeview
        # populated while its Notebook tab isn't yet the active one can fail
        # to render until something else forces a redraw, leaving the tab
        # looking empty at first.
        self.results_display_mode.set('all')
        self.notebook.select(self.tab_results)

        # Store and display all matching items
        self.last_all_results = []
        for slot, item in all_matches:
            item_spell = (item.get('Spell') or '').lower()
            
            # Check if this item provides any wanted spells
            for wanted in wanted_spells:
                if wanted in item_spell:
                    covered_spells.add(wanted)
            
            row = (
                '',
                slot.title(),
                item.get('Item', ''),
                item.get('Type', ''),
                item.get('Spell', ''),
                item.get('Sigil', ''),
                item.get('Level', ''),
                item.get('Mob', ''),
                item.get('Area', ''),
                '████████',
                ''
            )
            self.last_all_results.append(row)
            self.search_results_tv.insert('', 'end', values=row)
        self._autosize_results_columns()

        # Show coverage
        coverage = len(covered_spells)
        total = len(wanted_spells)
        self.search_status.config(
            text=f"All matches show {coverage}/{total} wanted spells | Spells covered: {', '.join(sorted(covered_spells))}")



def main():
    app = App()
    app.mainloop()


if __name__ == '__main__':
    main()
